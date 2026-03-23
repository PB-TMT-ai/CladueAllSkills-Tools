"""
Run All Tasks: Populate price difference columns in a single load-save cycle.

Usage:
  python tools/run_all.py                                               # Dec-25 defaults
  python tools/run_all.py "Price difference calculations - Jan26.xlsx" "Jan-26"

Auto-detects column layout from sheet names (supports Dec-25 and Jan-26+ formats).
Close Excel first. ~12 min.
"""

import os
import sys
import re
import time
import calendar
import openpyxl
from datetime import date, datetime

# ============================================================
# DEFAULTS -- overridden by command-line args
# ============================================================
DEFAULT_FILE = os.path.join(
    os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
    "Copy of Price difference calculations - Dec 25.xlsx")
DEFAULT_SHEET = "Dec-25"

SHEET_PINCODE = "Pincode"
SHEET_INSTRUCTIONS = "Instructions"
DATA_START_ROW = 4

# Lookup sheet columns (0-based for iter_rows) -- same positions in both
# "union book" and "Order" sheets
UB_C = 2     # Order ID
UB_S = 18    # Grade
UB_T = 19    # Diameter mm
UB_U = 20    # Form
UB_AD = 29   # Proposed Price Rs per MT
UB_BG = 58   # SFDC Comment

# Month abbreviation map for sheet-name parsing
MONTH_MAP = {v.upper()[:3]: k for k, v in enumerate(calendar.month_abbr) if v}


# ============================================================
# Layout detection
# ============================================================

def detect_layout(sheetnames):
    """Detect column layout from available sheet names."""
    if "union book" in sheetnames:
        return "dec25"
    elif "Order" in sheetnames:
        return "jan26"
    else:
        raise ValueError(f"Cannot detect layout. Sheets: {sheetnames}")


def get_layout_config(layout):
    """Return column config dict for the detected layout."""
    if layout == "dec25":
        return {
            "lookup_sheet": "union book",
            "lookup_data_start": 2,       # headers row 1, data row 2+
            # Main sheet input columns (1-based)
            "col_sentinel": 1,            # A  (col to detect end of data)
            "col_order_id": 3,            # C
            "col_grade": 8,               # H
            "col_diameter": 9,            # I
            "col_form": 10,               # J
            "col_pincode": 11,            # K
            "col_order_year": 12,         # L
            "col_order_month": 13,        # M
            "col_order_day": 14,          # N
            "col_invoice_date": 1,        # A
            "has_invoice_date": True,
            "has_order_ymd": True,
            # Output columns
            "col_out_inv_year": 16,       # P
            "col_out_inv_month": 17,      # Q
            "col_out_inv_day": 18,        # R
            "col_out_comment": 19,        # S
            "col_out_cluster": 20,        # T
            "col_out_proposed": 22,       # V
            "col_out_applicable": 23,     # W
        }
    elif layout == "jan26":
        return {
            "lookup_sheet": "Order",
            "lookup_data_start": 3,       # headers row 2, data row 3+
            # Main sheet input columns (1-based)
            "col_sentinel": 1,            # A  (Order ID)
            "col_order_id": 1,            # A
            "col_grade": 6,               # F
            "col_diameter": 7,            # G
            "col_form": 8,               # H
            "col_pincode": 9,             # I
            "col_order_day": 10,          # J  (day only, month from sheet name)
            "has_invoice_date": False,
            "has_order_ymd": False,
            # Output columns
            "col_out_comment": 12,        # L
            "col_out_cluster": 13,        # M
            "col_out_proposed": 14,       # N
            "col_out_applicable": 15,     # O
        }
    else:
        raise ValueError(f"Unknown layout: {layout}")


def parse_sheet_month_year(sheet_name):
    """Parse 'Jan-26' -> (1, 2026). Returns (month, year) or (None, None)."""
    parts = sheet_name.split("-")
    if len(parts) != 2:
        return None, None
    month = MONTH_MAP.get(parts[0].strip().upper()[:3])
    if month is None:
        return None, None
    try:
        year = 2000 + int(parts[1].strip())
    except ValueError:
        return None, None
    return month, year


# ============================================================
# Helper functions
# ============================================================

def parse_date_string(date_str):
    """Parse 'YYYY-MM-DD' string into (year, month, day) integers."""
    if not isinstance(date_str, str):
        return None, None, None
    parts = date_str.strip().split("-")
    if len(parts) != 3:
        return None, None, None
    try:
        return int(parts[0]), int(parts[1]), int(parts[2])
    except (ValueError, IndexError):
        return None, None, None


def normalize_key(order_id, grade, diameter, form):
    """Normalize all four fields for case/type-insensitive matching."""
    oid = str(order_id).strip() if order_id is not None else ""
    g = str(grade).strip().upper() if grade is not None else ""
    d = str(diameter).strip() if diameter is not None else ""
    f = str(form).strip().upper() if form is not None else ""
    return (oid, g, d, f)


def find_period_index(order_date, periods):
    """Find pricing period index for a date. Returns index or None."""
    if not periods:
        return None
    if order_date < periods[0][0]:
        return 0
    if order_date > periods[-1][1]:
        return len(periods) - 1
    for i, (d_from, d_to) in enumerate(periods):
        if d_from <= order_date <= d_to:
            return i
    return None


# ============================================================
# Pass 1: Build ALL lookups from read-only workbook
# ============================================================

def build_all_lookups(filepath, sheet_pricing, layout, cfg, sheet_month, sheet_year):
    """Single read-only pass: build all lookup dicts + pricing model."""
    print("\n[Pass 1] Building all lookups (read-only, data_only=True)...")
    t0 = time.time()
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)

    lookup_sheet = cfg["lookup_sheet"]
    lookup_start = cfg["lookup_data_start"]

    # --- Lookup sheet: Order ID -> SFDC Comment + Proposed Price ---
    ws_ub = wb[lookup_sheet]
    order_to_comment = {}
    proposed_price_lookup = {}
    ub_count = 0

    for row in ws_ub.iter_rows(min_row=lookup_start):
        order_id = row[UB_C].value
        if order_id is None:
            continue
        ub_count += 1

        # SFDC Comment
        comment = row[UB_BG].value
        if order_id not in order_to_comment:
            order_to_comment[order_id] = comment
        else:
            if isinstance(comment, str) and not isinstance(order_to_comment[order_id], str):
                order_to_comment[order_id] = comment

        # Proposed Price
        grade = row[UB_S].value
        diameter = row[UB_T].value
        form = row[UB_U].value
        price = row[UB_AD].value
        key = normalize_key(order_id, grade, diameter, form)
        if key not in proposed_price_lookup:
            proposed_price_lookup[key] = price

    print(f"  {lookup_sheet}: {ub_count} rows -> {len(order_to_comment)} Order IDs, "
          f"{len(proposed_price_lookup)} price combos")

    # --- Pincode: Pincode -> State ---
    ws_pin = wb[SHEET_PINCODE]
    pincode_to_state = {}
    pin_count = 0
    for row in ws_pin.iter_rows(min_row=2):
        pincode = row[0].value
        state = row[2].value
        if pincode is None:
            continue
        pin_count += 1
        if pincode not in pincode_to_state:
            pincode_to_state[pincode] = state

    print(f"  Pincode: {pin_count} rows -> {len(pincode_to_state)} unique pincodes")

    # --- Pricing model (read with data_only=True for computed values) ---
    print("  Reading pricing model...")
    model = read_pricing_model(wb, sheet_pricing, sheet_month, sheet_year)
    print(f"  {len(model['date_periods'])} periods:")
    for i, (df, dt) in enumerate(model["date_periods"]):
        base550 = model["delhi_fe550_bases"][i]
        base550d = model["delhi_fe550d_bases"][i]
        print(f"    {i+1}. {df} to {dt}  FE550={base550}  FE550D={base550d}")
    print(f"  {len(model['ladders'])} clusters, "
          f"dia extras={model['diameter_extras']}, "
          f"form extras={model['form_extras']}")

    wb.close()
    print(f"  Pass 1 completed in {time.time() - t0:.1f}s")

    return order_to_comment, pincode_to_state, proposed_price_lookup, model


def read_pricing_model(wb, sheet_pricing, sheet_month, sheet_year):
    """Read pricing model from workbook (data_only=True for computed values)."""
    ws = wb[sheet_pricing]

    # Read enough rows and columns
    rows = []
    for row in ws.iter_rows(min_col=1, max_col=22, values_only=True):
        rows.append(list(row))
        if len(rows) > 60:
            break

    # Detect period columns from row 3 (index 2): FE 550 at even offsets
    row3 = rows[2] if len(rows) > 2 else []
    fe550_cols = []   # indices into row for FE 550
    fe550d_cols = []  # indices into row for FE 550D
    for i in range(3, len(row3), 2):
        if row3[i] is not None:
            fe550_cols.append(i)
            if i + 1 < len(row3):
                fe550d_cols.append(i + 1)

    # Date periods -- may be date objects (Dec-25) or day integers (Jan-26+)
    date_periods = []
    for fc in fe550_cols:
        d_from = rows[2][fc]
        d_to = rows[3][fc]
        # Convert datetime -> date
        if isinstance(d_from, datetime):
            d_from = d_from.date()
        if isinstance(d_to, datetime):
            d_to = d_to.date()
        # Convert day-number integers to full dates using sheet month/year
        if isinstance(d_from, (int, float)) and sheet_month and sheet_year:
            d_from = date(sheet_year, sheet_month, int(d_from))
        if isinstance(d_to, (int, float)) and sheet_month and sheet_year:
            d_to = date(sheet_year, sheet_month, int(d_to))
        date_periods.append((d_from, d_to))

    # Delhi base prices - row 7 (index 6)
    delhi_fe550_bases = []
    delhi_fe550d_bases = []
    delhi_row = rows[6] if len(rows) > 6 else []
    for fc in fe550_cols:
        val = delhi_row[fc] if fc < len(delhi_row) else None
        delhi_fe550_bases.append(val if isinstance(val, (int, float)) else 0)
    for dc in fe550d_cols:
        val = delhi_row[dc] if dc < len(delhi_row) else None
        delhi_fe550d_bases.append(val if isinstance(val, (int, float)) else 0)

    # Ladders - rows 7+ until col A is empty
    ladders = {}
    for r in range(6, len(rows)):
        location = rows[r][0]
        if location is None:
            break
        ladder = rows[r][2]
        ladders[str(location).strip().upper()] = int(ladder) if ladder else 0

    # Instructions sheet - diameter and form extras
    diameter_extras = {}
    form_extras = {}
    try:
        ws_inst = wb[SHEET_INSTRUCTIONS]
        for row in ws_inst.iter_rows(min_col=1, max_col=2, values_only=True):
            if row[0] is None or row[1] is None:
                continue
            key = str(row[0]).strip()
            val = row[1]
            if not isinstance(val, (int, float)) or int(val) <= 0:
                continue
            val = int(val)
            key_upper = key.upper().replace(" ", "-")
            digits = re.sub(r"[^\d]", "", key)
            if key_upper in ("U-BEND", "FISH-BEND"):
                form_extras[key_upper] = val
            elif digits:
                diameter_extras[digits] = val
    except KeyError:
        print(f"  WARNING: '{SHEET_INSTRUCTIONS}' sheet not found, using no extras")

    return {
        "date_periods": date_periods,
        "delhi_fe550_bases": delhi_fe550_bases,
        "delhi_fe550d_bases": delhi_fe550d_bases,
        "ladders": ladders,
        "diameter_extras": diameter_extras,
        "form_extras": form_extras,
    }


# ============================================================
# Pass 2: Single read-write pass -- all tasks in one loop
# ============================================================

def run_all_tasks(filepath, sheet_main, cfg, model,
                  order_to_comment, pincode_to_state, proposed_price_lookup,
                  sheet_month, sheet_year):
    """Single read-write load: write all output columns in one row loop, save once."""
    print("\n[Pass 2] Opening workbook in read-write mode...")
    t0 = time.time()
    wb = openpyxl.load_workbook(filepath)
    print(f"  Workbook loaded in {time.time() - t0:.1f}s")

    ws = wb[sheet_main]

    has_invoice = cfg.get("has_invoice_date", False)
    has_ymd = cfg.get("has_order_ymd", False)
    col_sentinel = cfg["col_sentinel"]
    col_oid = cfg["col_order_id"]
    col_grade = cfg["col_grade"]
    col_diam = cfg["col_diameter"]
    col_form = cfg["col_form"]
    col_pin = cfg["col_pincode"]
    col_out_comment = cfg["col_out_comment"]
    col_out_cluster = cfg["col_out_cluster"]
    col_out_proposed = cfg["col_out_proposed"]
    col_out_applicable = cfg["col_out_applicable"]

    stats = {
        "rows": 0,
        "date_ok": 0, "date_fail": 0,
        "order_matched": 0, "order_missing": 0,
        "pin_matched": 0, "pin_missing": 0,
        "price_matched": 0, "price_missing": 0,
        "priced": 0, "placeholder": 0,
        "no_period": 0, "no_cluster": 0, "unknown_cluster": 0,
    }
    unknown_clusters = set()

    print("\n  Processing rows...")
    row_num = DATA_START_ROW
    while True:
        sentinel = ws.cell(row=row_num, column=col_sentinel).value
        if sentinel is None:
            break
        stats["rows"] += 1

        order_id = ws.cell(row=row_num, column=col_oid).value

        # ==== TASK 1a: Invoice date parts (Dec-25 layout only) ====
        if has_invoice:
            date_val = ws.cell(row=row_num, column=cfg["col_invoice_date"]).value
            year, month, day = parse_date_string(date_val)
            if year is not None:
                ws.cell(row=row_num, column=cfg["col_out_inv_year"], value=year)
                ws.cell(row=row_num, column=cfg["col_out_inv_month"], value=month)
                ws.cell(row=row_num, column=cfg["col_out_inv_day"], value=day)
                stats["date_ok"] += 1
            else:
                stats["date_fail"] += 1

        # ==== TASK 1b: SF comment ====
        if order_id is not None and order_id in order_to_comment:
            comment = order_to_comment[order_id]
            if isinstance(comment, int) and comment == 0:
                ws.cell(row=row_num, column=col_out_comment, value="")
            else:
                ws.cell(row=row_num, column=col_out_comment, value=comment)
            stats["order_matched"] += 1
        else:
            stats["order_missing"] += 1

        # ==== TASK 1c: Pricing cluster (state from pincode) ====
        pincode = ws.cell(row=row_num, column=col_pin).value
        cluster_val = None
        if pincode is not None and pincode in pincode_to_state:
            cluster_val = pincode_to_state[pincode]
            ws.cell(row=row_num, column=col_out_cluster, value=cluster_val)
            stats["pin_matched"] += 1
        else:
            stats["pin_missing"] += 1

        # ==== TASK 2: Proposed price ====
        grade = ws.cell(row=row_num, column=col_grade).value
        diameter = ws.cell(row=row_num, column=col_diam).value
        form = ws.cell(row=row_num, column=col_form).value
        pp_key = normalize_key(order_id, grade, diameter, form)
        pp = proposed_price_lookup.get(pp_key)
        if pp is not None:
            ws.cell(row=row_num, column=col_out_proposed, value=pp)
            stats["price_matched"] += 1
        else:
            stats["price_missing"] += 1

        # ==== TASK 3: Applicable price ====
        grade_str = str(grade).strip().upper() if grade else ""
        if grade_str == "0" or grade_str == "":
            stats["placeholder"] += 1
        else:
            # Build order date for period lookup
            order_date = None
            if has_ymd:
                # Dec-25: separate year/month/day columns
                o_year = ws.cell(row=row_num, column=cfg["col_order_year"]).value
                o_month = ws.cell(row=row_num, column=cfg["col_order_month"]).value
                o_day = ws.cell(row=row_num, column=cfg["col_order_day"]).value
                try:
                    mo = int(o_month)
                    if mo == 0:
                        raise ValueError
                    order_date = date(int(o_year), mo, int(o_day))
                except (ValueError, TypeError):
                    pass
            else:
                # Jan-26+: only day column, month/year from sheet name
                o_day = ws.cell(row=row_num, column=cfg["col_order_day"]).value
                if o_day is not None and sheet_month and sheet_year:
                    try:
                        order_date = date(sheet_year, sheet_month, int(o_day))
                    except (ValueError, TypeError):
                        pass

            if order_date is None:
                stats["no_period"] += 1
                row_num += 1
                if stats["rows"] % 500 == 0:
                    print(f"    {stats['rows']} rows...")
                continue

            pi = find_period_index(order_date, model["date_periods"])
            if pi is None:
                stats["no_period"] += 1
                row_num += 1
                if stats["rows"] % 500 == 0:
                    print(f"    {stats['rows']} rows...")
                continue

            # Cluster lookup
            cluster_key = str(cluster_val).strip().upper() if cluster_val else ""
            if not cluster_key:
                stats["no_cluster"] += 1
                row_num += 1
                if stats["rows"] % 500 == 0:
                    print(f"    {stats['rows']} rows...")
                continue
            if cluster_key not in model["ladders"]:
                stats["unknown_cluster"] += 1
                if cluster_key not in unknown_clusters:
                    unknown_clusters.add(cluster_key)
                    print(f"  WARNING: unknown cluster '{cluster_key}'")
                row_num += 1
                if stats["rows"] % 500 == 0:
                    print(f"    {stats['rows']} rows...")
                continue

            # Compute price: Delhi base (FE 550 or FE 550D) + ladder + extras
            if "550D" in grade_str:
                delhi_base = model["delhi_fe550d_bases"][pi]
            else:
                delhi_base = model["delhi_fe550_bases"][pi]

            w_price = delhi_base + model["ladders"][cluster_key]
            diam_str = str(diameter).strip() if diameter else ""
            w_price += model["diameter_extras"].get(diam_str, 0)
            form_str = str(form).strip().upper().replace(" ", "-") if form else ""
            w_price += model["form_extras"].get(form_str, 0)

            ws.cell(row=row_num, column=col_out_applicable, value=w_price)
            stats["priced"] += 1

        row_num += 1
        if stats["rows"] % 500 == 0:
            print(f"    {stats['rows']} rows...")

    t1 = time.time()
    print(f"\n  Processed {stats['rows']} rows in {t1 - t0:.1f}s")
    print("  Saving workbook...")
    wb.save(filepath)
    wb.close()
    print(f"  Saved in {time.time() - t1:.1f}s")

    return stats, unknown_clusters


# ============================================================
# Verification
# ============================================================

def verify(filepath, sheet_main, cfg):
    """Re-read saved file and print sample values + fill rates."""
    print("\n[Verification] Re-reading written values...")
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = wb[sheet_main]

    has_invoice = cfg.get("has_invoice_date", False)
    col_comment = cfg["col_out_comment"]
    col_cluster = cfg["col_out_cluster"]
    col_proposed = cfg["col_out_proposed"]
    col_applicable = cfg["col_out_applicable"]

    comment_filled = cluster_filled = proposed_filled = applicable_filled = total = 0
    inv_year_filled = 0

    for row in ws.iter_rows(min_row=DATA_START_ROW):
        if row[0].value is None:
            break
        total += 1
        if row[col_comment - 1].value is not None:
            comment_filled += 1
        if row[col_cluster - 1].value is not None:
            cluster_filled += 1
        if row[col_proposed - 1].value is not None:
            proposed_filled += 1
        if row[col_applicable - 1].value is not None:
            applicable_filled += 1
        if has_invoice and row[cfg["col_out_inv_year"] - 1].value is not None:
            inv_year_filled += 1

        if total <= 5:
            vals = (f"Comment={row[col_comment-1].value}, "
                    f"Cluster={row[col_cluster-1].value}, "
                    f"Proposed={row[col_proposed-1].value}, "
                    f"Applicable={row[col_applicable-1].value}")
            print(f"  Row {total + 3}: {vals}")

    pct = lambda f: f"{100 * f / total:.1f}%" if total else "N/A"
    print(f"\n  Fill rates ({total} rows):")
    if has_invoice:
        print(f"    Invoice year:      {inv_year_filled}/{total} ({pct(inv_year_filled)})")
    print(f"    SF comment:        {comment_filled}/{total} ({pct(comment_filled)})")
    print(f"    Pricing cluster:   {cluster_filled}/{total} ({pct(cluster_filled)})")
    print(f"    Proposed Price:    {proposed_filled}/{total} ({pct(proposed_filled)})")
    print(f"    Applicable Price:  {applicable_filled}/{total} ({pct(applicable_filled)})")
    wb.close()


# ============================================================
# Main
# ============================================================

if __name__ == "__main__":
    filepath = sys.argv[1] if len(sys.argv) > 1 else DEFAULT_FILE
    sheet = sys.argv[2] if len(sys.argv) > 2 else DEFAULT_SHEET

    if not os.path.isabs(filepath):
        filepath = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), filepath)

    sheet_pricing = f"{sheet} pricing"
    sheet_month, sheet_year = parse_sheet_month_year(sheet)

    print("=" * 60)
    print("Price Difference: All Tasks")
    print(f"  File:    {os.path.basename(filepath)}")
    print(f"  Sheet:   {sheet}")
    print(f"  Pricing: {sheet_pricing}")
    if sheet_month and sheet_year:
        print(f"  Month:   {calendar.month_abbr[sheet_month]} {sheet_year}")
    print("=" * 60)

    if not os.path.exists(filepath):
        print(f"ERROR: File not found: {filepath}")
        sys.exit(1)

    # Detect layout from sheet names
    wb_tmp = openpyxl.load_workbook(filepath, read_only=True)
    layout = detect_layout(wb_tmp.sheetnames)
    wb_tmp.close()
    cfg = get_layout_config(layout)
    print(f"  Layout:  {layout} (lookup sheet: '{cfg['lookup_sheet']}')")

    t_start = time.time()

    # Pass 1: read-only -- build all lookups + pricing model
    order_to_comment, pincode_to_state, proposed_price_lookup, model = \
        build_all_lookups(filepath, sheet_pricing, layout, cfg, sheet_month, sheet_year)

    # Pass 2: read-write -- all tasks in one row loop, one save
    stats, unknown_clusters = run_all_tasks(
        filepath, sheet, cfg, model,
        order_to_comment, pincode_to_state, proposed_price_lookup,
        sheet_month, sheet_year)

    total_time = time.time() - t_start

    # Summary
    print(f"\n{'=' * 60}")
    print(f"SUMMARY ({total_time:.0f}s total)")
    print(f"{'=' * 60}")
    print(f"  Rows:              {stats['rows']}")
    if cfg.get("has_invoice_date"):
        print(f"  --- Task 1 (date parts) ---")
        print(f"  Date parsed:       {stats['date_ok']}/{stats['rows']}")
    print(f"  --- Lookups ---")
    print(f"  Order ID matched:  {stats['order_matched']}/{stats['rows']}")
    print(f"  Pincode matched:   {stats['pin_matched']}/{stats['rows']}")
    pct2 = 100 * stats["price_matched"] / stats["rows"] if stats["rows"] else 0
    print(f"  --- Proposed Price ---")
    print(f"  Matched:           {stats['price_matched']}/{stats['rows']} ({pct2:.1f}%)")
    pct3 = 100 * stats["priced"] / stats["rows"] if stats["rows"] else 0
    print(f"  --- Applicable Price ---")
    print(f"  Priced:            {stats['priced']}/{stats['rows']} ({pct3:.1f}%)")
    print(f"  Placeholder:       {stats['placeholder']}")
    print(f"  No period:         {stats['no_period']}")
    print(f"  No cluster:        {stats['no_cluster']}")
    print(f"  Unknown cluster:   {stats['unknown_cluster']} {unknown_clusters or ''}")
    print("=" * 60)

    verify(filepath, sheet, cfg)
    print(f"\nAll tasks complete in {total_time:.0f}s.")
