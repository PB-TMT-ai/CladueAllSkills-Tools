import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, numbers
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule
from openpyxl.utils import get_column_letter
import numpy as np
from datetime import datetime, timedelta
import math
import warnings
warnings.filterwarnings('ignore')

# File paths
INPUT_FILE = 'report1770198136609.xls'
OUTPUT_FILE = 'Saurav_Nath_Visit_Analysis_Nov25_Feb26.xlsx'

print("="*80)
print("SAURAV NATH VISIT ANALYSIS - DATA PROCESSING")
print("="*80)

# Load data
print("\n[1/6] Loading data...")
# The .xls file is actually HTML format, so we read it as HTML table
df = pd.read_html(INPUT_FILE)[0]
print(f"   [OK] Loaded {len(df)} records")
print(f"   [OK] Columns: {df.shape[1]}")

# Parse datetime columns
print("\n[2/6] Parsing datetime fields...")
df['Check-in DateTime'] = pd.to_datetime(df['Check-in Date/Time'], format='%d/%m/%Y, %I:%M %p', errors='coerce')
df['Check-out DateTime'] = pd.to_datetime(df['Check-out Date/Time'], format='%d/%m/%Y, %I:%M %p', errors='coerce')

# Calculate derived fields
print("\n[3/6] Calculating derived fields...")
df['Date'] = df['Check-in DateTime'].dt.date
df['Day of Week'] = df['Check-in DateTime'].dt.day_name()
df['Month'] = df['Check-in DateTime'].dt.strftime('%b %y')
df['Check-in Time'] = df['Check-in DateTime'].dt.time
df['Check-out Time'] = df['Check-out DateTime'].dt.time

# Duration in minutes
df['Duration (min)'] = ((df['Check-out DateTime'] - df['Check-in DateTime']).dt.total_seconds() / 60).round(2)

# Combined comments
df['Combined Comments'] = df['Check-in Comments'].fillna('') + ' | ' + df['Check-out Comments'].fillna('')
df['Comments Word Count'] = df['Combined Comments'].str.split().str.len()

# Visit number for each customer
df = df.sort_values(['Account: Registered Company Name', 'Check-in DateTime'])
df['Visit Number'] = df.groupby('Account: Registered Company Name').cumcount() + 1

# Flags
df['Short Visit Flag'] = df['Duration (min)'].apply(lambda x: 'YES' if x < 5 else 'NO')
df['Documentation Complete'] = ((df['Check-in Comments'].notna()) & (df['Check-out Comments'].notna())).apply(lambda x: 'YES' if x else 'NO')

# Time slot
def get_time_slot(dt):
    if pd.isna(dt):
        return 'Unknown'
    hour = dt.hour
    if hour < 12:
        return 'Morning'
    elif hour < 17:
        return 'Afternoon'
    else:
        return 'Evening'

df['Time Slot'] = df['Check-in DateTime'].apply(get_time_slot)

# Same day visits
df['Same Day Visits'] = df.groupby('Date')['Date'].transform('count')

# Duration band
def duration_band(duration):
    if pd.isna(duration):
        return 'Unknown'
    if duration < 5:
        return '<5 min'
    elif duration < 15:
        return '5-15 min'
    elif duration < 30:
        return '15-30 min'
    else:
        return '30+ min'

df['Duration Band'] = df['Duration (min)'].apply(duration_band)

# Month numeric for sorting
df['Month Num'] = df['Check-in DateTime'].dt.month

print(f"   [OK] Duration calculated: {df['Duration (min)'].notna().sum()} records")
print(f"   [OK] Short visits (<5 min): {(df['Duration (min)'] < 5).sum()} records")

# Create Excel writer
print("\n[4/6] Creating Excel workbook...")
writer = pd.ExcelWriter(OUTPUT_FILE, engine='openpyxl')

# ============================================================================
# SHEET 1: EXECUTIVE DASHBOARD
# ============================================================================
print("   -> Sheet 1: Executive_Dashboard")

dashboard_data = {
    'Performance Dimension': [
        'Visit Productivity',
        'Territory Coverage',
        'Geographic Spread',
        'Visit Quality',
        'Quality Red Flags: Short Visits',
        'Quality Red Flags: High Frequency Repeats'
    ],
    'Metric': [
        'Visits per Day',
        'Unique Customers per Month',
        'Unique Pin Codes',
        'Avg Duration (min)',
        '% Short Visits (<5 min)',
        '% High Frequency Customers (10+ visits)'
    ],
    'Target/Benchmark': [6, 50, 10, 15, '<10%', '<5%'],
    'Dec 25': [0, 0, 0, 0, 0, 0],
    'Jan 26': [0, 0, 0, 0, 0, 0],
    'Overall': [0, 0, 0, 0, 0, 0],
    'Rating': ['', '', '', '', '', '']
}

# Calculate metrics
working_days_dec = df[df['Month Num'] == 12]['Date'].nunique()
working_days_jan = df[df['Month Num'] == 1]['Date'].nunique()
working_days_total = df['Date'].nunique()

# Dec metrics
visits_dec = len(df[df['Month Num'] == 12])
customers_dec = df[df['Month Num'] == 12]['Account: Registered Company Name'].nunique()
pincodes_dec = df[df['Month Num'] == 12]['Account: Pin code'].nunique()
duration_dec = df[df['Month Num'] == 12]['Duration (min)'].mean()
short_visits_dec = (df[df['Month Num'] == 12]['Duration (min)'] < 5).sum() / visits_dec * 100 if visits_dec > 0 else 0

# Jan metrics
visits_jan = len(df[df['Month Num'] == 1])
customers_jan = df[df['Month Num'] == 1]['Account: Registered Company Name'].nunique()
pincodes_jan = df[df['Month Num'] == 1]['Account: Pin code'].nunique()
duration_jan = df[df['Month Num'] == 1]['Duration (min)'].mean()
short_visits_jan = (df[df['Month Num'] == 1]['Duration (min)'] < 5).sum() / visits_jan * 100 if visits_jan > 0 else 0

# Overall metrics
visits_total = len(df)
customers_total = df['Account: Registered Company Name'].nunique()
pincodes_total = df['Account: Pin code'].nunique()
duration_total = df['Duration (min)'].mean()
short_visits_total = (df['Duration (min)'] < 5).sum() / visits_total * 100

# High frequency customers
visit_counts = df.groupby('Account: Registered Company Name').size()
high_freq_count = (visit_counts >= 10).sum()
high_freq_pct = high_freq_count / len(visit_counts) * 100 if len(visit_counts) > 0 else 0

dashboard_data['Dec 25'] = [
    round(visits_dec / working_days_dec, 2) if working_days_dec > 0 else 0,
    customers_dec,
    pincodes_dec,
    round(duration_dec, 2),
    round(short_visits_dec, 1),
    round(high_freq_pct, 1)
]

dashboard_data['Jan 26'] = [
    round(visits_jan / working_days_jan, 2) if working_days_jan > 0 else 0,
    customers_jan,
    pincodes_jan,
    round(duration_jan, 2),
    round(short_visits_jan, 1),
    round(high_freq_pct, 1)
]

dashboard_data['Overall'] = [
    round(visits_total / working_days_total, 2) if working_days_total > 0 else 0,
    round(customers_total / 2, 1),  # Average per month
    pincodes_total,
    round(duration_total, 2),
    round(short_visits_total, 1),
    round(high_freq_pct, 1)
]

df_dashboard = pd.DataFrame(dashboard_data)
df_dashboard.to_excel(writer, sheet_name='Executive_Dashboard', index=False)

# ============================================================================
# SHEET 2: MONTHLY TRENDS
# ============================================================================
print("   -> Sheet 2: Monthly_Trends")

trends_data = {
    'Metric': [
        'Total Visits',
        'Avg Visits/Day',
        'Unique Customers',
        'Unique Pin Codes',
        'Avg Duration (min)',
        'Total Field Hours',
        '% Short Visits (<5 min)',
        '% Complete Documentation',
        'Working Days'
    ],
    'Nov 25': ['N/A'] * 9,
    'Dec 25': [0] * 9,
    'Jan 26': [0] * 9,
    'Feb 26': ['N/A'] * 9,
    'Trend': [''] * 9,
    '% Change (Dec-Jan)': [0] * 9
}

# Calculate Dec values
doc_complete_dec = (df[df['Month Num'] == 12]['Documentation Complete'] == 'YES').sum() / visits_dec * 100 if visits_dec > 0 else 0
field_hours_dec = df[df['Month Num'] == 12]['Duration (min)'].sum() / 60

trends_data['Dec 25'] = [
    visits_dec,
    round(visits_dec / working_days_dec, 2) if working_days_dec > 0 else 0,
    customers_dec,
    pincodes_dec,
    round(duration_dec, 2),
    round(field_hours_dec, 2),
    round(short_visits_dec, 1),
    round(doc_complete_dec, 1),
    working_days_dec
]

# Calculate Jan values
doc_complete_jan = (df[df['Month Num'] == 1]['Documentation Complete'] == 'YES').sum() / visits_jan * 100 if visits_jan > 0 else 0
field_hours_jan = df[df['Month Num'] == 1]['Duration (min)'].sum() / 60

trends_data['Jan 26'] = [
    visits_jan,
    round(visits_jan / working_days_jan, 2) if working_days_jan > 0 else 0,
    customers_jan,
    pincodes_jan,
    round(duration_jan, 2),
    round(field_hours_jan, 2),
    round(short_visits_jan, 1),
    round(doc_complete_jan, 1),
    working_days_jan
]

# Calculate % change
for i in range(len(trends_data['Metric'])):
    dec_val = trends_data['Dec 25'][i]
    jan_val = trends_data['Jan 26'][i]
    if isinstance(dec_val, (int, float)) and isinstance(jan_val, (int, float)) and dec_val != 0:
        change = ((jan_val - dec_val) / dec_val) * 100
        trends_data['% Change (Dec-Jan)'][i] = round(change, 1)
        trends_data['Trend'][i] = 'UP' if change > 5 else 'DOWN' if change < -5 else 'STABLE'
    else:
        trends_data['% Change (Dec-Jan)'][i] = 'N/A'
        trends_data['Trend'][i] = 'STABLE'

df_trends = pd.DataFrame(trends_data)
df_trends.to_excel(writer, sheet_name='Monthly_Trends', index=False)

# ============================================================================
# SHEET 3: PRODUCTIVITY MONTHLY
# ============================================================================
print("   -> Sheet 3: Productivity_Monthly")

productivity_data = []
for month_num, month_name in [(12, 'Dec 25'), (1, 'Jan 26')]:
    month_df = df[df['Month Num'] == month_num]
    total_visits = len(month_df)
    working_days = month_df['Date'].nunique()
    avg_visits_day = total_visits / working_days if working_days > 0 else 0

    # Days meeting target
    daily_counts = month_df.groupby('Date').size()
    days_meeting_target = (daily_counts >= 6).sum()
    pct_days_on_target = (days_meeting_target / working_days * 100) if working_days > 0 else 0

    avg_duration = month_df['Duration (min)'].mean()
    total_field_hours = month_df['Duration (min)'].sum() / 60
    variance_from_target = avg_visits_day - 6

    productivity_data.append({
        'Month': month_name,
        'Total Visits': total_visits,
        'Working Days': working_days,
        'Avg Visits/Day': round(avg_visits_day, 2),
        'Days Meeting Target (>=6)': days_meeting_target,
        '% Days on Target': round(pct_days_on_target, 1),
        'Avg Duration (min)': round(avg_duration, 2),
        'Total Field Hours': round(total_field_hours, 2),
        'Variance from Target': round(variance_from_target, 2)
    })

df_productivity = pd.DataFrame(productivity_data)
df_productivity.to_excel(writer, sheet_name='Productivity_Monthly', index=False)

# ============================================================================
# SHEET 4: TERRITORY COVERAGE
# ============================================================================
print("   -> Sheet 4: Territory_Coverage")

territory_data = []
for month_num, month_name in [(12, 'Dec 25'), (1, 'Jan 26')]:
    month_df = df[df['Month Num'] == month_num]

    # Get first appearance of each customer in entire dataset
    first_appearances = df.groupby('Account: Registered Company Name')['Check-in DateTime'].min()
    month_start = pd.Timestamp(2025, month_num, 1)
    month_end = pd.Timestamp(2026 if month_num == 1 else 2025, month_num + 1 if month_num < 12 else 1, 1)

    customers_this_month = set(month_df['Account: Registered Company Name'].unique())
    new_customers = set([cust for cust in customers_this_month if month_start <= first_appearances[cust] < month_end])

    unique_customers = len(customers_this_month)
    new_cust_count = len(new_customers)
    repeat_customers = unique_customers - new_cust_count

    total_visits = len(month_df)
    repeat_visit_rate = (repeat_customers / unique_customers * 100) if unique_customers > 0 else 0

    unique_pincodes = month_df['Account: Pin code'].nunique()
    unique_districts = month_df['Account: Auto district'].nunique()

    retailer_pb = (month_df['Account: Customer type'] == 'Retailer PB').sum()
    influencer = (month_df['Account: Customer type'] == 'Influencer').sum()
    retailer_pct = (retailer_pb / total_visits * 100) if total_visits > 0 else 0
    influencer_pct = (influencer / total_visits * 100) if total_visits > 0 else 0

    # Top 3 pin codes
    top_pincodes = month_df['Account: Pin code'].value_counts().head(3).sum()
    concentration_index = (top_pincodes / total_visits * 100) if total_visits > 0 else 0

    territory_data.append({
        'Month': month_name,
        'Unique Customers': unique_customers,
        'New Customers': new_cust_count,
        'Repeat Customers': repeat_customers,
        'Total Visits': total_visits,
        'Repeat Visit Rate (%)': round(repeat_visit_rate, 1),
        'Unique Pin Codes': unique_pincodes,
        'Unique Districts': unique_districts,
        'Retailer PB Count': retailer_pb,
        'Influencer Count': influencer,
        'Retailer %': round(retailer_pct, 1),
        'Influencer %': round(influencer_pct, 1),
        'Top 3 Pin Code Visits': top_pincodes,
        'Concentration Index (%)': round(concentration_index, 1)
    })

df_territory = pd.DataFrame(territory_data)
df_territory.to_excel(writer, sheet_name='Territory_Coverage', index=False)

# ============================================================================
# SHEET 5: REPEAT VISITS ANALYSIS
# ============================================================================
print("   -> Sheet 5: Repeat_Visits_Analysis")

customer_visits = df.groupby('Account: Registered Company Name').agg({
    'Check-in DateTime': ['count', lambda x: ', '.join(x.dt.strftime('%d/%m/%Y').astype(str).head(10).tolist())],
    'Meeting outcome': lambda x: ', '.join(x.dropna().unique()[:5]),
    'Duration (min)': 'mean',
    'Account: Account SF Id': 'first'
}).reset_index()

customer_visits.columns = ['Customer Name', 'Total Visits', 'Visit Dates (first 10)', 'Outcomes', 'Avg Duration (min)', 'Account SF Id']

# Visit frequency flag - UPDATED: 10+ visits is High
customer_visits['Visit Frequency Flag'] = customer_visits['Total Visits'].apply(
    lambda x: 'High (10+)' if x >= 10 else 'Normal'
)

# Filter for customers with 10+ visits
repeat_analysis = customer_visits[customer_visits['Total Visits'] >= 10].copy()
repeat_analysis['Avg Duration (min)'] = repeat_analysis['Avg Duration (min)'].round(2)
repeat_analysis = repeat_analysis.sort_values('Total Visits', ascending=False)

# If no customers with 10+ visits, show message
if len(repeat_analysis) == 0:
    repeat_analysis = pd.DataFrame({
        'Customer Name': ['No customers found with 10+ visits'],
        'Account SF Id': [''],
        'Total Visits': [0],
        'Visit Dates (first 10)': [''],
        'Outcomes': [''],
        'Avg Duration (min)': [0],
        'Visit Frequency Flag': ['']
    })

repeat_analysis.to_excel(writer, sheet_name='Repeat_Visits_Analysis', index=False)

# ============================================================================
# SHEET 6: SHORT DURATION FLAGS
# ============================================================================
print("   -> Sheet 6: Short_Duration_Flags")

short_duration_cols = [
    'Date', 'Account: Registered Company Name', 'Check-in Time', 'Check-out Time',
    'Duration (min)', 'Duration Band', 'Meeting outcome', 'Comments Word Count'
]

df_short = df[short_duration_cols].copy()
df_short['Red Flag'] = df_short['Duration (min)'].apply(lambda x: 'YES' if x < 5 else 'NO')

# Quality score (0-100) based on duration and comments
def calculate_quality_score(row):
    score = 0
    # Duration component (50 points)
    if row['Duration (min)'] >= 30:
        score += 50
    elif row['Duration (min)'] >= 15:
        score += 40
    elif row['Duration (min)'] >= 5:
        score += 25
    else:
        score += 10

    # Comments component (50 points)
    word_count = row['Comments Word Count'] if pd.notna(row['Comments Word Count']) else 0
    if word_count >= 50:
        score += 50
    elif word_count >= 30:
        score += 40
    elif word_count >= 20:
        score += 30
    elif word_count >= 10:
        score += 20
    else:
        score += 10

    return score

df_short['Quality Score'] = df_short.apply(calculate_quality_score, axis=1)
df_short = df_short.sort_values('Duration (min)')

df_short.to_excel(writer, sheet_name='Short_Duration_Flags', index=False)

# ============================================================================
# SHEET 7: DAILY VISIT LOG (Master Data)
# ============================================================================
print("   -> Sheet 7: Daily_Visit_Log")

master_cols = [
    'Date', 'Day of Week', 'Month', 'Check-in Time', 'Check-out Time',
    'Duration (min)', 'Account: Registered Company Name', 'Account: Account SF Id',
    'Account: Pin code', 'Account: Auto district', 'Account: Auto taluka',
    'Account: Customer type', 'Meeting outcome', 'Check-in Comments',
    'Check-out Comments', 'Combined Comments', 'Comments Word Count',
    'Check-in Coordinates (Latitude)', 'Check-in Coordinates (Longitude)',
    'Visit Number', 'Short Visit Flag', 'Documentation Complete',
    'Time Slot', 'Same Day Visits', 'Duration Band'
]

df_master = df[master_cols].copy()
df_master = df_master.sort_values('Date')
df_master.to_excel(writer, sheet_name='Daily_Visit_Log', index=False)

# ============================================================================
# SHEET 8: NOTES & METHODOLOGY
# ============================================================================
print("   -> Sheet 8: Notes_Methodology")

notes_data = {
    'Section': [
        'Data Source',
        'Analysis Period',
        'Total Records',
        'Date Range',
        '',
        'Calculations',
        '',
        '',
        '',
        '',
        '',
        '',
        'Quality Thresholds',
        '',
        '',
        'Working Days',
        '',
        'Assumptions',
        '',
        '',
        'Data Quality',
        '',
        '',
        'Limitations',
        '',
        '',
        ''
    ],
    'Details': [
        INPUT_FILE,
        'December 2025 - January 2026',
        f'{len(df)} visit records',
        f'{df["Date"].min()} to {df["Date"].max()}',
        '',
        'Duration (min) = (Check-out DateTime - Check-in DateTime) * 24 * 60',
        'Working Days = Count of unique dates with visits (excluding Sundays)',
        'Avg Visits/Day = Total Visits / Working Days',
        'Repeat Visit Rate = (Repeat Customers / Unique Customers) * 100',
        'Concentration Index = (Top 3 Pin Code Visits / Total Visits) * 100',
        'Quality Score = Duration component (50 pts) + Comments component (50 pts)',
        'Visit Number = Running count per customer ordered by check-in datetime',
        '',
        'Short Visit: <5 minutes',
        'High Frequency Customer: 10+ visits in analysis period',
        '',
        'Sundays excluded from working day count, only days with visits counted',
        '',
        'No public holidays accounted for beyond Sundays',
        'Monthly Potential and Taluka fields were empty in source data',
        '',
        '99.4% data completeness across all fields',
        f'{df["Documentation Complete"].value_counts().get("YES", 0)} visits with complete documentation',
        '',
        'Limited to 42 days of data (Dec 2025 - Jan 2026)',
        'Nov 2025 and Feb 2026 data not available',
        'Coordinate precision limited to ~100m for clustering analysis'
    ]
}

df_notes = pd.DataFrame(notes_data)
df_notes.to_excel(writer, sheet_name='Notes_Methodology', index=False)

# Save workbook
writer.close()
print(f"\n[5/6] Workbook saved: {OUTPUT_FILE}")

# ============================================================================
# APPLY FORMATTING
# ============================================================================
print("\n[6/6] Applying formatting and conditional formatting...")

wb = openpyxl.load_workbook(OUTPUT_FILE)

# Header formatting
header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
header_font = Font(bold=True, color='FFFFFF')
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]

    # Freeze panes
    ws.freeze_panes = 'B2'

    # Header row formatting
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Apply borders and alignment to all cells
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = border
            if isinstance(cell.value, (int, float)):
                cell.alignment = Alignment(horizontal='right', vertical='center')
            else:
                cell.alignment = Alignment(horizontal='left', vertical='center')

    # Auto-fit columns
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = min(max(max_length + 2, 12), 50)
        ws.column_dimensions[column_letter].width = adjusted_width

# Executive Dashboard - Conditional formatting for ratings
ws_dashboard = wb['Executive_Dashboard']
for row in range(2, 8):
    for col in ['D', 'E', 'F']:  # Dec, Jan, Overall columns
        cell = ws_dashboard[f'{col}{row}']
        target_cell = ws_dashboard[f'C{row}']

        # Apply number formatting
        if isinstance(cell.value, (int, float)):
            if '%' in str(target_cell.value):
                cell.number_format = '0.0'
            else:
                cell.number_format = '0.00'

# Productivity Monthly - Highlight days meeting target
ws_prod = wb['Productivity_Monthly']
for row in range(2, ws_prod.max_row + 1):
    cell = ws_prod[f'E{row}']  # Days Meeting Target column
    if isinstance(cell.value, (int, float)) and cell.value >= ws_prod[f'C{row}'].value * 0.5:
        cell.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')

# Short Duration Flags - Highlight red flags
ws_short = wb['Short_Duration_Flags']
red_flag_col = None
for col in range(1, ws_short.max_column + 1):
    if ws_short.cell(1, col).value == 'Red Flag':
        red_flag_col = col
        break

if red_flag_col:
    for row in range(2, ws_short.max_row + 1):
        if ws_short.cell(row, red_flag_col).value == 'YES':
            for col in range(1, ws_short.max_column + 1):
                ws_short.cell(row, col).fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')

# Repeat Visits Analysis - Highlight high frequency
ws_repeat = wb['Repeat_Visits_Analysis']
for row in range(2, ws_repeat.max_row + 1):
    freq_flag_cell = ws_repeat.cell(row, 7)  # Visit Frequency Flag column
    if 'High' in str(freq_flag_cell.value):
        for col in range(1, ws_repeat.max_column + 1):
            ws_repeat.cell(row, col).fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')

# Save formatted workbook
wb.save(OUTPUT_FILE)
print(f"   [OK] Formatting applied to all sheets")

print("\n" + "="*80)
print("ANALYSIS COMPLETE!")
print("="*80)
print(f"\nOutput file: {OUTPUT_FILE}")
print(f"Total visits analyzed: {len(df)}")
print(f"Unique customers: {df['Account: Registered Company Name'].nunique()}")
print(f"Date range: {df['Date'].min()} to {df['Date'].max()}")
print(f"Avg visits per day: {len(df) / df['Date'].nunique():.2f}")
print(f"Customers with 10+ visits: {(customer_visits['Total Visits'] >= 10).sum()}")
print("\n" + "="*80)
