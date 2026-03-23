"""
Visit Gap Analysis Tool
=======================
Reads visit data from Excel, computes inter-visit gaps per person,
and produces a dynamic Excel workbook with COUNTIFS formulas that
reference editable threshold cells.

Usage:
    python tools/visit_gap_analysis.py

Input:  Data/*.xlsx (auto-detected)
Output: Visit_Gap_Analysis.xlsx
"""

import glob
import sys
from datetime import datetime

import pandas as pd
import openpyxl
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, numbers
from openpyxl.utils import get_column_letter


# ---------------------------------------------------------------------------
# 1. Configuration
# ---------------------------------------------------------------------------
DATA_DIR = "D:/Visit Analysis/Data"
OUTPUT_PATH = "D:/Visit Analysis/Visit_Gap_Analysis_v3.xlsx"

DEFAULT_THRESHOLDS = [2, 5, 10, 30, 60]  # minutes (5 thresholds for 6 buckets)

# ---------------------------------------------------------------------------
# 2. Read & parse source data
# ---------------------------------------------------------------------------
def find_input_file():
    files = glob.glob(f"{DATA_DIR}/*.xlsx")
    if not files:
        print(f"ERROR: No .xlsx files found in {DATA_DIR}")
        sys.exit(1)
    if len(files) > 1:
        print(f"WARNING: Multiple files found, using most recent: {files[-1]}")
    return files[-1]


def parse_datetime(s):
    if pd.isna(s) or str(s).strip() == "":
        return None
    try:
        return datetime.strptime(str(s).strip(), "%d/%m/%Y, %I:%M %p")
    except ValueError:
        return None


def load_data(filepath):
    print(f"Reading: {filepath}")
    df = pd.read_excel(filepath)
    print(f"  Rows: {len(df)}, Columns: {len(df.columns)}")

    df["checkin_dt"] = df["Check-in Date/Time"].apply(parse_datetime)
    df["checkout_dt"] = df["Check-out Date/Time"].apply(parse_datetime)

    null_checkins = df["checkin_dt"].isna().sum()
    null_checkouts = df["checkout_dt"].isna().sum()
    print(f"  Null checkins: {null_checkins}, Null checkouts: {null_checkouts}")

    # Drop rows with no checkin (unusable)
    df = df.dropna(subset=["checkin_dt"]).copy()
    return df


# ---------------------------------------------------------------------------
# 3. Compute inter-visit gaps
# ---------------------------------------------------------------------------
def compute_primary_states(df):
    """Compute primary (most-visited) state per person. Normalizes case."""
    state_col = "Account: Auto state"
    df_valid = df.dropna(subset=[state_col]).copy()
    df_valid["state_upper"] = df_valid[state_col].str.upper().str.strip()

    # Count visits per person per normalized state
    counts = df_valid.groupby(["Visit: Created By", "state_upper"]).size().reset_index(name="n")
    # Pick the state with most visits per person
    idx = counts.groupby("Visit: Created By")["n"].idxmax()
    primary = counts.loc[idx].set_index("Visit: Created By")["state_upper"].to_dict()

    # Title-case for display
    primary = {k: v.title() for k, v in primary.items()}

    print(f"  Primary states computed for {len(primary)} people")
    return primary


def compute_gaps(df):
    gap_rows = []

    for person, grp in df.groupby("Visit: Created By"):
        grp = grp.sort_values("checkin_dt").reset_index(drop=True)

        for i in range(1, len(grp)):
            prev_ci = grp.loc[i - 1, "checkin_dt"]
            prev_co = grp.loc[i - 1, "checkout_dt"]
            curr_ci = grp.loc[i, "checkin_dt"]

            # Determine effective checkout and flag
            if pd.isna(prev_co):
                effective_co = prev_ci
                flag = "NULL_CHECKOUT"
            elif prev_co < prev_ci:
                effective_co = prev_ci
                flag = "ANOMALY"
            else:
                effective_co = prev_co
                flag = "OK"

            gap_minutes = (curr_ci - effective_co).total_seconds() / 60

            # Cross-day check (only upgrade OK -> CROSS_DAY)
            if flag == "OK" and effective_co.date() != curr_ci.date():
                flag = "CROSS_DAY"

            # Negative gap sanity check (shouldn't happen after fixes, but guard)
            if gap_minutes < 0:
                flag = "ANOMALY"
                gap_minutes = 0

            month_str = curr_ci.strftime("%Y-%m")

            gap_rows.append({
                "person": person,
                "month": month_str,
                "gap_min": round(gap_minutes, 1),
                "flag": flag,
                "prev_co": effective_co,
                "curr_ci": curr_ci,
            })

    print(f"  Computed {len(gap_rows)} gap rows")

    # Flag distribution
    from collections import Counter
    flags = Counter(r["flag"] for r in gap_rows)
    for f, c in sorted(flags.items()):
        print(f"    {f}: {c}")

    return gap_rows


# ---------------------------------------------------------------------------
# 4. Build Excel workbook
# ---------------------------------------------------------------------------
def build_workbook(gap_rows, df, primary_states):
    wb = openpyxl.Workbook()

    # --- Detect months and people ---
    months_set = sorted(set(r["month"] for r in gap_rows))
    month_labels = []
    for m in months_set:
        dt = datetime.strptime(m, "%Y-%m")
        month_labels.append(dt.strftime("%b %Y"))

    people = sorted(set(r["person"] for r in gap_rows))
    visit_counts = df.groupby("Visit: Created By").size().to_dict()

    num_months = len(months_set)
    num_buckets = 6
    num_cols_per_month = 7  # 6 buckets + 1 subtotal

    print(f"  Months: {month_labels}")
    print(f"  People: {len(people)}")

    # ===================================================================
    # SHEET 1: Config
    # ===================================================================
    ws_config = wb.active
    ws_config.title = "Config"

    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    white_font = Font(bold=True, size=11, color="FFFFFF")

    ws_config["A1"] = "Threshold Name"
    ws_config["B1"] = "Value (minutes)"
    ws_config["A1"].font = white_font
    ws_config["B1"].font = white_font
    ws_config["A1"].fill = header_fill
    ws_config["B1"].fill = header_fill

    ws_config["A2"] = "Bucket 1 Upper Limit"
    ws_config["B2"] = DEFAULT_THRESHOLDS[0]
    ws_config["A3"] = "Bucket 2 Upper Limit"
    ws_config["B3"] = DEFAULT_THRESHOLDS[1]
    ws_config["A4"] = "Bucket 3 Upper Limit"
    ws_config["B4"] = DEFAULT_THRESHOLDS[2]
    ws_config["A5"] = "Bucket 4 Upper Limit"
    ws_config["B5"] = DEFAULT_THRESHOLDS[3]
    ws_config["A6"] = "Bucket 5 Upper Limit"
    ws_config["B6"] = DEFAULT_THRESHOLDS[4]

    ws_config["A8"] = "Instructions:"
    ws_config["A9"] = "Change the values in column B to adjust time buckets."
    ws_config["A10"] = "The Summary sheet formulas will recalculate automatically."
    ws_config["A8"].font = Font(bold=True, italic=True)

    ws_config.column_dimensions["A"].width = 28
    ws_config.column_dimensions["B"].width = 20

    # Highlight threshold cells
    thresh_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    for row in range(2, 7):  # rows 2-6 for 5 thresholds
        ws_config.cell(row=row, column=2).fill = thresh_fill
        ws_config.cell(row=row, column=2).border = thin_border
        ws_config.cell(row=row, column=2).font = Font(bold=True, size=12)

    # Named ranges for thresholds (5 thresholds for 6 buckets)
    for name, cell_ref in [
        ("Thresh1", "$B$2"),
        ("Thresh2", "$B$3"),
        ("Thresh3", "$B$4"),
        ("Thresh4", "$B$5"),
        ("Thresh5", "$B$6"),
    ]:
        dn = DefinedName(name, attr_text=f"Config!{cell_ref}")
        wb.defined_names.add(dn)

    # ===================================================================
    # SHEET 2: Data
    # ===================================================================
    ws_data = wb.create_sheet("Data")

    data_headers = ["State", "Person", "Month", "Gap_Minutes", "Flag", "Prev_Checkout", "Curr_Checkin"]
    for col_idx, h in enumerate(data_headers, 1):
        cell = ws_data.cell(row=1, column=col_idx, value=h)
        cell.font = white_font
        cell.fill = header_fill

    for row_idx, gap in enumerate(gap_rows, 2):
        ws_data.cell(row=row_idx, column=1, value=primary_states.get(gap["person"], "Unknown"))
        ws_data.cell(row=row_idx, column=2, value=gap["person"])
        ws_data.cell(row=row_idx, column=3, value=gap["month"])
        c = ws_data.cell(row=row_idx, column=4, value=gap["gap_min"])
        c.number_format = "0.0"
        ws_data.cell(row=row_idx, column=5, value=gap["flag"])
        ws_data.cell(row=row_idx, column=6, value=gap["prev_co"])
        ws_data.cell(row=row_idx, column=7, value=gap["curr_ci"])

    last_data_row = len(gap_rows) + 1

    # Column widths
    ws_data.column_dimensions["A"].width = 22
    ws_data.column_dimensions["B"].width = 28
    ws_data.column_dimensions["C"].width = 10
    ws_data.column_dimensions["D"].width = 14
    ws_data.column_dimensions["E"].width = 16
    ws_data.column_dimensions["F"].width = 20
    ws_data.column_dimensions["G"].width = 20

    # Auto-filter
    ws_data.auto_filter.ref = f"A1:G{last_data_row}"

    # Freeze header
    ws_data.freeze_panes = "A2"

    # Named ranges for data columns (B=Person, C=Month, D=Gap_Minutes, E=Flag)
    range_defs = {
        "GapPerson": f"Data!$B$2:$B${last_data_row}",
        "GapMonth": f"Data!$C$2:$C${last_data_row}",
        "GapMinutes": f"Data!$D$2:$D${last_data_row}",
        "GapFlag": f"Data!$E$2:$E${last_data_row}",
    }
    for name, ref in range_defs.items():
        dn = DefinedName(name, attr_text=ref)
        wb.defined_names.add(dn)

    # ===================================================================
    # SHEET 3: Summary
    # ===================================================================
    ws_sum = wb.create_sheet("Summary")

    # --- Styling definitions ---
    month_colors = ["4472C4", "548235", "BF8F00"]  # blue, green, gold
    subtotal_color = "8DB4E2"  # lighter blue for subtotal columns
    total_color = "7030A0"  # purple

    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="center")

    # Column offset: State=1, Person=2, then month groups start at 3
    COL_STATE = 1
    COL_PERSON = 2
    MONTH_START = 3  # first month group starts here

    # --- Row 1: Merged headers for State and Person ---
    dark_fill = PatternFill(start_color="333333", end_color="333333", fill_type="solid")
    dark_font = Font(bold=True, size=11, color="FFFFFF")

    cell_state = ws_sum.cell(row=1, column=COL_STATE, value="State")
    cell_state.font = dark_font
    cell_state.fill = dark_fill
    cell_state.alignment = center_align
    ws_sum.merge_cells(start_row=1, start_column=COL_STATE, end_row=2, end_column=COL_STATE)

    cell_person = ws_sum.cell(row=1, column=COL_PERSON, value="Person")
    cell_person.font = dark_font
    cell_person.fill = dark_fill
    cell_person.alignment = center_align
    ws_sum.merge_cells(start_row=1, start_column=COL_PERSON, end_row=2, end_column=COL_PERSON)

    for m_idx, m_label in enumerate(month_labels):
        start_col = MONTH_START + m_idx * num_cols_per_month
        end_col = start_col + num_cols_per_month - 1

        color = month_colors[m_idx % len(month_colors)]
        fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        font = Font(bold=True, size=11, color="FFFFFF")

        ws_sum.merge_cells(
            start_row=1, start_column=start_col,
            end_row=1, end_column=end_col,
        )
        cell = ws_sum.cell(row=1, column=start_col, value=m_label)
        cell.font = font
        cell.fill = fill
        cell.alignment = center_align

        # Fill merged cells
        for c in range(start_col + 1, end_col + 1):
            ws_sum.cell(row=1, column=c).fill = fill

    # Total Visits column
    total_col = MONTH_START + num_months * num_cols_per_month
    cell_tv = ws_sum.cell(row=1, column=total_col, value="Total Visits")
    cell_tv.font = Font(bold=True, size=11, color="FFFFFF")
    cell_tv.fill = PatternFill(start_color=total_color, end_color=total_color, fill_type="solid")
    cell_tv.alignment = center_align
    ws_sum.merge_cells(start_row=1, start_column=total_col, end_row=2, end_column=total_col)

    # --- Row 2: Sub-headers (dynamic formulas + subtotal) ---
    for m_idx in range(num_months):
        base_col = MONTH_START + m_idx * num_cols_per_month
        # Lighter shade for sub-headers
        light_colors = ["D6E4F0", "E2EFDA", "FFF2CC"]
        light_fill = PatternFill(
            start_color=light_colors[m_idx % len(light_colors)],
            end_color=light_colors[m_idx % len(light_colors)],
            fill_type="solid",
        )
        sub_font = Font(bold=True, size=10)
        subtotal_fill = PatternFill(
            start_color=subtotal_color, end_color=subtotal_color, fill_type="solid"
        )

        # Bucket 1: <Thresh1 (e.g. <2 min)
        c1 = ws_sum.cell(row=2, column=base_col)
        c1.value = '="<"&Thresh1&" min"'
        c1.font = sub_font
        c1.fill = light_fill
        c1.alignment = center_align

        # Bucket 2: Thresh1 - Thresh2 (e.g. 2-5 min)
        c2 = ws_sum.cell(row=2, column=base_col + 1)
        c2.value = '=Thresh1&"-"&Thresh2&" min"'
        c2.font = sub_font
        c2.fill = light_fill
        c2.alignment = center_align

        # Bucket 3: Thresh2 - Thresh3 (e.g. 5-10 min)
        c3 = ws_sum.cell(row=2, column=base_col + 2)
        c3.value = '=Thresh2&"-"&Thresh3&" min"'
        c3.font = sub_font
        c3.fill = light_fill
        c3.alignment = center_align

        # Bucket 4: Thresh3 - Thresh4 (e.g. 10-30 min)
        c4 = ws_sum.cell(row=2, column=base_col + 3)
        c4.value = '=Thresh3&"-"&Thresh4&" min"'
        c4.font = sub_font
        c4.fill = light_fill
        c4.alignment = center_align

        # Bucket 5: Thresh4 - Thresh5 (e.g. 30-60 min)
        c5 = ws_sum.cell(row=2, column=base_col + 4)
        c5.value = '=Thresh4&"-"&Thresh5&" min"'
        c5.font = sub_font
        c5.fill = light_fill
        c5.alignment = center_align

        # Bucket 6: >=Thresh5 (e.g. >=60 min)
        c6 = ws_sum.cell(row=2, column=base_col + 5)
        c6.value = '=">="&Thresh5&" min"'
        c6.font = sub_font
        c6.fill = light_fill
        c6.alignment = center_align

        # Subtotal (col base_col+6)
        c7 = ws_sum.cell(row=2, column=base_col + 6)
        c7.value = "Subtotal"
        c7.font = Font(bold=True, size=10, color="FFFFFF")
        c7.fill = subtotal_fill
        c7.alignment = center_align

    # --- Helper: build COUNTIFS fragment for a bucket ---
    def _countifs(person_ref, m_key, lower_expr, upper_expr):
        """Return the COUNTIFS(...) string for a single bucket."""
        if upper_expr is None:
            # Last bucket: >= threshold (e.g. >=60 min)
            return (
                f'COUNTIFS(GapPerson,{person_ref},GapMonth,"{m_key}",'
                f'GapMinutes,">="&{lower_expr},GapFlag,"OK")'
            )
        return (
            f'COUNTIFS(GapPerson,{person_ref},GapMonth,"{m_key}",'
            f'GapMinutes,">="&{lower_expr},GapMinutes,"<"&{upper_expr},GapFlag,"OK")'
        )

    # Bucket definitions: (lower_expr, upper_expr) for 6 buckets
    bucket_defs = [
        ("0",       "Thresh1"),  # <2 min
        ("Thresh1", "Thresh2"),  # 2-5 min
        ("Thresh2", "Thresh3"),  # 5-10 min
        ("Thresh3", "Thresh4"),  # 10-30 min
        ("Thresh4", "Thresh5"),  # 30-60 min
        ("Thresh5", None),       # >=60 min
    ]

    # --- Rows 3+: Person rows with COUNTIFS formulas ---
    alt_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

    for p_idx, person in enumerate(people):
        row = 3 + p_idx
        is_alt = p_idx % 2 == 1
        row_fill = alt_fill if is_alt else PatternFill(fill_type=None)

        # State
        cell_state = ws_sum.cell(row=row, column=COL_STATE, value=primary_states.get(person, "Unknown"))
        cell_state.font = Font(size=10)
        cell_state.alignment = left_align
        if is_alt:
            cell_state.fill = row_fill

        # Person name
        cell_name = ws_sum.cell(row=row, column=COL_PERSON, value=person)
        cell_name.font = Font(size=10)
        cell_name.alignment = left_align
        if is_alt:
            cell_name.fill = row_fill

        person_ref = f"$B${row}"

        for m_idx, m_key in enumerate(months_set):
            base_col = MONTH_START + m_idx * num_cols_per_month
            subtotal_col_idx = base_col + 6
            sub_letter = get_column_letter(subtotal_col_idx)
            sub_ref = f"${sub_letter}${row}"

            # Subtotal column: total OK gaps for this person+month (numeric)
            f_sub = (
                f'=COUNTIFS(GapPerson,{person_ref},GapMonth,"{m_key}",GapFlag,"OK")'
            )
            c_sub = ws_sum.cell(row=row, column=subtotal_col_idx, value=f_sub)
            c_sub.alignment = center_align
            c_sub.font = Font(bold=True, size=10)
            if is_alt:
                c_sub.fill = row_fill

            # 6 bucket columns: "xx (yy%)" format
            for b_idx, (lower, upper) in enumerate(bucket_defs):
                col_idx = base_col + b_idx
                cf = _countifs(person_ref, m_key, lower, upper)
                # Formula: COUNTIFS(...)&" ("&TEXT(IFERROR(COUNTIFS(...)/subtotal,0),"0%")&")"
                formula = (
                    f'={cf}'
                    f'&" ("&TEXT(IFERROR({cf}/{sub_ref},0),"0%")&")"'
                )
                c = ws_sum.cell(row=row, column=col_idx, value=formula)
                c.alignment = center_align
                if is_alt:
                    c.fill = row_fill

        # Total Visits (static value from source data)
        tv = ws_sum.cell(row=row, column=total_col, value=visit_counts.get(person, 0))
        tv.alignment = center_align
        tv.font = Font(bold=True, size=10)
        if is_alt:
            tv.fill = row_fill

    # --- Grand Total Row ---
    total_row = 3 + len(people)
    total_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
    total_font = Font(bold=True, size=11)

    # Merge State + Person cells for TOTAL label
    ws_sum.merge_cells(start_row=total_row, start_column=COL_STATE, end_row=total_row, end_column=COL_PERSON)
    cell_total_label = ws_sum.cell(row=total_row, column=COL_STATE, value="TOTAL")
    cell_total_label.font = total_font
    cell_total_label.fill = total_fill
    cell_total_label.alignment = center_align
    ws_sum.cell(row=total_row, column=COL_PERSON).fill = total_fill

    for m_idx, m_key in enumerate(months_set):
        base_col = MONTH_START + m_idx * num_cols_per_month
        subtotal_col_idx = base_col + 6  # 7th col in month group (after 6 buckets)
        sub_letter = get_column_letter(subtotal_col_idx)
        sub_ref = f"${sub_letter}${total_row}"

        # Subtotal total: SUM of subtotal column (numeric)
        sub_col_letter = get_column_letter(subtotal_col_idx)
        f_sub_total = f"=SUM({sub_col_letter}3:{sub_col_letter}{total_row - 1})"
        c_sub = ws_sum.cell(row=total_row, column=subtotal_col_idx, value=f_sub_total)
        c_sub.font = total_font
        c_sub.fill = total_fill
        c_sub.alignment = center_align

        # Bucket totals: numeric COUNTIFS with percentage
        person_ref = '"*"'  # wildcard to match all
        for b_idx, (lower, upper) in enumerate(bucket_defs):
            col_idx = base_col + b_idx
            cf = _countifs(person_ref, m_key, lower, upper)
            formula = (
                f'={cf}'
                f'&" ("&TEXT(IFERROR({cf}/{sub_ref},0),"0%")&")"'
            )
            c = ws_sum.cell(row=total_row, column=col_idx, value=formula)
            c.font = total_font
            c.fill = total_fill
            c.alignment = center_align

    # Total Visits grand total
    tv_letter = get_column_letter(total_col)
    f_tv_total = f"=SUM({tv_letter}3:{tv_letter}{total_row - 1})"
    c_tv = ws_sum.cell(row=total_row, column=total_col, value=f_tv_total)
    c_tv.font = total_font
    c_tv.fill = total_fill
    c_tv.alignment = center_align

    # --- Column widths ---
    ws_sum.column_dimensions["A"].width = 22  # State
    ws_sum.column_dimensions["B"].width = 28  # Person
    for col in range(MONTH_START, total_col + 1):
        ws_sum.column_dimensions[get_column_letter(col)].width = 14

    # --- Freeze panes (freeze State + Person columns and header rows) ---
    ws_sum.freeze_panes = "C3"

    # --- Thin borders for data area ---
    for row in range(1, total_row + 1):
        for col in range(1, total_col + 1):
            ws_sum.cell(row=row, column=col).border = thin_border

    # ===================================================================
    # SHEET 4: About (Documentation)
    # ===================================================================
    ws_about = wb.create_sheet("About")

    title_font = Font(bold=True, size=14, color="4472C4")
    section_font = Font(bold=True, size=12, color="333333")
    body_font = Font(size=11)
    note_font = Font(size=11, italic=True, color="666666")

    ws_about.column_dimensions["A"].width = 4
    ws_about.column_dimensions["B"].width = 80

    content = [
        (title_font, "Visit Gap Analysis - Documentation"),
        (None, ""),
        (section_font, "1. Purpose"),
        (body_font, "This workbook analyzes the time gaps between consecutive field visits"),
        (body_font, "for each sales team member. A 'gap' is the duration from checkout of one"),
        (body_font, "visit to checkin of the next visit by the same person."),
        (None, ""),
        (section_font, "2. How Gaps Are Calculated"),
        (body_font, "- Visits are grouped by person ('Visit: Created By')"),
        (body_font, "- Within each person, visits are sorted by checkin time"),
        (body_font, "- Gap = checkin_time[next visit] - checkout_time[current visit]"),
        (body_font, "- Gaps are measured in minutes"),
        (body_font, "- Only same-day gaps with valid data (Flag = 'OK') are counted in Summary"),
        (None, ""),
        (section_font, "3. Time Buckets (6 buckets, 5 thresholds)"),
        (body_font, "Default thresholds: 2, 5, 10, 30, 60 minutes"),
        (body_font, "  Bucket 1: < 2 min   (Thresh1)"),
        (body_font, "  Bucket 2: 2 - 5 min  (Thresh1 to Thresh2)"),
        (body_font, "  Bucket 3: 5 - 10 min (Thresh2 to Thresh3)"),
        (body_font, "  Bucket 4: 10 - 30 min (Thresh3 to Thresh4)"),
        (body_font, "  Bucket 5: 30 - 60 min (Thresh4 to Thresh5)"),
        (body_font, "  Bucket 6: >= 60 min  (Thresh5)"),
        (body_font, "All 5 thresholds are editable on the Config sheet (cells B2:B6)."),
        (None, ""),
        (section_font, "4. Edge Cases Handled"),
        (body_font, "NULL_CHECKOUT (136 visits):"),
        (note_font, "  Checkout time was missing. Used checkin time as fallback."),
        (note_font, "  These gaps are flagged and excluded from Summary counts."),
        (body_font, "ANOMALY (18 visits):"),
        (note_font, "  Checkout time was earlier than checkin (data entry error)."),
        (note_font, "  Used checkin time as fallback. Excluded from Summary counts."),
        (body_font, "CROSS_DAY (3,148 gaps):"),
        (note_font, "  Gap spans across different calendar days (overnight/weekend)."),
        (note_font, "  Excluded from Summary to avoid inflating the 60+ min bucket."),
        (body_font, "First visit per person:"),
        (note_font, "  No gap is generated (no preceding visit to measure from)."),
        (body_font, "Zero-duration visits (5,442):"),
        (note_font, "  Checkout equals checkin. Processed normally; gap is measured from checkin."),
        (None, ""),
        (section_font, "5. How to Use the Config Sheet"),
        (body_font, "- Go to the 'Config' sheet"),
        (body_font, "- Change values in column B (Bucket 1-5 Upper Limits, cells B2:B6)"),
        (body_font, "- Example: Change Thresh1 from 2 to 3, and '<2 min' columns become '<3 min'"),
        (body_font, "- All Summary formulas recalculate automatically"),
        (None, ""),
        (section_font, "6. Column Format in Summary"),
        (body_font, "Bucket cells show: xx (yy%)"),
        (note_font, "  xx  = number of gaps falling in that time bucket"),
        (note_font, "  yy% = xx divided by the month subtotal (% of that month's gaps)"),
        (body_font, "Subtotal column: total same-day valid gaps for that person in that month"),
        (body_font, "Total Visits: total original visit records from source data"),
        (None, ""),
        (section_font, "7. Data Sheet"),
        (body_font, "Contains one row per computed gap with columns:"),
        (body_font, "  State, Person, Month, Gap_Minutes, Flag, Prev_Checkout, Curr_Checkin"),
        (body_font, "Use the auto-filter to explore gap data by person, month, or flag."),
        (body_font, "The Summary sheet COUNTIFS formulas reference this data via named ranges."),
    ]

    for row_idx, (font, text) in enumerate(content, 1):
        cell = ws_about.cell(row=row_idx, column=2, value=text)
        if font:
            cell.font = font

    # --- Set Summary as the active sheet ---
    wb.active = wb.sheetnames.index("Summary")

    return wb


# ---------------------------------------------------------------------------
# 5. Main
# ---------------------------------------------------------------------------
def main():
    print("=" * 60)
    print("Visit Gap Analysis Tool")
    print("=" * 60)

    # Step 1: Load data
    input_file = find_input_file()
    df = load_data(input_file)

    # Step 2: Compute gaps and primary states
    print("\nComputing inter-visit gaps...")
    gap_rows = compute_gaps(df)
    print("\nComputing primary states...")
    primary_states = compute_primary_states(df)

    # Step 3: Build workbook
    print("\nBuilding Excel workbook...")
    wb = build_workbook(gap_rows, df, primary_states)

    # Step 4: Save
    wb.save(OUTPUT_PATH)
    print(f"\nSaved: {OUTPUT_PATH}")
    print("Done!")


if __name__ == "__main__":
    main()
