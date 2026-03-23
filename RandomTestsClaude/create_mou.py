import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# ===========================
# SHEET 1: MOU SLAB TABLE
# ===========================
ws1 = wb.active
ws1.title = "MOU Slab Structure"

# Colors
jsw_blue = "1F4E79"
jsw_light_blue = "D6E4F0"
jsw_gold = "FFC000"
white_c = "FFFFFF"
black_c = "000000"
light_gray = "F2F2F2"
border_color = "8DB4E2"

# Fonts
title_font = Font(name="Calibri", size=18, bold=True, color=white_c)
subtitle_font = Font(name="Calibri", size=13, bold=True, color=jsw_blue)
header_font = Font(name="Calibri", size=11, bold=True, color=white_c)
data_font = Font(name="Calibri", size=11, color=black_c)
data_bold = Font(name="Calibri", size=11, bold=True, color=black_c)
highlight_font = Font(name="Calibri", size=11, bold=True, color="C00000")
note_font = Font(name="Calibri", size=10, italic=True, color="404040")
product_font = Font(name="Calibri", size=11, color=jsw_blue)

# Borders
thin_border = Border(
    left=Side(style="thin", color=border_color),
    right=Side(style="thin", color=border_color),
    top=Side(style="thin", color=border_color),
    bottom=Side(style="thin", color=border_color),
)

# Fills
blue_fill = PatternFill(start_color=jsw_blue, end_color=jsw_blue, fill_type="solid")
light_blue_fill = PatternFill(start_color=jsw_light_blue, end_color=jsw_light_blue, fill_type="solid")
gold_fill = PatternFill(start_color=jsw_gold, end_color=jsw_gold, fill_type="solid")
gray_fill = PatternFill(start_color=light_gray, end_color=light_gray, fill_type="solid")
white_fill = PatternFill(start_color=white_c, end_color=white_c, fill_type="solid")

# Column widths
ws1.column_dimensions["A"].width = 4
ws1.column_dimensions["B"].width = 10
ws1.column_dimensions["C"].width = 30
ws1.column_dimensions["D"].width = 25
ws1.column_dimensions["E"].width = 35
ws1.column_dimensions["F"].width = 4

# Row 1-2: Title banner
ws1.merge_cells("A1:F2")
cell = ws1["A1"]
cell.value = "JSW ONE TMT"
cell.font = title_font
cell.fill = blue_fill
cell.alignment = Alignment(horizontal="center", vertical="center")
ws1.row_dimensions[1].height = 25
ws1.row_dimensions[2].height = 25

# Row 3: Subtitle
ws1.merge_cells("A3:F3")
cell = ws1["A3"]
cell.value = "PROJECT CUSTOMER MOU - INCENTIVE STRUCTURE FY 2025-26"
cell.font = subtitle_font
cell.fill = light_blue_fill
cell.alignment = Alignment(horizontal="center", vertical="center")
ws1.row_dimensions[3].height = 30

# Row 4: blank
ws1.row_dimensions[4].height = 10

# Row 5: Product info
ws1.merge_cells("B5:E5")
cell = ws1["B5"]
cell.value = "Product: JSW ONE TMT Bars (Straight Length / Coils / Cut & Bend)"
cell.font = product_font
cell.alignment = Alignment(horizontal="left", vertical="center")

# Row 6: Date
ws1.merge_cells("B6:E6")
cell = ws1["B6"]
cell.value = "MOU Period: 1st April 2025 to 31st March 2026"
cell.font = product_font
cell.alignment = Alignment(horizontal="left", vertical="center")

# Row 7: blank
ws1.row_dimensions[7].height = 15

# Row 8: Table header
headers = ["Slab", "Annual MOU Volume (MT)", "MOU Benefit (Rs/MT)", "Special Reward"]
cols = ["B", "C", "D", "E"]
ws1.row_dimensions[8].height = 35
for col, header in zip(cols, headers):
    cell = ws1[f"{col}8"]
    cell.value = header
    cell.font = header_font
    cell.fill = blue_fill
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = thin_border

# Data rows
slab_data = [
    ["1", "10,000 - 12,500", 200, "-"],
    ["2", "12,501 - 15,000", 250, "-"],
    ["3", "15,001 - 20,000", 300, "Thailand Trip for 4 Persons"],
    ["4", "20,001 - 25,000", 350, "Europe Trip for 4 Persons"],
]

for idx, (slab, volume, rate, reward) in enumerate(slab_data):
    row = 9 + idx
    ws1.row_dimensions[row].height = 30
    fill = gray_fill if idx % 2 == 0 else white_fill

    cell = ws1[f"B{row}"]
    cell.value = slab
    cell.font = data_bold
    cell.fill = fill
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = thin_border

    cell = ws1[f"C{row}"]
    cell.value = volume
    cell.font = data_font
    cell.fill = fill
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = thin_border

    cell = ws1[f"D{row}"]
    cell.value = rate
    cell.font = data_bold
    cell.fill = fill
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.number_format = "#,##0"
    cell.border = thin_border

    cell = ws1[f"E{row}"]
    cell.value = reward
    cell.font = highlight_font if reward != "-" else data_font
    cell.fill = gold_fill if reward != "-" else fill
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = thin_border

# Blank row
ws1.row_dimensions[13].height = 10

# Notes
ws1.merge_cells("B14:E14")
cell = ws1["B14"]
cell.value = "Note: MOU Benefit is payable in Rs per MT on total dispatched quantity within the signed slab."
cell.font = note_font
cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

ws1.merge_cells("B15:E15")
cell = ws1["B15"]
cell.value = "Trip rewards are redeemable only after 100% achievement of annual MOU volume."
cell.font = note_font
cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

ws1.merge_cells("B16:E16")
cell = ws1["B16"]
cell.value = "Maximum payout capped at 110% of signed MOU quantity."
cell.font = note_font
cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

# Footer
ws1.merge_cells("B18:E18")
cell = ws1["B18"]
cell.value = "Confidential - JSW ONE TMT | For Authorized Use Only"
cell.font = Font(name="Calibri", size=9, italic=True, color="808080")
cell.alignment = Alignment(horizontal="center", vertical="center")

# Print settings
ws1.sheet_properties.pageSetUpPr = openpyxl.worksheet.properties.PageSetupProperties(fitToPage=True)
ws1.page_setup.fitToWidth = 1
ws1.page_setup.fitToHeight = 1
ws1.page_setup.orientation = "landscape"


# ===========================
# SHEET 2: TERMS & CONDITIONS
# ===========================
ws2 = wb.create_sheet("Terms & Conditions")

ws2.column_dimensions["A"].width = 4
ws2.column_dimensions["B"].width = 6
ws2.column_dimensions["C"].width = 90
ws2.column_dimensions["D"].width = 4

# Row 1-2: Title
ws2.merge_cells("A1:D2")
cell = ws2["A1"]
cell.value = "JSW ONE TMT - PROJECT MOU"
cell.font = title_font
cell.fill = blue_fill
cell.alignment = Alignment(horizontal="center", vertical="center")
ws2.row_dimensions[1].height = 25
ws2.row_dimensions[2].height = 25

# Row 3: Subtitle
ws2.merge_cells("A3:D3")
cell = ws2["A3"]
cell.value = "TERMS & CONDITIONS"
cell.font = subtitle_font
cell.fill = light_blue_fill
cell.alignment = Alignment(horizontal="center", vertical="center")
ws2.row_dimensions[3].height = 30

ws2.row_dimensions[4].height = 10

# T&C sections
sections = [
    ("CONSISTENCY & MINIMUM LIFTING", [
        "Minimum quarterly lifting shall be 20% of the annual MOU quantity (evenly spread across 4 quarters).",
        "Minimum monthly lifting shall be 6% of the annual MOU quantity.",
        "Lean Period Relaxation: Up to 2 months in a year (non-consecutive) where the monthly minimum can be relaxed to 4% instead of 6%.",
        "Lean period months must fall within Q2 (Jul-Sep) or Q3 (Oct-Dec) only.",
    ]),
    ("PAYMENT & SETTLEMENT", [
        "MOU benefit payout: 50% at half-year (on achieving 50% of signed volume), balance 50% at year-end on 100% completion.",
        "Trip rewards are redeemable only after 100% annual volume achievement.",
        "All payouts shall be processed within 30 days of meeting the qualification criteria.",
    ]),
    ("VOLUME & QUALIFICATION", [
        "MOU is applicable for JSW ONE TMT Bars (Straight Length / Coils / Cut & Bend) only.",
        "Volumes shall be counted on dispatched (billed) quantities. Auction and complaint quantities are excluded.",
        "Maximum payout shall be capped at 110% of the signed MOU quantity.",
        "MOU quantity shall be signed in multiples of 100 MT.",
    ]),
    ("ENHANCEMENT", [
        "One-time enhancement of MOU quantity is allowed after Q2 completion (between 1st - 15th October 2025).",
        "Enhanced quantity shall be mutually agreed. The incentive rate shall remain as per the original signed slab.",
    ]),
    ("GENERAL", [
        "MOU validity: 1st April 2025 to 31st March 2026.",
        "MOU shall be signed by both parties on or before 30th June 2025.",
        "JSW reserves the right to supply from any of its plant / stockyard / service center locations.",
        "Dispatch mode (Rail / Road / Sea) shall be at JSW's discretion.",
        "JSW may terminate the MOU with 7 days written notice in case of any breach or misconduct by the customer.",
        "All unpaid incentives and trip benefits shall stand forfeited upon termination.",
        "This MOU shall be governed by the Laws of India.",
    ]),
]

row = 5
section_num = 1
for section_title, points in sections:
    # Section header
    ws2.merge_cells(f"B{row}:C{row}")
    cell = ws2[f"B{row}"]
    cell.value = f"{section_num}. {section_title}"
    cell.font = Font(name="Calibri", size=11, bold=True, color=white_c)
    cell.fill = blue_fill
    cell.alignment = Alignment(horizontal="left", vertical="center")
    cell.border = thin_border
    ws2.row_dimensions[row].height = 28
    row += 1

    # Points
    for i, point in enumerate(points):
        letter = chr(97 + i)

        cell = ws2[f"B{row}"]
        cell.value = f"{letter}."
        cell.font = data_font
        cell.alignment = Alignment(horizontal="center", vertical="top")
        cell.fill = gray_fill if i % 2 == 0 else white_fill
        cell.border = thin_border

        cell = ws2[f"C{row}"]
        cell.value = point
        cell.font = data_font
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        cell.fill = gray_fill if i % 2 == 0 else white_fill
        cell.border = thin_border
        ws2.row_dimensions[row].height = 30
        row += 1

    row += 1
    section_num += 1

# Footer
row += 1
ws2.merge_cells(f"B{row}:C{row}")
cell = ws2[f"B{row}"]
cell.value = "Confidential - JSW ONE TMT | For Authorized Use Only"
cell.font = Font(name="Calibri", size=9, italic=True, color="808080")
cell.alignment = Alignment(horizontal="center", vertical="center")

# Print settings
ws2.sheet_properties.pageSetUpPr = openpyxl.worksheet.properties.PageSetupProperties(fitToPage=True)
ws2.page_setup.fitToWidth = 1
ws2.page_setup.orientation = "portrait"

# Save
output_path = "D:/RandomTestsClaude/JSW_ONE_TMT_Project_MOU_FY2025-26.xlsx"
wb.save(output_path)
print(f"File saved successfully: {output_path}")
