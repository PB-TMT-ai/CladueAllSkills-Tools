import sys
import io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

from copy import copy
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

INPUT_FILE = r"C:\Users\2750834\Downloads\Distributor wise meets data.xlsx"
OUTPUT_FILE = r"C:\Users\2750834\Downloads\Distributor wise meets data - Updated.xlsx"

# ── Name mapping: monthly tab variants -> standard first-tab names ──
NAME_MAP = {
    # Nov inconsistencies
    "NIYATI UDYOG": "Niyati Udyog Private Limited",
    "SSTL INDIA PVT. LMT.": "SSTL",
    "R K Enterprises": "R. K. ENTERPRISES",
    "R K Enterprises ": "R. K. ENTERPRISES",
    "Goldwin Ispat": "GOLDWIN ISPAT PRIVATE LIMITED",
    "Kharakia Merals Pvt Ltd": "KHARAKIA METALS PRIVATE LIMITED",
    "Kharakia Merals Pvt Ltd ": "KHARAKIA METALS PRIVATE LIMITED",
    "Sachdeva ispat": "SACHDEVA ISPAT PVT LTD",
    "Sachdeva ispat ": "SACHDEVA ISPAT PVT LTD",
    " Real Dreams Enterprises": "Real dreams enterprises",
    "Global Steel ": "Global Steel",
    "Global iron trading Pvt Ltd": "GLOBAL IRONTRADING PVT. LTD.",
    "Lalsar steel": "Lalsar Steels",
    "kedia pipes pvt Ltd": "KEDIA PIPES PVT LTD",
    "Dinesh Steel pvt ltd": "DINESH STEEL PRIVATE LIMITED",
    "Dinesh Steel pvt ltd ": "DINESH STEEL PRIVATE LIMITED",
    # Dec inconsistencies
    "GLOBAL STEEL": "Global Steel",
    "KEDIA PIPES PRIVATE LMITED": "KEDIA PIPES PVT LTD",
    # Jan inconsistencies
    "SSTL (INDIA) PRIVATE LIMITED": "SSTL",
    # Feb inconsistencies
    "M/s REAL DREAMS ENTERPRISES": "Real dreams enterprises",
    "OPTRIX Infra LLP": "OPTRIX INFRA LLP",
    "SILIGURI BUILDERS STORES": "Siliguri Builders Stores",
}

# ── First-tab duplicate consolidation: duplicates -> canonical name ──
FIRST_TAB_DUPES = {
    "NERIUM MULTICOM LLP": "NERIUM MULTICOMS LLP",
    "NERIUM MULTICOM": "NERIUM MULTICOMS LLP",
    "MAHADEVI METALS PVT. LTD.": "MAHADEVI METAL PVT LTD",
    "MEWARAM MADHUSUDAN PRASAD & SONS": "MEWARAM MADHUSUDAN PRASAD AND SONS",
    "LALSAR STEEL": "Lalsar Steels",
    "Maheshwari Traders ": "Maheshwari Traders",
}

# Also map monthly tab names to canonical (in case monthly tabs use the duplicate form)
NAME_MAP.update({
    "NERIUM MULTICOM LLP": "NERIUM MULTICOMS LLP",
    "NERIUM MULTICOM": "NERIUM MULTICOMS LLP",
})

NEW_DISTRIBUTORS = ["GOLDWIN ISPAT PRIVATE LIMITED", "Tanvi Sales", "SHALIKA ENTERPRISES PRIVATE LIMITED"]

# ── Load workbook ──
wb = load_workbook(INPUT_FILE)
ws_main = wb["Meets Apr to Oct"]

# Read master data into a list of lists for easier manipulation
master_data = []
for row in ws_main.iter_rows(min_row=1, values_only=True):
    master_data.append(list(row))

print("=" * 70)
print("STEP 1: CONSOLIDATING FIRST-TAB DUPLICATES")
print("=" * 70)

# Consolidate duplicates in first tab (rows index 2+ are data, 0=totals, 1=headers)
rows_to_remove = []
for idx in range(2, len(master_data)):
    name = master_data[idx][0]
    if name is None:
        continue
    name_str = str(name)
    if name_str in FIRST_TAB_DUPES:
        canonical = FIRST_TAB_DUPES[name_str]
        # Find the canonical row
        for cidx in range(2, len(master_data)):
            if cidx in rows_to_remove:
                continue
            cname = str(master_data[cidx][0] or "").strip()
            if cname == canonical:
                # Sum Apr-Oct values (columns 1-7)
                for col in range(1, 8):
                    val1 = master_data[cidx][col] if master_data[cidx][col] is not None else 0
                    val2 = master_data[idx][col] if master_data[idx][col] is not None else 0
                    master_data[cidx][col] = int(val1) + int(val2)
                print(f"  Merged '{name_str}' into '{canonical}' (summed Apr-Oct)")
                rows_to_remove.append(idx)
                break

# Remove duplicate rows (in reverse order to keep indices valid)
for idx in sorted(rows_to_remove, reverse=True):
    master_data.pop(idx)

print(f"  Removed {len(rows_to_remove)} duplicate rows. Now {len(master_data) - 2} distributors.")

# Build name->row_index lookup for the master tab
name_to_row = {}
for idx in range(2, len(master_data)):
    name = str(master_data[idx][0] or "").strip()
    name_to_row[name] = idx

# ── Step 2: Process monthly tabs ──
print("\n" + "=" * 70)
print("STEP 2: MAPPING MONTHLY TAB DATA")
print("=" * 70)

month_info = {"Nov": 8, "Dec": 9, "Jan": 10, "Feb": 11}  # 0-indexed column
mapped_names_log = []

for sheet_name, col_idx in month_info.items():
    ws_month = wb[sheet_name]
    print(f"\n--- {sheet_name} ---")

    # Build consolidated meets dict
    meets_dict = {}
    for row in ws_month.iter_rows(min_row=2, values_only=True):  # skip header
        raw_name = row[0]
        meets_val = row[1]
        if raw_name is None or str(raw_name).strip() == "" or str(raw_name).strip() == "0":
            continue

        raw_name_str = str(raw_name)
        meets_val = int(meets_val) if meets_val is not None else 0

        # Apply name mapping
        if raw_name_str in NAME_MAP:
            standard_name = NAME_MAP[raw_name_str]
            mapped_names_log.append((sheet_name, raw_name_str, standard_name))
            print(f"  MAPPED: '{raw_name_str}' -> '{standard_name}'")
        elif raw_name_str.strip() in NAME_MAP:
            standard_name = NAME_MAP[raw_name_str.strip()]
            mapped_names_log.append((sheet_name, raw_name_str, standard_name))
            print(f"  MAPPED: '{raw_name_str}' -> '{standard_name}'")
        else:
            standard_name = raw_name_str.strip()

        # Check first-tab duplicate aliases
        if standard_name in FIRST_TAB_DUPES:
            canonical = FIRST_TAB_DUPES[standard_name]
            print(f"  ALIAS: '{standard_name}' -> '{canonical}'")
            standard_name = canonical

        meets_dict[standard_name] = meets_dict.get(standard_name, 0) + meets_val

    # Show consolidated values
    total = sum(meets_dict.values())
    print(f"\n  Consolidated {sheet_name} data ({len(meets_dict)} distributors, total={total}):")
    for name, val in sorted(meets_dict.items()):
        print(f"    {name}: {val}")

    # Fill into master data
    filled = set()
    for name, val in meets_dict.items():
        if name in name_to_row:
            row_idx = name_to_row[name]
            # Ensure row has enough columns
            while len(master_data[row_idx]) <= col_idx:
                master_data[row_idx].append(None)
            master_data[row_idx][col_idx] = val
            filled.add(name)

    unfilled = set(meets_dict.keys()) - filled
    for name in unfilled:
        if name in NEW_DISTRIBUTORS:
            continue
        print(f"  WARNING: UNFILLED '{name}' not found in master tab!")

# ── Step 3: Add new distributors ──
print("\n" + "=" * 70)
print("STEP 3: ADDING NEW DISTRIBUTORS")
print("=" * 70)

# Collect new distributor data from monthly tabs
new_dist_data = {name: {"Nov": 0, "Dec": 0, "Jan": 0, "Feb": 0} for name in NEW_DISTRIBUTORS}

for sheet_name in ["Nov", "Dec", "Jan", "Feb"]:
    ws_month = wb[sheet_name]
    for row in ws_month.iter_rows(min_row=2, values_only=True):
        raw_name = row[0]
        meets_val = row[1]
        if raw_name is None:
            continue
        raw_name_str = str(raw_name).strip()
        mapped = NAME_MAP.get(raw_name_str, NAME_MAP.get(str(raw_name), raw_name_str))
        if mapped in NEW_DISTRIBUTORS:
            new_dist_data[mapped][sheet_name] += int(meets_val) if meets_val is not None else 0

for name in NEW_DISTRIBUTORS:
    new_row = [name] + [0]*7 + [
        new_dist_data[name]["Nov"],
        new_dist_data[name]["Dec"],
        new_dist_data[name]["Jan"],
        new_dist_data[name]["Feb"]
    ]
    master_data.append(new_row)
    print(f"  Added: {name} -> Nov:{new_dist_data[name]['Nov']}, Dec:{new_dist_data[name]['Dec']}, Jan:{new_dist_data[name]['Jan']}, Feb:{new_dist_data[name]['Feb']}")

# ── Step 4: Fill NaN with 0 for Nov-Feb, update totals ──
for idx in range(2, len(master_data)):
    while len(master_data[idx]) < 12:
        master_data[idx].append(None)
    for col in [8, 9, 10, 11]:
        if master_data[idx][col] is None:
            master_data[idx][col] = 0

# Ensure totals row has enough columns
while len(master_data[0]) < 12:
    master_data[0].append(None)
while len(master_data[1]) < 12:
    master_data[1].append(None)

# Update totals in row 0
for col in [8, 9, 10, 11]:
    total = sum(int(master_data[idx][col]) for idx in range(2, len(master_data)))
    master_data[0][col] = total

print(f"\n  Totals -> Nov:{master_data[0][8]}, Dec:{master_data[0][9]}, Jan:{master_data[0][10]}, Feb:{master_data[0][11]}")

# ── Step 5: Write output Excel ──
from openpyxl import Workbook

wb_out = Workbook()
ws_out = wb_out.active
ws_out.title = "Meets Apr to Oct"

yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
green_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")

# Build mapped names per month for highlighting
mapped_per_month = {}
for sheet, raw, standard in mapped_names_log:
    mapped_per_month.setdefault(sheet, set()).add(standard)

month_to_col_1based = {"Nov": 9, "Dec": 10, "Jan": 11, "Feb": 12}

for r_idx, row_data in enumerate(master_data):
    for c_idx, val in enumerate(row_data):
        cell = ws_out.cell(row=r_idx + 1, column=c_idx + 1, value=val)

    # Apply highlighting for data rows (row index 2+, i.e. Excel row 3+)
    if r_idx >= 2:
        dist_name = str(row_data[0] or "").strip()

        # Yellow for mapped names
        for month, col in month_to_col_1based.items():
            if dist_name in mapped_per_month.get(month, set()):
                ws_out.cell(row=r_idx + 1, column=col).fill = yellow_fill

        # Green for new distributors
        if dist_name in NEW_DISTRIBUTORS:
            for c in range(1, 13):
                ws_out.cell(row=r_idx + 1, column=c).fill = green_fill

wb_out.save(OUTPUT_FILE)
print(f"\nSaved to: {OUTPUT_FILE}")
print("  Yellow highlights = mapped from inconsistent name")
print("  Green highlights = new distributor added")

# ── Final Summary ──
print("\n" + "=" * 70)
print("DISCREPANCY SUMMARY")
print("=" * 70)
print(f"\nFirst-tab duplicates consolidated: {len(rows_to_remove)}")
print(f"Name mappings applied across monthly tabs: {len(mapped_names_log)}")
print(f"New distributors added: {len(NEW_DISTRIBUTORS)}")
print(f"\nFinal distributor count: {len(master_data) - 2}")
print(f"Monthly totals: Nov={master_data[0][8]}, Dec={master_data[0][9]}, Jan={master_data[0][10]}, Feb={master_data[0][11]}")
