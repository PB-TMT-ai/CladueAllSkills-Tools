"""
Parser for PB SOP Links Confluence.xlsx.
Reads SOP name → Confluence URL mappings and provides fuzzy matching.
"""

import re
import openpyxl


def parse_confluence_links(filepath: str) -> dict:
    """Read the Confluence links Excel and return {sop_name: url} mapping."""
    wb = openpyxl.load_workbook(filepath)
    ws = wb[wb.sheetnames[0]]
    links = {}
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=False):
        name_cell = row[0]
        url_cell = row[1]
        if name_cell.value and url_cell.value:
            name = str(name_cell.value).strip()
            # Prefer the hyperlink target if available, else use cell value
            if url_cell.hyperlink and url_cell.hyperlink.target:
                url = url_cell.hyperlink.target
            else:
                url = str(url_cell.value).strip()
            links[name] = url
    wb.close()
    return links


def _normalize(text: str) -> str:
    """Normalize a string for fuzzy matching: lowercase, strip numbers/special chars."""
    text = text.lower()
    # Decode common HTML entities before stripping
    text = text.replace('&amp;', '&').replace('&lt;', '<').replace('&gt;', '>')
    text = text.replace('&#39;', "'").replace('&quot;', '"')
    text = re.sub(r'^[\d\-\+\.\s]+', '', text)  # strip leading numbers/dashes
    text = re.sub(r'\(.*?\)', '', text)  # strip parenthetical notes
    text = re.sub(r'[^a-z0-9\s]', '', text)  # keep only alphanumeric + spaces
    text = re.sub(r'\s+', ' ', text).strip()
    return text


# Known fallback mappings for names that won't match automatically
_FALLBACK_MAP = {
    "nfluencer data management sop  private brands": "Influencer data management SOP - Private Brands",
}


def match_confluence_link(doc_title: str, links: dict) -> str | None:
    """Find the best Confluence URL match for a document title.

    Uses normalized substring matching with fallback for known mismatches.
    Returns the URL or None if no match found.
    """
    norm_title = _normalize(doc_title)

    # Direct normalized match
    for link_name, url in links.items():
        norm_link = _normalize(link_name)
        if norm_title == norm_link:
            return url

    # Substring match (title contains link name or vice versa)
    for link_name, url in links.items():
        norm_link = _normalize(link_name)
        if norm_link in norm_title or norm_title in norm_link:
            return url

    # Keyword overlap: require at least 3 matching words AND
    # the overlap must cover at least 60% of the shorter text's words
    # to prevent false matches on common terms like "pb retailer influencer"
    title_words = set(norm_title.split()) - {'the', 'and', 'or', 'of', 'in', 'on', 'to', 'a', 'for'}
    best_match = None
    best_ratio = 0
    for link_name, url in links.items():
        norm_link = _normalize(link_name)
        link_words = set(norm_link.split()) - {'the', 'and', 'or', 'of', 'in', 'on', 'to', 'a', 'for'}
        overlap = title_words & link_words
        if len(overlap) >= 3:
            shorter_len = min(len(title_words), len(link_words))
            ratio = len(overlap) / shorter_len if shorter_len else 0
            if ratio >= 0.6 and ratio > best_ratio:
                best_ratio = ratio
                best_match = url
    if best_match:
        return best_match

    # Fallback for known mismatches
    for fallback_key, canonical_name in _FALLBACK_MAP.items():
        if fallback_key in norm_title:
            norm_canonical = _normalize(canonical_name)
            for link_name, url in links.items():
                if _normalize(link_name) == norm_canonical or fallback_key in _normalize(link_name):
                    return url

    return None
