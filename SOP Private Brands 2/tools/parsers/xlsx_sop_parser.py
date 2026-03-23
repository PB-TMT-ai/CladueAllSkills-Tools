"""
Parser for .xlsx SOP source documents (e.g., SOP Marketing Activities.xlsx).
Reads columnar activity data and returns a structured SOPDocument.
"""

import os
import sys
import openpyxl

sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..'))
from models.sop_data import SOPDocument


def parse_xlsx_sop(filepath: str) -> SOPDocument:
    """Parse an Excel SOP source document and return a SOPDocument."""
    wb = openpyxl.load_workbook(filepath)
    ws = wb[wb.sheetnames[0]]
    filename = os.path.basename(filepath)

    # Find the header row (first row with 'Sr No' or 'Activity' in it)
    header_row_idx = None
    headers = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=min(5, ws.max_row), values_only=True), 1):
        row_text = ' '.join(str(c or '') for c in row).lower()
        if 'sr no' in row_text or 'activity' in row_text:
            header_row_idx = row_idx
            headers = [str(c or '').strip() for c in row]
            break

    if header_row_idx is None:
        wb.close()
        return SOPDocument(filename=filename, title=filename.replace('.xlsx', ''))

    # Read data rows
    tables_data = [headers]
    raw_lines = []

    for row in ws.iter_rows(min_row=header_row_idx + 1, max_row=ws.max_row, values_only=True):
        cells = [str(c or '').strip() for c in row]
        if any(c for c in cells):
            tables_data.append(cells)
            raw_lines.append(' | '.join(c for c in cells if c))

    # Extract title from filename
    title = filename.replace('.xlsx', '').replace('_', ' ')

    wb.close()

    return SOPDocument(
        filename=filename,
        title=title,
        doc_type="",
        purpose="Marketing activities SOP covering requirement gathering through invoice processing",
        stakeholders={"Marketing": "", "Business Team": "", "Finance": "", "Vendor": ""},
        steps=[],
        tables=[tables_data],
        headings=[(1, title)],
        raw_text='\n'.join(raw_lines),
    )
