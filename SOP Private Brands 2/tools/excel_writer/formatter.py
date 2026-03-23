"""
Excel formatting constants matching the Construct.xlsx reference template.
Colors, fonts, alignments, borders, and column widths.
"""

from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# --- Fills ---
HEADER_FILL = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
SECTION_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
JOURNEY_FILL = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")

# --- Fonts ---
HEADER_FONT = Font(name="Calibri", size=10, bold=True, color="000000")
SECTION_FONT = Font(name="Calibri", size=10, bold=True, color="000000")
DATA_FONT = Font(name="Calibri", size=10, bold=False, color="000000")
LINK_FONT = Font(name="Calibri", size=10, bold=False, color="0563C1", underline="single")

# --- Alignment ---
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
SECTION_ALIGN = Alignment(horizontal="left", vertical="center", wrap_text=False)
JOURNEY_ALIGN = Alignment(horizontal="center", vertical="center", text_rotation=90, wrap_text=False)
DATA_ALIGN = Alignment(horizontal="left", vertical="top", wrap_text=True)
DATA_CENTER_ALIGN = Alignment(horizontal="center", vertical="top", wrap_text=True)

# --- Borders ---
THIN_BORDER = Border(
    top=Side(style="thin"),
    bottom=Side(style="thin"),
    left=Side(style="thin"),
    right=Side(style="thin"),
)

# --- Column widths (approximate match to Construct.xlsx) ---
COLUMN_WIDTHS = {
    'A': 10,      # Journey
    'B': 6,       # Sr. No.
    'C': 28,      # Activity
    'D': 55,      # Description
    'E': 21,      # Owner
    'F': 18,      # Interface
    'G': 23,      # Sign off
    'H': 70,      # Steps
    'I': 14,      # Flow Type
    'J': 50,      # SOP Link
    'K': 50,      # Remarks
    'L': 30,      # Assignee notes
}

# --- Column headers ---
COLUMN_HEADERS = [
    "Journey",
    "Sr. No.",
    "Activity",
    "Description",
    "Owner",
    "Interface",
    "Sign off",
    "Steps",
    "Flow Type",
    "SOP Link",
    "Remarks",
    "",  # Blank column L (assignee notes in reference)
]

# --- Row height ---
HEADER_ROW_HEIGHT = 30
DATA_ROW_HEIGHT = 15
SECTION_ROW_HEIGHT = 20
