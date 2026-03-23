"""
Excel writer: generates the master SOP Excel file with proper formatting,
merged cells, and hyperlinks matching the Construct.xlsx reference template.
Sheet tab name: "Private Brands".
"""

import os
import sys
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..'))
from excel_writer.formatter import (
    HEADER_FILL, SECTION_FILL, JOURNEY_FILL,
    HEADER_FONT, SECTION_FONT, DATA_FONT, LINK_FONT,
    HEADER_ALIGN, SECTION_ALIGN, JOURNEY_ALIGN, DATA_ALIGN, DATA_CENTER_ALIGN,
    THIN_BORDER, COLUMN_WIDTHS, COLUMN_HEADERS,
    HEADER_ROW_HEIGHT, DATA_ROW_HEIGHT, SECTION_ROW_HEIGHT,
)
from models.sop_data import Journey, Section, Activity


def write_master_excel(journeys: list, output_path: str, sheet_name: str = "Private Brands"):
    """Generate the master SOP Excel file.

    Args:
        journeys: List of Journey objects (Pre-Order, Order, Post Order)
        output_path: Full path for the output .xlsx file
        sheet_name: Name for the worksheet tab (default: "Private Brands")
    """
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    # Write header row
    _write_header(ws)

    # Track current row (start after header)
    current_row = 2
    global_sr_no = 1

    for journey in journeys:
        journey_start_row = current_row

        for section in journey.sections:
            if not section.activities:
                continue

            # Write section header row
            current_row = _write_section_header(ws, current_row, section.label)

            for activity in section.activities:
                # Calculate how many rows this activity needs
                num_rows = _calc_activity_rows(activity)

                # Write activity data
                _write_activity(ws, current_row, global_sr_no, activity, num_rows)

                # Merge cells for multi-row activities
                if num_rows > 1:
                    _merge_activity_cells(ws, current_row, num_rows)

                current_row += num_rows
                global_sr_no += 1

        # Merge Journey column (A) across all rows for this journey
        journey_end_row = current_row - 1
        if journey_end_row >= journey_start_row:
            if journey_end_row > journey_start_row:
                ws.merge_cells(f'A{journey_start_row}:A{journey_end_row}')
            cell = ws.cell(row=journey_start_row, column=1)
            cell.value = journey.name
            cell.font = HEADER_FONT
            cell.fill = JOURNEY_FILL
            cell.alignment = JOURNEY_ALIGN
            cell.border = THIN_BORDER

    # Apply column widths
    for col_letter, width in COLUMN_WIDTHS.items():
        ws.column_dimensions[col_letter].width = width

    # Freeze top row
    ws.freeze_panes = 'A2'

    # Save
    os.makedirs(os.path.dirname(output_path) or '.', exist_ok=True)
    wb.save(output_path)
    wb.close()

    return output_path


def _write_header(ws):
    """Write the header row (row 1)."""
    for col_idx, header_text in enumerate(COLUMN_HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header_text)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER

    ws.row_dimensions[1].height = HEADER_ROW_HEIGHT


def _write_section_header(ws, row: int, section_label: str) -> int:
    """Write a section header row (merged across B:D, green fill).

    Returns the next available row.
    """
    # Merge B:D for section label
    ws.merge_cells(f'B{row}:D{row}')
    cell = ws.cell(row=row, column=2, value=section_label)
    cell.font = SECTION_FONT
    cell.fill = SECTION_FILL
    cell.alignment = SECTION_ALIGN
    cell.border = THIN_BORDER

    # Apply fill and border to all columns in the section row
    for col in range(1, 13):
        c = ws.cell(row=row, column=col)
        c.fill = SECTION_FILL
        c.border = THIN_BORDER

    ws.row_dimensions[row].height = SECTION_ROW_HEIGHT
    return row + 1


def _calc_activity_rows(activity: Activity) -> int:
    """Calculate how many Excel rows an activity needs."""
    rows_needed = max(
        1,
        len(activity.steps),
        len(activity.description_details) + 1,
        len(activity.remarks_details) + 1,
    )
    return rows_needed


def _write_activity(ws, start_row: int, sr_no: int, activity: Activity, num_rows: int):
    """Write a single activity across one or more rows."""

    # --- Row 1: Main data in all columns ---
    # Column B: Sr. No.
    ws.cell(row=start_row, column=2, value=sr_no).font = DATA_FONT
    ws.cell(row=start_row, column=2).alignment = DATA_CENTER_ALIGN

    # Column C: Activity name
    ws.cell(row=start_row, column=3, value=activity.activity_name).font = DATA_FONT
    ws.cell(row=start_row, column=3).alignment = DATA_ALIGN

    # Column D: Description (main line)
    ws.cell(row=start_row, column=4, value=activity.description).font = DATA_FONT
    ws.cell(row=start_row, column=4).alignment = DATA_ALIGN

    # Column E: Owner
    ws.cell(row=start_row, column=5, value=activity.owner).font = DATA_FONT
    ws.cell(row=start_row, column=5).alignment = DATA_ALIGN

    # Column F: Interface
    ws.cell(row=start_row, column=6, value=activity.interface).font = DATA_FONT
    ws.cell(row=start_row, column=6).alignment = DATA_ALIGN

    # Column G: Sign off
    ws.cell(row=start_row, column=7, value=activity.sign_off).font = DATA_FONT
    ws.cell(row=start_row, column=7).alignment = DATA_ALIGN

    # Column I: Flow Type
    ws.cell(row=start_row, column=9, value=activity.flow_type).font = DATA_FONT
    ws.cell(row=start_row, column=9).alignment = DATA_CENTER_ALIGN

    # Column J: SOP Link (as hyperlink if URL)
    if activity.sop_link and activity.sop_link.startswith('http'):
        cell_j = ws.cell(row=start_row, column=10, value=activity.sop_link)
        cell_j.hyperlink = activity.sop_link
        cell_j.font = LINK_FONT
    else:
        ws.cell(row=start_row, column=10, value=activity.sop_link or "").font = DATA_FONT
    ws.cell(row=start_row, column=10).alignment = DATA_ALIGN

    # Column K: Remarks
    ws.cell(row=start_row, column=11, value=activity.remarks or "").font = DATA_FONT
    ws.cell(row=start_row, column=11).alignment = DATA_ALIGN

    # Column L: Assignee notes
    ws.cell(row=start_row, column=12, value=activity.assignee_notes or "").font = DATA_FONT
    ws.cell(row=start_row, column=12).alignment = DATA_ALIGN

    # --- Column H: Steps (one per row) ---
    for i, step in enumerate(activity.steps):
        row_idx = start_row + i
        ws.cell(row=row_idx, column=8, value=step).font = DATA_FONT
        ws.cell(row=row_idx, column=8).alignment = DATA_ALIGN

    # --- Column D: Description details (sub-items on rows 2+) ---
    for i, detail in enumerate(activity.description_details):
        row_idx = start_row + 1 + i
        ws.cell(row=row_idx, column=4, value=detail).font = DATA_FONT
        ws.cell(row=row_idx, column=4).alignment = DATA_ALIGN

    # --- Column K: Remarks details (sub-items on rows 2+) ---
    for i, remark in enumerate(activity.remarks_details):
        row_idx = start_row + 1 + i
        ws.cell(row=row_idx, column=11, value=remark).font = DATA_FONT
        ws.cell(row=row_idx, column=11).alignment = DATA_ALIGN

    # --- Apply borders to all cells ---
    for r in range(start_row, start_row + num_rows):
        for c in range(1, 13):
            ws.cell(row=r, column=c).border = THIN_BORDER
        ws.row_dimensions[r].height = DATA_ROW_HEIGHT


def _merge_activity_cells(ws, start_row: int, num_rows: int):
    """Merge vertically across all rows for columns that span the activity.

    Merged: B (Sr. No.), C (Activity), E (Owner), F (Interface),
            G (Sign off), I (Flow Type), J (SOP Link), K (Remarks), L (notes)
    NOT merged: A (Journey - handled separately), D (Description), H (Steps)
    """
    merge_columns = [2, 3, 5, 6, 7, 9, 10, 11, 12]
    end_row = start_row + num_rows - 1

    for col in merge_columns:
        col_letter = get_column_letter(col)
        ws.merge_cells(f'{col_letter}{start_row}:{col_letter}{end_row}')
