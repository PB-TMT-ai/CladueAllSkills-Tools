#!/usr/bin/env python3
"""
Fix duplicate Description/Steps in JSW_ONE_PB_SOPs_Master.xlsx.

97 activities have identical content in Description (col D) and Steps (col H).
This script splits them:
  - Case A (multi-line/multi-sentence): D = first sentence, H = keep full text
  - Case B (short single-sentence): D = keep as-is, H = clear
"""

import openpyxl
import sys
import io

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

EXCEL_PATH = "output/JSW_ONE_PB_SOPs_Master.xlsx"


def derive_short_description(full_text):
    """Extract first sentence/line as concise description."""
    d = str(full_text).strip()
    first_line = d.split("\n")[0].strip()
    # If first line starts with bullet, keep just the first line
    if first_line.startswith("- ") or first_line.startswith("* "):
        result = first_line
    elif ". " in first_line:
        # Take up to first sentence boundary
        result = first_line.split(". ")[0] + "."
    else:
        result = first_line
    # Cap at 200 chars
    if len(result) > 200:
        result = result[:197] + "..."
    return result


def fix_duplicate_desc_steps(ws):
    """Fix all activities where Description == Steps."""
    case_a = 0  # multi-line: shorten D, keep H
    case_b = 0  # short: keep D, clear H

    for row_idx in range(2, ws.max_row + 1):
        sr = ws.cell(row_idx, 2).value
        if not sr or not str(sr).isdigit():
            continue

        desc = ws.cell(row_idx, 4).value
        steps = ws.cell(row_idx, 8).value

        if not desc or not steps:
            continue

        desc_str = str(desc).strip()
        steps_str = str(steps).strip()

        if desc_str != steps_str:
            continue

        # This activity has D == H, fix it
        short_desc = derive_short_description(desc_str)

        if short_desc != desc_str:
            # Case A: content can be split
            ws.cell(row_idx, 4).value = short_desc
            # H stays as-is (full text)
            case_a += 1
        else:
            # Case B: short content, no meaningful split
            # Keep D, clear H
            ws.cell(row_idx, 8).value = None
            case_b += 1

    return case_a, case_b


def main():
    # Load workbook preserving formatting (no backup per user request)
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb["Private Brands"]

    print("Fixing duplicate Description/Steps...")
    case_a, case_b = fix_duplicate_desc_steps(ws)

    print(f"  Case A (D shortened, H kept): {case_a}")
    print(f"  Case B (D kept, H cleared): {case_b}")
    print(f"  Total fixed: {case_a + case_b}")

    wb.save(EXCEL_PATH)
    print(f"\nSaved: {EXCEL_PATH}")
    wb.close()


if __name__ == "__main__":
    main()
