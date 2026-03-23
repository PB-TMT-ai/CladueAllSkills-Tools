"""
One-time script to insert Beat Planning activity into the existing
Private Brands master Excel at Pre - Order > B. SALES PLANNING.

Usage:
    python tools/add_beat_planning.py

Strategy: Since openpyxl's insert_rows() corrupts merged cell ranges,
this script rebuilds the sheet by copying all existing data to a new sheet
with the Beat Planning rows inserted at the correct position (row 30).
"""

import os
import sys
from copy import copy
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Formatting constants (matching formatter.py)
SECTION_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
SECTION_FONT = Font(name="Calibri", size=10, bold=True, color="000000")
SECTION_ALIGN = Alignment(horizontal="left", vertical="center", wrap_text=False)

DATA_FONT = Font(name="Calibri", size=10, bold=False, color="000000")
DATA_ALIGN = Alignment(horizontal="left", vertical="top", wrap_text=True)
DATA_CENTER_ALIGN = Alignment(horizontal="center", vertical="top", wrap_text=True)
LINK_FONT = Font(name="Calibri", size=10, bold=False, color="0563C1", underline="single")

THIN_BORDER = Border(
    top=Side(style="thin"),
    bottom=Side(style="thin"),
    left=Side(style="thin"),
    right=Side(style="thin"),
)

DATA_ROW_HEIGHT = 15
SECTION_ROW_HEIGHT = 20

INSERT_AT = 30        # Insert new rows before this row
TOTAL_INSERT = 7      # 1 section header + 6 step rows
MAX_COL = 12          # Columns A-L

# Beat Planning data from Construct.xlsx
BEAT_PLANNING = {
    "activity_name": "Beat Planning",
    "description": "Monthly visit planning",
    "owner": "Sales and Business analyst",
    "interface": "Excel and salesforce",
    "sign_off": "Sales",
    "flow_type": "Negative",
    "sop_link": (
        "https://jsw-my.sharepoint.com/:x:/g/personal/"
        "sachin_kumar6_jsw_in/IQDYXpTBtgFyS766CXR0AAYEATOrksf0l-d837obC2tuTu0?e=ieExNt"
    ),
    "remarks": (
        "Like MS Teams, calender blocking should be available to sales owners in SF. "
        "Regular notifications to be enabled for the cadence."
    ),
    "steps": [
        "1. Business analyst shares the list of customers to all the sales members in Excel in the last week.",
        "2. Sales associates are required to plan the daily visits against the accounts",
        "3. Once the team has updated the visits, BA will upload the beat plan from the backend.",
        "4. This will be then reflected in SF calendar.",
        "5. Associates are allowed to mark visits only against the accounts where beat is planned.",
        "6. If anyone is visiting new account, one need to create a beat on SF.",
    ],
}


def _copy_cell(src_cell, dst_cell):
    """Copy value and formatting from one cell to another."""
    dst_cell.value = src_cell.value
    if src_cell.has_style:
        dst_cell.font = copy(src_cell.font)
        dst_cell.fill = copy(src_cell.fill)
        dst_cell.alignment = copy(src_cell.alignment)
        dst_cell.border = copy(src_cell.border)
        dst_cell.number_format = src_cell.number_format
    # Copy hyperlink if present
    if src_cell.hyperlink:
        dst_cell.hyperlink = copy(src_cell.hyperlink)


def main():
    project_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    excel_path = os.path.join(project_root, "output", "JSW_ONE_PB_SOPs_Master.xlsx")

    if not os.path.exists(excel_path):
        print(f"ERROR: Excel file not found at {excel_path}")
        sys.exit(1)

    print(f"Loading: {excel_path}")
    wb_src = load_workbook(excel_path)
    ws_src = wb_src["Private Brands"]
    old_max_row = ws_src.max_row
    print(f"  Source rows: {old_max_row}")

    # --- Step 1: Read all existing merge ranges ---
    print("  Reading merge ranges...")
    old_merges = []
    for mr in ws_src.merged_cells.ranges:
        old_merges.append((mr.min_row, mr.min_col, mr.max_row, mr.max_col))

    # --- Step 2: Create new workbook with same structure ---
    print("  Creating new workbook...")
    wb_dst = Workbook()
    ws_dst = wb_dst.active
    ws_dst.title = "Private Brands"

    # Copy column widths
    for col_letter in ['A','B','C','D','E','F','G','H','I','J','K','L']:
        if col_letter in ws_src.column_dimensions:
            ws_dst.column_dimensions[col_letter].width = ws_src.column_dimensions[col_letter].width

    # Copy freeze panes
    ws_dst.freeze_panes = ws_src.freeze_panes

    # --- Step 3: Copy rows 1 to INSERT_AT-1 (rows 1-29) as-is ---
    print(f"  Copying rows 1-{INSERT_AT - 1} (before insertion point)...")
    for row in range(1, INSERT_AT):
        for col in range(1, MAX_COL + 1):
            _copy_cell(ws_src.cell(row=row, column=col), ws_dst.cell(row=row, column=col))
        # Copy row height
        if row in ws_src.row_dimensions:
            ws_dst.row_dimensions[row].height = ws_src.row_dimensions[row].height

    # --- Step 4: Write B. SALES PLANNING section header (row 30) ---
    print(f"  Writing section header at row {INSERT_AT}: B. SALES PLANNING")
    cell = ws_dst.cell(row=INSERT_AT, column=2, value="B. SALES PLANNING")
    cell.font = SECTION_FONT
    cell.fill = SECTION_FILL
    cell.alignment = SECTION_ALIGN
    cell.border = THIN_BORDER
    for col in range(1, MAX_COL + 1):
        c = ws_dst.cell(row=INSERT_AT, column=col)
        c.fill = SECTION_FILL
        c.border = THIN_BORDER
    ws_dst.row_dimensions[INSERT_AT].height = SECTION_ROW_HEIGHT

    # --- Step 5: Write Beat Planning activity (rows 31-36) ---
    act_start = INSERT_AT + 1  # 31
    act_end = act_start + len(BEAT_PLANNING["steps"]) - 1  # 36
    bp = BEAT_PLANNING
    print(f"  Writing Beat Planning activity (rows {act_start}-{act_end})")

    # Row 31: all columns
    ws_dst.cell(row=act_start, column=2, value=4).font = DATA_FONT
    ws_dst.cell(row=act_start, column=2).alignment = DATA_CENTER_ALIGN

    ws_dst.cell(row=act_start, column=3, value=bp["activity_name"]).font = DATA_FONT
    ws_dst.cell(row=act_start, column=3).alignment = DATA_ALIGN

    ws_dst.cell(row=act_start, column=4, value=bp["description"]).font = DATA_FONT
    ws_dst.cell(row=act_start, column=4).alignment = DATA_ALIGN

    ws_dst.cell(row=act_start, column=5, value=bp["owner"]).font = DATA_FONT
    ws_dst.cell(row=act_start, column=5).alignment = DATA_ALIGN

    ws_dst.cell(row=act_start, column=6, value=bp["interface"]).font = DATA_FONT
    ws_dst.cell(row=act_start, column=6).alignment = DATA_ALIGN

    ws_dst.cell(row=act_start, column=7, value=bp["sign_off"]).font = DATA_FONT
    ws_dst.cell(row=act_start, column=7).alignment = DATA_ALIGN

    ws_dst.cell(row=act_start, column=9, value=bp["flow_type"]).font = DATA_FONT
    ws_dst.cell(row=act_start, column=9).alignment = DATA_CENTER_ALIGN

    cell_j = ws_dst.cell(row=act_start, column=10, value=bp["sop_link"])
    cell_j.hyperlink = bp["sop_link"]
    cell_j.font = LINK_FONT
    cell_j.alignment = DATA_ALIGN

    ws_dst.cell(row=act_start, column=11, value=bp["remarks"]).font = DATA_FONT
    ws_dst.cell(row=act_start, column=11).alignment = DATA_ALIGN

    # Steps (one per row in column H)
    for i, step in enumerate(bp["steps"]):
        r = act_start + i
        ws_dst.cell(row=r, column=8, value=step).font = DATA_FONT
        ws_dst.cell(row=r, column=8).alignment = DATA_ALIGN

    # Borders and row height for activity rows
    for r in range(act_start, act_end + 1):
        for c in range(1, MAX_COL + 1):
            ws_dst.cell(row=r, column=c).border = THIN_BORDER
        ws_dst.row_dimensions[r].height = DATA_ROW_HEIGHT

    # --- Step 6: Copy rows INSERT_AT to old_max_row → shifted by TOTAL_INSERT ---
    print(f"  Copying rows {INSERT_AT}-{old_max_row} -> {INSERT_AT + TOTAL_INSERT}-{old_max_row + TOTAL_INSERT}...")
    for row in range(INSERT_AT, old_max_row + 1):
        dst_row = row + TOTAL_INSERT
        for col in range(1, MAX_COL + 1):
            src_cell = ws_src.cell(row=row, column=col)
            dst_cell = ws_dst.cell(row=dst_row, column=col)
            _copy_cell(src_cell, dst_cell)
        # Copy row height
        if row in ws_src.row_dimensions:
            ws_dst.row_dimensions[dst_row].height = ws_src.row_dimensions[row].height

    # --- Step 7: Recreate merge ranges ---
    print("  Recreating merge ranges...")

    # Merges for rows before insertion point: copy as-is
    for min_r, min_c, max_r, max_c in old_merges:
        if max_r < INSERT_AT:
            # Entirely before insertion — keep as-is
            ws_dst.merge_cells(
                start_row=min_r, start_column=min_c,
                end_row=max_r, end_column=max_c
            )
        elif min_r >= INSERT_AT:
            # Entirely at or after insertion — shift by TOTAL_INSERT
            ws_dst.merge_cells(
                start_row=min_r + TOTAL_INSERT, start_column=min_c,
                end_row=max_r + TOTAL_INSERT, end_column=max_c
            )
        else:
            # Spans the insertion point — expand by TOTAL_INSERT
            ws_dst.merge_cells(
                start_row=min_r, start_column=min_c,
                end_row=max_r + TOTAL_INSERT, end_column=max_c
            )

    # Add new merges: section header B30:D30
    ws_dst.merge_cells("B30:D30")

    # Add new merges: activity columns B,C,E,F,G,I,J,K,L across rows 31-36
    merge_cols = [2, 3, 5, 6, 7, 9, 10, 11, 12]
    for col in merge_cols:
        cl = get_column_letter(col)
        ws_dst.merge_cells(f"{cl}{act_start}:{cl}{act_end}")

    # --- Step 8: Renumber serial numbers ---
    print("  Renumbering serial numbers...")
    # All activities from old row 30 onward (now at row 37+) have Sr. No. >= 4
    # Increment by 1 since Beat Planning takes Sr. No. 4
    for row in range(INSERT_AT + TOTAL_INSERT, ws_dst.max_row + 1):
        cell = ws_dst.cell(row=row, column=2)
        if cell.value is not None and isinstance(cell.value, (int, float)):
            old_val = int(cell.value)
            if old_val >= 4:
                cell.value = old_val + 1

    # --- Step 9: Save ---
    print(f"  Saving to: {excel_path}")
    wb_src.close()
    wb_dst.save(excel_path)
    wb_dst.close()

    new_max_row = old_max_row + TOTAL_INSERT
    print(f"\nDone! Beat Planning added successfully.")
    print(f"  Rows: {old_max_row} -> {new_max_row}")
    print(f"  Activities: 129 -> 130")
    print(f"  Location: Pre - Order > B. SALES PLANNING (rows {INSERT_AT}-{act_end})")


if __name__ == "__main__":
    main()
