"""
JSW Steel TMT Marketing Budget Model V2 - Management Review Edition
4-tab model: Executive Summary | Obj1 Helix | Obj2 JSW ONE | Budget Control & Review
"""

import os
from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Border, Side, Alignment, Protection, numbers
)
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

# === OUTPUT PATH ===
OUTPUT_DIR = os.path.dirname(__file__)
OUTPUT_PATH = os.path.join(OUTPUT_DIR, "JSW_TMT_Budget_V2.xlsx")

# === STYLE CONSTANTS ===
COLORS = {
    'header_dark': '1F4E79',
    'header_med': '2E75B6',
    'header_light': '9DC3E6',
    'input_yellow': 'FFF2CC',
    'calc_green': 'E2EFDA',
    'output_blue': 'D6E4F0',
    'total_gray': 'D9D9D9',
    'alert_red': 'FFC7CE',
    'ok_green': 'C6EFCE',
    'white': 'FFFFFF',
    'helix_blue': 'DAEEF3',
    'jsw_green': 'EBF1DE',
    'jsw_orange': 'FDE9D9',
    'kpi_bg': 'F2F2F2',
    'dark_text': '1F4E79',
}

THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
HEADER_FONT = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
HEADER_FONT_LG = Font(name='Calibri', bold=True, color='FFFFFF', size=14)
TITLE_FONT = Font(name='Calibri', bold=True, size=12, color='1F4E79')
BOLD_FONT = Font(name='Calibri', bold=True, size=11)
BOLD_FONT_SM = Font(name='Calibri', bold=True, size=10)
NORMAL_FONT = Font(name='Calibri', size=10)
SMALL_FONT = Font(name='Calibri', size=9, italic=True, color='808080')
KPI_FONT = Font(name='Calibri', bold=True, size=16, color='1F4E79')
KPI_LABEL_FONT = Font(name='Calibri', size=9, color='808080')

QUARTERS = ['Q1 (Apr-Jun)', 'Q2 (Jul-Sep)', 'Q3 (Oct-Dec)', 'Q4 (Jan-Mar)']
QUARTERS_SHORT = ['Q1', 'Q2', 'Q3', 'Q4']
TAB4 = "'Budget Control'!"


# === HELPER FUNCTIONS ===

def fill(color_key):
    return PatternFill(start_color=COLORS[color_key], end_color=COLORS[color_key], fill_type='solid')


def style_header_row(ws, row, max_col, color_key='header_dark', font=HEADER_FONT):
    for c in range(1, max_col + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = fill(color_key)
        cell.font = font
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)


def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def write_row(ws, row, values, start_col=1, font=NORMAL_FONT, color_key='white', num_fmt=None):
    for i, v in enumerate(values):
        cell = ws.cell(row=row, column=start_col + i, value=v)
        cell.font = font
        cell.fill = fill(color_key)
        cell.border = THIN_BORDER
        cell.alignment = Alignment(vertical='center', wrap_text=True)
        if num_fmt:
            cell.number_format = num_fmt


def write_section_header(ws, row, text, max_col, color_key='header_med'):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=max_col)
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = HEADER_FONT
    cell.fill = fill(color_key)
    cell.border = THIN_BORDER
    cell.alignment = Alignment(horizontal='left', vertical='center')
    for c in range(2, max_col + 1):
        ws.cell(row=row, column=c).fill = fill(color_key)
        ws.cell(row=row, column=c).border = THIN_BORDER


def style_cell(ws, row, col, color_key='white', font=NORMAL_FONT, num_fmt=None, locked=True):
    cell = ws.cell(row=row, column=col)
    cell.fill = fill(color_key)
    cell.font = font
    cell.border = THIN_BORDER
    if num_fmt:
        cell.number_format = num_fmt
    cell.protection = Protection(locked=locked)
    return cell


# === ATL & BTL ACTIVITY DATA ===

ATL_ACTIVITIES = [
    ('Impact Wall Painting', 15, 'Rs/sq ft', '20,000 sq ft per district'),
    ('FM Radio Campaign', 250000, 'Rs/campaign', 'Per district per quarter'),
    ('Hoardings (Flex/Billboard)', 50000, 'Rs/hoarding', '5 per priority district'),
    ('Digital Marketing Campaign', 500000, 'Rs/campaign', 'Social + YouTube + Display'),
    ('Local Events / Trade Fairs', 500000, 'Rs/event', 'Zone-level events'),
    ('Van Campaign (Mobile)', 7000, 'Rs/day', '2 months per state'),
    ('Syndicate Launch Events', 500000, 'Rs/event', 'Per state launch'),
    ('Transit Auto Branding', 700, 'Rs/auto', 'Auto rickshaw wraps'),
    ('Tea Stall Branding', 4000, 'Rs/stall', '25 stalls per district'),
    ('Paper Advertisement', 100000, 'Rs/insert', 'Regional dailies'),
]

BTL_ACTIVITIES = [
    ('NLB (Name/Lit Board)', 1200, 'Rs/board', '40/sqft x 30 sqft'),
    ('GSB (Glow Sign Board)', 6510, 'Rs/board', '180x30 + 65x18'),
    ('Counter Wall Painting', 10000, 'Rs/counter', 'Retailer counter branding'),
    ('Inshop Branding', 40000, 'Rs/counter', 'Premium in-store display'),
    ('Architect/Engineer Meet', 3500, 'Rs/person', '50 pax per meet'),
    ('Contractor Meet', 1000, 'Rs/person', '50 pax per meet'),
    ('Mason Meet', 300, 'Rs/person', '30 pax per meet'),
    ('Dealer Certificates', 400, 'Rs/certificate', 'Recognition boards'),
    ('Construction Site Gate Board', 1500, 'Rs/board', 'Per active site'),
    ('POP Materials Kit', 3000, 'Rs/dealer', 'Standees, banners, flyers'),
    ('Architect Gift/Engagement', 500, 'Rs/gift', 'Relationship building'),
    ('Distributor Meet (Grand)', 100000, 'Rs/person', 'Annual convention'),
    ('Retailer Loyalty Program', 2000, 'Rs/retailer/qtr', 'Volume-linked rewards'),
    ('Mason Training Camp', 5000, 'Rs/camp', '20 masons per camp'),
    ('Competitive Conversion Scheme', 10000, 'Rs/retailer', 'Switch-over incentive'),
]

# Helix quantities for each activity (annual)
HELIX_ATL_QTY = [
    ('=250*20000', 'sqft across 250 districts'),         # Wall Painting
    ('=ROUND(250*0.05,0)*4', '5% districts x 4 qtrs'),   # FM Radio
    ('=250*5', '5 per district'),                          # Hoardings
    (4, '4 campaigns/year'),                               # Digital
    (8, '8 zonal events'),                                 # Events
    ('=10*60', '10 vans x 60 days'),                       # Van
    (15, '15 state launches'),                             # Syndicate
    (500, '500 autos branded'),                            # Transit
    ('=ROUND(250*0.1,0)*25', '10% dist x 25 stalls'),    # Tea Stall
    (12, '12 insertions/year'),                            # Paper
]

HELIX_BTL_QTY = [
    ('=ROUND(2015*0.5,0)', '50% active retailers'),       # NLB
    ('=ROUND(2015*0.25,0)', '25% active retailers'),      # GSB
    ('=ROUND(2015*0.2,0)', '20% active retailers'),       # Counter
    ('=ROUND(2015*0.1,0)', '10% active retailers'),       # Inshop
    ('=ROUND(250*0.5,0)*4', '50% dist x 4 qtrs meets'),  # Architect Meet (meets, 50 pax each)
    ('=250*4', 'all dist x 4 qtrs meets'),                # Contractor Meet
    ('=250*4', 'all dist x 4 qtrs meets'),                # Mason Meet
    (2015, 'all active retailers'),                        # Dealer Certs
    ('=ROUND(250*0.3,0)', '30% districts'),               # Gate Board
    (42, 'all distributors'),                              # POP Kit
    ('=ROUND(250*0.5,0)*50', '50% dist x 50 architects'), # Architect Gift
    (42, 'all distributors'),                              # Distributor Meet
    ('=ROUND(1310*0.3,0)*4', '30% transacting x 4 qtrs'), # Retailer Loyalty
    ('=ROUND(250*0.2,0)', '20% districts'),               # Mason Training
    ('=ROUND(2015*0.05,0)', '5% retailers'),              # Competitive Conversion
]

# JSW ONE East (Sub-2A) quantities
EAST_ATL_QTY = [
    ('=120*20000', 'sqft across 120 East districts'),
    ('=ROUND(120*0.05,0)*4', '5% dist x 4 qtrs'),
    ('=120*3', '3 per district'),
    (4, '4 campaigns/year'),
    (4, '4 zonal events'),
    ('=4*60', '4 vans x 60 days (4 states)'),
    (4, '4 state launches'),
    (300, '300 autos'),
    ('=ROUND(120*0.1,0)*25', '10% dist x 25 stalls'),
    (8, '8 insertions/year'),
]

EAST_BTL_QTY = [
    ('=ROUND(954*0.5,0)', '50% of 954 East retailers'),
    ('=ROUND(954*0.25,0)', '25%'),
    ('=ROUND(954*0.2,0)', '20%'),
    ('=ROUND(954*0.1,0)', '10%'),
    ('=ROUND(120*0.5,0)*4', '50% dist x 4 qtrs'),
    ('=120*4', 'all dist x 4 qtrs'),
    ('=120*4', 'all dist x 4 qtrs'),
    (954, 'all East retailers'),
    ('=ROUND(120*0.3,0)', '30% districts'),
    (34, '60% of 56 distributors'),
    ('=ROUND(120*0.5,0)*50', '50% dist x 50 arch'),
    (34, '34 East distributors'),
    ('=ROUND(620*0.3,0)*4', '30% transacting x 4 qtrs'),
    ('=ROUND(120*0.2,0)', '20% districts'),
    ('=ROUND(954*0.05,0)', '5% retailers'),
]

# JSW ONE UP+HR (Sub-2B) quantities
UPHR_ATL_QTY = [
    ('=74*20000', 'sqft across 74 UP+HR districts'),
    ('=ROUND(74*0.05,0)*4', '5% dist x 4 qtrs'),
    ('=74*3', '3 per district'),
    (4, '4 campaigns/year'),
    (4, '4 zonal events'),
    ('=2*60', '2 vans x 60 days'),
    (2, '2 state launches'),
    (200, '200 autos'),
    ('=ROUND(74*0.1,0)*25', '10% dist x 25 stalls'),
    (6, '6 insertions/year'),
]

UPHR_BTL_QTY = [
    ('=ROUND(636*0.5,0)', '50% of 636 UP+HR retailers'),
    ('=ROUND(636*0.25,0)', '25%'),
    ('=ROUND(636*0.2,0)', '20%'),
    ('=ROUND(636*0.1,0)', '10%'),
    ('=ROUND(74*0.5,0)*4', '50% dist x 4 qtrs'),
    ('=74*4', 'all dist x 4 qtrs'),
    ('=74*4', 'all dist x 4 qtrs'),
    (636, 'all UP+HR retailers'),
    ('=ROUND(74*0.3,0)', '30% districts'),
    (22, '40% of 56 distributors'),
    ('=ROUND(74*0.5,0)*50', '50% dist x 50 arch'),
    (22, '22 UP+HR distributors'),
    ('=ROUND(413*0.3,0)*4', '30% transacting x 4 qtrs'),
    ('=ROUND(74*0.2,0)', '20% districts'),
    ('=ROUND(636*0.05,0)', '5% retailers'),
]


# ============================================================
# TAB 4: BUDGET CONTROL & REVIEW (built first - input layer)
# ============================================================
def build_tab4(wb):
    """Input layer: editable parameters, unit costs, scenario analysis, quarterly tracking."""
    ws = wb.create_sheet("Budget Control")
    ws.sheet_properties.tabColor = "FFC000"
    max_col = 10
    set_col_widths(ws, [4, 38, 18, 14, 14, 14, 14, 14, 14, 35])

    # ---- TITLE ----
    ws.merge_cells('A1:J1')
    ws.cell(row=1, column=1, value="BUDGET CONTROL & REVIEW").font = HEADER_FONT_LG
    ws.cell(row=1, column=1).fill = fill('header_dark')
    for c in range(2, max_col + 1):
        ws.cell(row=1, column=c).fill = fill('header_dark')
    ws.merge_cells('A2:J2')
    ws.cell(row=2, column=1, value="Yellow cells are editable. All changes cascade to objective tabs and executive summary.").font = SMALL_FONT

    refs = {}  # store cell references for cross-tab formulas

    # ======== SECTION 1: BUDGET PARAMETERS ========
    r = 4
    write_section_header(ws, r, "SECTION 1: BUDGET PARAMETERS", max_col, 'header_dark')
    r += 1
    write_row(ws, r, ['#', 'Parameter', 'Value', 'Unit', '', '', '', '', '', 'Notes'], font=HEADER_FONT, color_key='header_med')
    r += 1

    # Budget params
    params = [
        ('Total Marketing Budget', 50000000, 'Rs.', '#,##0', 'input_yellow', 'Rs. 5 Crores for FY27'),
        ('Objective 1 - Helix Share %', 0.60, '%', '0.0%', 'input_yellow', '60% for national launch'),
        ('Objective 2 - JSW ONE Share %', None, '%', '0.0%', 'calc_green', 'Auto = 1 - Helix%'),
        ('  Sub-2A East Share (of Obj2)', 0.60, '%', '0.0%', 'input_yellow', '60% of JSW ONE budget'),
        ('  Sub-2B UP+HR Share (of Obj2)', None, '%', '0.0%', 'calc_green', 'Auto = 1 - East%'),
        ('Contingency Reserve %', 0.05, '%', '0.0%', 'input_yellow', '5% held back'),
        ('ATL Share % (Helix)', 0.40, '%', '0.0%', 'input_yellow', '40% ATL, 60% BTL'),
        ('BTL Share % (Helix)', None, '%', '0.0%', 'calc_green', 'Auto = 1 - ATL%'),
    ]
    param_start = r
    for i, (label, val, unit, fmt, clr, note) in enumerate(params):
        row = param_start + i
        ws.cell(row=row, column=1, value=i + 1).font = NORMAL_FONT
        ws.cell(row=row, column=1).border = THIN_BORDER
        ws.cell(row=row, column=2, value=label).font = BOLD_FONT if i == 0 else NORMAL_FONT
        ws.cell(row=row, column=2).border = THIN_BORDER
        cell_v = style_cell(ws, row, 3, clr, BOLD_FONT, fmt, clr != 'input_yellow')
        if val is not None:
            cell_v.value = val
        ws.cell(row=row, column=4, value=unit).font = NORMAL_FONT
        ws.cell(row=row, column=4).border = THIN_BORDER
        ws.cell(row=row, column=10, value=note).font = SMALL_FONT

    # Auto-calc formulas
    ws.cell(row=param_start + 2, column=3).value = f'=1-C{param_start + 1}'  # JSW ONE %
    ws.cell(row=param_start + 4, column=3).value = f'=1-C{param_start + 3}'  # Sub-2B %
    ws.cell(row=param_start + 7, column=3).value = f'=1-C{param_start + 6}'  # BTL %

    refs['total_budget'] = f'C{param_start}'
    refs['helix_pct'] = f'C{param_start + 1}'
    refs['jswone_pct'] = f'C{param_start + 2}'
    refs['east_pct'] = f'C{param_start + 3}'
    refs['uphr_pct'] = f'C{param_start + 4}'
    refs['contingency_pct'] = f'C{param_start + 5}'
    refs['atl_pct'] = f'C{param_start + 6}'
    refs['btl_pct'] = f'C{param_start + 7}'

    r = param_start + len(params) + 1

    # ---- DERIVED BUDGETS ----
    write_section_header(ws, r, "DERIVED BUDGET AMOUNTS", max_col, 'header_light')
    r += 1
    derived_start = r
    derived = [
        ('Helix Budget (after contingency)',
         f'={refs["total_budget"]}*{refs["helix_pct"]}*(1-{refs["contingency_pct"]})'),
        ('JSW ONE Total Budget (after contingency)',
         f'={refs["total_budget"]}*{refs["jswone_pct"]}*(1-{refs["contingency_pct"]})'),
        ('  Sub-2A East Budget',
         f'=C{derived_start + 1}*{refs["east_pct"]}'),
        ('  Sub-2B UP+HR Budget',
         f'=C{derived_start + 1}*{refs["uphr_pct"]}'),
        ('Contingency Reserve',
         f'={refs["total_budget"]}*{refs["contingency_pct"]}'),
        ('Total Allocated (Verification)',
         f'=C{derived_start}+C{derived_start + 1}+C{derived_start + 4}'),
    ]
    for i, (label, formula) in enumerate(derived):
        row = derived_start + i
        ws.cell(row=row, column=1, value='').border = THIN_BORDER
        ws.cell(row=row, column=2, value=label).font = BOLD_FONT
        ws.cell(row=row, column=2).border = THIN_BORDER
        style_cell(ws, row, 3, 'calc_green', BOLD_FONT, '#,##0')
        ws.cell(row=row, column=3).value = formula

    refs['helix_budget'] = f'C{derived_start}'
    refs['jswone_budget'] = f'C{derived_start + 1}'
    refs['east_budget'] = f'C{derived_start + 2}'
    refs['uphr_budget'] = f'C{derived_start + 3}'
    refs['contingency_amt'] = f'C{derived_start + 4}'
    refs['total_allocated'] = f'C{derived_start + 5}'

    r = derived_start + len(derived) + 1

    # ---- VERIFICATION CHECKS ----
    write_section_header(ws, r, "VERIFICATION CHECKS", max_col, 'header_light')
    r += 1
    checks = [
        ('Obj1 + Obj2 = 100%?',
         f'=IF(ABS({refs["helix_pct"]}+{refs["jswone_pct"]}-1)<0.001,"YES","NO")'),
        ('Sub-2A + Sub-2B = Obj2?',
         f'=IF(ABS(C{derived_start + 2}+C{derived_start + 3}-C{derived_start + 1})<1,"YES","NO")'),
        ('Total = Budget?',
         f'=IF(ABS({refs["total_allocated"]}-{refs["total_budget"]})<1,"YES","NO")'),
    ]
    check_start = r
    for i, (label, formula) in enumerate(checks):
        row = check_start + i
        ws.cell(row=row, column=2, value=label).font = BOLD_FONT
        ws.cell(row=row, column=2).border = THIN_BORDER
        ws.cell(row=row, column=3, value=formula).font = BOLD_FONT
        ws.cell(row=row, column=3).border = THIN_BORDER
        # Conditional formatting
        ref = f'C{row}'
        ws.conditional_formatting.add(ref,
            CellIsRule(operator='equal', formula=['"YES"'],
                       fill=PatternFill(fgColor='C6EFCE'), font=Font(color='006100', bold=True)))
        ws.conditional_formatting.add(ref,
            CellIsRule(operator='equal', formula=['"NO"'],
                       fill=PatternFill(fgColor='FFC7CE'), font=Font(color='9C0006', bold=True)))

    r = check_start + len(checks) + 1

    # ---- QUARTERLY ALLOCATION ----
    write_section_header(ws, r, "QUARTERLY ALLOCATION", max_col, 'header_light')
    r += 1
    q_start = r
    q_vals = [
        ('Q1 (Apr-Jun) %', 0.25, 'Launch push + early summer'),
        ('Q2 (Jul-Sep) %', 0.15, 'Monsoon slowdown'),
        ('Q3 (Oct-Dec) %', 0.35, 'Peak construction season'),
        ('Q4 (Jan-Mar) %', 0.25, 'Year-end push'),
    ]
    for i, (label, val, note) in enumerate(q_vals):
        row = q_start + i
        ws.cell(row=row, column=2, value=label).font = NORMAL_FONT
        ws.cell(row=row, column=2).border = THIN_BORDER
        cell_v = style_cell(ws, row, 3, 'input_yellow', BOLD_FONT, '0.0%', False)
        cell_v.value = val
        ws.cell(row=row, column=10, value=note).font = SMALL_FONT

    refs['q1_pct'] = f'C{q_start}'
    refs['q2_pct'] = f'C{q_start + 1}'
    refs['q3_pct'] = f'C{q_start + 2}'
    refs['q4_pct'] = f'C{q_start + 3}'

    # Quarterly sum check
    row = q_start + 4
    ws.cell(row=row, column=2, value='Quarterly Sum = 100%?').font = BOLD_FONT
    ws.cell(row=row, column=2).border = THIN_BORDER
    ws.cell(row=row, column=3).value = f'=IF(ABS(C{q_start}+C{q_start+1}+C{q_start+2}+C{q_start+3}-1)<0.001,"YES","NO")'
    ws.cell(row=row, column=3).font = BOLD_FONT
    ws.cell(row=row, column=3).border = THIN_BORDER
    ws.conditional_formatting.add(f'C{row}',
        CellIsRule(operator='equal', formula=['"YES"'],
                   fill=PatternFill(fgColor='C6EFCE'), font=Font(color='006100', bold=True)))
    ws.conditional_formatting.add(f'C{row}',
        CellIsRule(operator='equal', formula=['"NO"'],
                   fill=PatternFill(fgColor='FFC7CE'), font=Font(color='9C0006', bold=True)))

    r = q_start + 6

    # ---- VOLUME & NETWORK ASSUMPTIONS ----
    write_section_header(ws, r, "VOLUME & NETWORK ASSUMPTIONS", max_col, 'header_light')
    r += 1
    vol_start = r
    vol_params = [
        ('Helix - Distributors', 42, 'nos', 'input_yellow'),
        ('Helix - Annual Volume Target', 121000, 'MT', 'input_yellow'),
        ('Helix - Active Retailers', 2015, 'nos', 'input_yellow'),
        ('Helix - Transacting Retailers', 1310, 'nos', 'input_yellow'),
        ('Helix - Priority Districts', 250, 'nos', 'input_yellow'),
        ('Helix - Selling Price/MT', 48868, 'Rs/MT', 'input_yellow'),
        ('JSW ONE - Distributors', 56, 'nos', 'input_yellow'),
        ('JSW ONE - Annual Volume Target', 268000, 'MT', 'input_yellow'),
        ('JSW ONE - Active Retailers', 1590, 'nos', 'input_yellow'),
        ('JSW ONE - Transacting Retailers', 1033, 'nos', 'input_yellow'),
        ('JSW ONE - Priority Districts', 194, 'nos', 'input_yellow'),
        ('JSW ONE - Selling Price/MT', 50000, 'Rs/MT', 'input_yellow'),
        ('JSW ONE - East Volume Split', 0.60, '%', 'input_yellow'),
        ('JSW ONE - UP+HR Volume Split', None, '%', 'calc_green'),
        ('Sub-2B Growth Multiplier', 2.0, 'x', 'input_yellow'),
    ]
    for i, (label, val, unit, clr) in enumerate(vol_params):
        row = vol_start + i
        ws.cell(row=row, column=2, value=label).font = NORMAL_FONT
        ws.cell(row=row, column=2).border = THIN_BORDER
        fmt = '0.0%' if unit == '%' else ('0.0' if unit == 'x' else '#,##0')
        cell_v = style_cell(ws, row, 3, clr, BOLD_FONT, fmt, clr != 'input_yellow')
        if val is not None:
            cell_v.value = val
        ws.cell(row=row, column=4, value=unit).font = NORMAL_FONT
        ws.cell(row=row, column=4).border = THIN_BORDER

    # Auto-calc UP+HR volume split
    ws.cell(row=vol_start + 13, column=3).value = f'=1-C{vol_start + 12}'

    refs['helix_dist'] = f'C{vol_start}'
    refs['helix_vol'] = f'C{vol_start + 1}'
    refs['helix_retailers'] = f'C{vol_start + 2}'
    refs['helix_transacting'] = f'C{vol_start + 3}'
    refs['helix_districts'] = f'C{vol_start + 4}'
    refs['helix_price'] = f'C{vol_start + 5}'
    refs['jswone_dist'] = f'C{vol_start + 6}'
    refs['jswone_vol'] = f'C{vol_start + 7}'
    refs['jswone_retailers'] = f'C{vol_start + 8}'
    refs['jswone_transacting'] = f'C{vol_start + 9}'
    refs['jswone_districts'] = f'C{vol_start + 10}'
    refs['jswone_price'] = f'C{vol_start + 11}'
    refs['east_vol_split'] = f'C{vol_start + 12}'
    refs['uphr_vol_split'] = f'C{vol_start + 13}'
    refs['growth_mult'] = f'C{vol_start + 14}'

    r = vol_start + len(vol_params) + 1

    # ---- EAST STATE ALLOCATION ----
    write_section_header(ws, r, "EAST STATE BUDGET ALLOCATION", max_col, 'header_light')
    r += 1
    east_start = r
    states = [
        ('Bihar', 0.30, 70, '30% of East budget'),
        ('West Bengal', 0.30, 159, '30% - largest dealer network'),
        ('Jharkhand', 0.20, 61, '20% of East budget'),
        ('Odisha', 0.20, 48, '20% of East budget'),
    ]
    write_row(ws, r, ['', 'State', 'Budget Share %', 'Dealers'], font=HEADER_FONT, color_key='header_med')
    r += 1
    for i, (state, share, dealers, note) in enumerate(states):
        row = r + i
        ws.cell(row=row, column=2, value=state).font = NORMAL_FONT
        ws.cell(row=row, column=2).border = THIN_BORDER
        cell_s = style_cell(ws, row, 3, 'input_yellow', BOLD_FONT, '0.0%', False)
        cell_s.value = share
        ws.cell(row=row, column=4, value=dealers).font = NORMAL_FONT
        ws.cell(row=row, column=4).border = THIN_BORDER
        ws.cell(row=row, column=10, value=note).font = SMALL_FONT

    refs['east_states_start'] = r

    r = r + len(states) + 1

    # ======== SECTION 2: ACTIVITY UNIT COSTS ========
    write_section_header(ws, r, "SECTION 2: ACTIVITY UNIT COSTS (Master Price List)", max_col, 'header_dark')
    r += 1
    write_row(ws, r, ['#', 'Activity', 'Unit Cost (Rs.)', 'Unit', '', '', '', '', '', 'Remarks'], font=HEADER_FONT, color_key='header_med')
    r += 1

    # ATL
    write_section_header(ws, r, "ATL ACTIVITIES (Above The Line)", max_col, 'header_light')
    r += 1
    atl_cost_start = r
    for i, (name, cost, unit, remark) in enumerate(ATL_ACTIVITIES):
        row = atl_cost_start + i
        ws.cell(row=row, column=1, value=i + 1).font = NORMAL_FONT
        ws.cell(row=row, column=1).border = THIN_BORDER
        ws.cell(row=row, column=2, value=name).font = NORMAL_FONT
        ws.cell(row=row, column=2).border = THIN_BORDER
        cell_c = style_cell(ws, row, 3, 'input_yellow', BOLD_FONT, '#,##0', False)
        cell_c.value = cost
        ws.cell(row=row, column=4, value=unit).font = NORMAL_FONT
        ws.cell(row=row, column=4).border = THIN_BORDER
        ws.cell(row=row, column=10, value=remark).font = SMALL_FONT

    refs['atl_cost_start'] = atl_cost_start

    r = atl_cost_start + len(ATL_ACTIVITIES)

    # BTL
    write_section_header(ws, r, "BTL ACTIVITIES (Below The Line)", max_col, 'header_light')
    r += 1
    btl_cost_start = r
    for i, (name, cost, unit, remark) in enumerate(BTL_ACTIVITIES):
        row = btl_cost_start + i
        ws.cell(row=row, column=1, value=len(ATL_ACTIVITIES) + i + 1).font = NORMAL_FONT
        ws.cell(row=row, column=1).border = THIN_BORDER
        ws.cell(row=row, column=2, value=name).font = NORMAL_FONT
        ws.cell(row=row, column=2).border = THIN_BORDER
        cell_c = style_cell(ws, row, 3, 'input_yellow', BOLD_FONT, '#,##0', False)
        cell_c.value = cost
        ws.cell(row=row, column=4, value=unit).font = NORMAL_FONT
        ws.cell(row=row, column=4).border = THIN_BORDER
        ws.cell(row=row, column=10, value=remark).font = SMALL_FONT

    refs['btl_cost_start'] = btl_cost_start

    r = btl_cost_start + len(BTL_ACTIVITIES) + 1

    # ======== SECTION 3: SCENARIO ANALYSIS ========
    write_section_header(ws, r, "SECTION 3: SCENARIO ANALYSIS", max_col, 'header_dark')
    r += 1
    write_row(ws, r, ['', 'Scenario', 'Budget (Rs.)', 'Helix Alloc', 'JSW ONE Alloc',
                       'Helix Cost/MT', 'JSW ONE Cost/MT', 'Blended Cost/MT'],
              font=HEADER_FONT, color_key='header_med')
    r += 1
    refs['scenario_start'] = r
    scenarios = [
        ('Conservative (-20%)', 0.80),
        ('Cautious (-10%)', 0.90),
        ('Base Case (100%)', 1.00),
        ('Aggressive (+10%)', 1.10),
        ('Stretch (+20%)', 1.20),
    ]
    for i, (label, mult) in enumerate(scenarios):
        row = r + i
        ws.cell(row=row, column=2, value=label).font = BOLD_FONT if mult == 1.0 else NORMAL_FONT
        ws.cell(row=row, column=2).border = THIN_BORDER
        # Budget = Total * multiplier
        style_cell(ws, row, 3, 'calc_green', NORMAL_FONT, '#,##0')
        ws.cell(row=row, column=3).value = f'={refs["total_budget"]}*{mult}'
        # Helix alloc
        style_cell(ws, row, 4, 'calc_green', NORMAL_FONT, '#,##0')
        ws.cell(row=row, column=4).value = f'=C{row}*{refs["helix_pct"]}*(1-{refs["contingency_pct"]})'
        # JSW ONE alloc
        style_cell(ws, row, 5, 'calc_green', NORMAL_FONT, '#,##0')
        ws.cell(row=row, column=5).value = f'=C{row}*{refs["jswone_pct"]}*(1-{refs["contingency_pct"]})'
        # Helix Cost/MT
        style_cell(ws, row, 6, 'calc_green', NORMAL_FONT, '#,##0')
        ws.cell(row=row, column=6).value = f'=IF({refs["helix_vol"]}>0,D{row}/{refs["helix_vol"]},0)'
        # JSW ONE Cost/MT
        style_cell(ws, row, 7, 'calc_green', NORMAL_FONT, '#,##0')
        ws.cell(row=row, column=7).value = f'=IF({refs["jswone_vol"]}>0,E{row}/{refs["jswone_vol"]},0)'
        # Blended
        style_cell(ws, row, 8, 'calc_green', NORMAL_FONT, '#,##0')
        ws.cell(row=row, column=8).value = f'=IF(({refs["helix_vol"]}+{refs["jswone_vol"]})>0,(D{row}+E{row})/({refs["helix_vol"]}+{refs["jswone_vol"]}),0)'

        if mult == 1.0:
            for c in range(2, 9):
                ws.cell(row=row, column=c).font = BOLD_FONT

    r = r + len(scenarios) + 1

    # ======== SECTION 4: QUARTERLY ACTUALS vs PLAN ========
    write_section_header(ws, r, "SECTION 4: QUARTERLY ACTUALS vs PLAN TRACKER", max_col, 'header_dark')
    r += 1
    refs['review_start'] = r

    for qi in range(4):
        q_label = QUARTERS[qi]
        q_pct_ref = refs[f'q{qi+1}_pct']

        write_section_header(ws, r, f"{q_label} REVIEW", max_col, 'header_med')
        r += 1
        write_row(ws, r, ['', 'Metric', 'Obj1: Helix', 'Obj2A: East', 'Obj2B: UP+HR', 'Total', 'Remarks'],
                  font=HEADER_FONT, color_key='header_light')
        r += 1

        # Planned row
        ws.cell(row=r, column=2, value='Planned Spend').font = BOLD_FONT
        ws.cell(row=r, column=2).border = THIN_BORDER
        for c in range(3, 6):
            style_cell(ws, r, c, 'calc_green', NORMAL_FONT, '#,##0')
        # Helix planned = Helix budget * Q%
        ws.cell(row=r, column=3).value = f'={refs["helix_budget"]}*{q_pct_ref}'
        # East planned
        ws.cell(row=r, column=4).value = f'={refs["east_budget"]}*{q_pct_ref}'
        # UPHR planned
        ws.cell(row=r, column=5).value = f'={refs["uphr_budget"]}*{q_pct_ref}'
        # Total
        style_cell(ws, r, 6, 'calc_green', BOLD_FONT, '#,##0')
        ws.cell(row=r, column=6).value = f'=C{r}+D{r}+E{r}'
        planned_row = r
        r += 1

        # Actual row
        ws.cell(row=r, column=2, value='Actual Spend').font = BOLD_FONT
        ws.cell(row=r, column=2).border = THIN_BORDER
        for c in range(3, 6):
            style_cell(ws, r, c, 'input_yellow', NORMAL_FONT, '#,##0', False)
            ws.cell(row=r, column=c).value = 0
        style_cell(ws, r, 6, 'calc_green', BOLD_FONT, '#,##0')
        ws.cell(row=r, column=6).value = f'=C{r}+D{r}+E{r}'
        actual_row = r
        r += 1

        # Variance row
        ws.cell(row=r, column=2, value='Variance (Plan - Actual)').font = BOLD_FONT
        ws.cell(row=r, column=2).border = THIN_BORDER
        for c in range(3, 7):
            style_cell(ws, r, c, 'calc_green', BOLD_FONT, '#,##0')
            col_l = get_column_letter(c)
            ws.cell(row=r, column=c).value = f'={col_l}{planned_row}-{col_l}{actual_row}'
            # Red if negative
            ws.conditional_formatting.add(f'{col_l}{r}',
                CellIsRule(operator='lessThan', formula=['0'],
                           fill=PatternFill(fgColor='FFC7CE'), font=Font(color='9C0006', bold=True)))
        r += 1

        # Variance % row
        ws.cell(row=r, column=2, value='Deviation %').font = NORMAL_FONT
        ws.cell(row=r, column=2).border = THIN_BORDER
        for c in range(3, 7):
            style_cell(ws, r, c, 'calc_green', NORMAL_FONT, '0.0%')
            col_l = get_column_letter(c)
            ws.cell(row=r, column=c).value = f'=IF({col_l}{planned_row}<>0,{col_l}{r-1}/{col_l}{planned_row},0)'
            # Red if >15%
            ws.conditional_formatting.add(f'{col_l}{r}',
                FormulaRule(formula=[f'ABS({col_l}{r})>0.15'],
                            fill=PatternFill(fgColor='FFC7CE'), font=Font(color='9C0006')))
        r += 2

    return ws, refs


# ============================================================
# TAB 2: OBJECTIVE 1 - ONE HELIX LAUNCH
# ============================================================
def build_tab2(wb, refs):
    """Complete Helix national launch story: overview, activities, volumes, efficiency."""
    ws = wb.create_sheet("Obj1 - Helix Launch")
    ws.sheet_properties.tabColor = "4472C4"
    max_col = 10
    set_col_widths(ws, [4, 32, 12, 14, 16, 14, 14, 14, 14, 30])
    # Cols: A=# B=Activity C=Qty D=Unit Cost E=Annual Cost F=Q1 G=Q2 H=Q3 I=Q4 J=Remarks

    # ---- TITLE ----
    ws.merge_cells('A1:J1')
    ws.cell(row=1, column=1, value="OBJECTIVE 1: ONE HELIX TMT - NATIONAL LAUNCH").font = HEADER_FONT_LG
    ws.cell(row=1, column=1).fill = fill('header_dark')
    for c in range(2, max_col + 1):
        ws.cell(row=1, column=c).fill = fill('header_dark')

    # ---- SECTION 1: BRAND OVERVIEW ----
    r = 3
    write_section_header(ws, r, "BRAND OVERVIEW", max_col, 'header_med')
    r += 1
    overview = [
        ('Brand', 'One Helix TMT (FE550 Grade Premium Steel)'),
        ('Geography', 'Pan-India, 250 Priority Districts'),
        ('Network', f'={TAB4}{refs["helix_dist"]} & " Distributors, " & {TAB4}{refs["helix_retailers"]} & " Active Retailers"'),
        ('Budget Allocation', None),  # formula
        ('ATL : BTL Split', None),  # formula
        ('Volume Target', None),
        ('Selling Price', None),
    ]
    for i, (label, val) in enumerate(overview):
        row = r + i
        ws.cell(row=row, column=2, value=label).font = BOLD_FONT
        ws.cell(row=row, column=2).border = THIN_BORDER
        ws.cell(row=row, column=2).fill = fill('helix_blue')
        cell_v = ws.cell(row=row, column=3)
        cell_v.border = THIN_BORDER
        cell_v.fill = fill('calc_green')
        if i == 0:
            ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=6)
            cell_v.value = val
            cell_v.font = BOLD_FONT
        elif i == 1:
            ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=6)
            cell_v.value = val
            cell_v.font = NORMAL_FONT
        elif i == 2:
            ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=6)
            cell_v.value = val
            cell_v.font = NORMAL_FONT
        elif i == 3:
            ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=6)
            cell_v.value = f'="Rs. " & TEXT({TAB4}{refs["helix_budget"]},"##,##,##0") & " (" & TEXT({TAB4}{refs["helix_pct"]},"0%") & " of total)"'
            cell_v.font = BOLD_FONT
        elif i == 4:
            ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=6)
            cell_v.value = f'="ATL " & TEXT({TAB4}{refs["atl_pct"]},"0%") & " : BTL " & TEXT({TAB4}{refs["btl_pct"]},"0%")'
            cell_v.font = NORMAL_FONT
        elif i == 5:
            ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=6)
            cell_v.value = f'=TEXT({TAB4}{refs["helix_vol"]},"##,##0") & " MT (Annual)"'
            cell_v.font = BOLD_FONT
        elif i == 6:
            ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=6)
            cell_v.value = f'="Rs. " & TEXT({TAB4}{refs["helix_price"]},"##,##0") & " per MT"'
            cell_v.font = NORMAL_FONT
        for mc in range(4, 7):
            ws.cell(row=row, column=mc).fill = fill('calc_green')
            ws.cell(row=row, column=mc).border = THIN_BORDER

    r = r + len(overview) + 1

    # ---- SECTION 2: ACTIVITY PLAN ----
    write_section_header(ws, r, "ACTIVITY PLAN WITH BUDGET", max_col, 'header_med')
    r += 1
    write_row(ws, r, ['#', 'Activity', 'Quantity', 'Unit Cost (Rs.)', 'Annual Cost (Rs.)',
                       'Q1', 'Q2', 'Q3', 'Q4', 'Basis'],
              font=HEADER_FONT, color_key='header_light')
    r += 1

    # ATL Section
    write_section_header(ws, r, "ATL ACTIVITIES", max_col, 'helix_blue')
    ws.cell(row=r, column=1).font = Font(name='Calibri', bold=True, size=11, color='1F4E79')
    r += 1
    atl_start = r
    atl_cost_ref = refs['atl_cost_start']

    for i in range(len(ATL_ACTIVITIES)):
        row = atl_start + i
        cost_row = atl_cost_ref + i
        ws.cell(row=row, column=1, value=i + 1).font = NORMAL_FONT
        ws.cell(row=row, column=1).border = THIN_BORDER
        ws.cell(row=row, column=2, value=ATL_ACTIVITIES[i][0]).font = NORMAL_FONT
        ws.cell(row=row, column=2).border = THIN_BORDER

        # Quantity
        qty_val = HELIX_ATL_QTY[i][0]
        style_cell(ws, row, 3, 'calc_green', NORMAL_FONT, '#,##0')
        ws.cell(row=row, column=3).value = qty_val

        # Unit Cost (linked to Tab 4)
        style_cell(ws, row, 4, 'output_blue', NORMAL_FONT, '#,##0')
        ws.cell(row=row, column=4).value = f'={TAB4}C{cost_row}'

        # Annual Cost = Qty * Unit Cost
        style_cell(ws, row, 5, 'calc_green', BOLD_FONT, '#,##0')
        ws.cell(row=row, column=5).value = f'=C{row}*D{row}'

        # Quarterly split
        for qi, q_ref in enumerate(['q1_pct', 'q2_pct', 'q3_pct', 'q4_pct']):
            col = 6 + qi
            style_cell(ws, row, col, 'calc_green', NORMAL_FONT, '#,##0')
            ws.cell(row=row, column=col).value = f'=E{row}*{TAB4}{refs[q_ref]}'

        # Remarks
        ws.cell(row=row, column=10, value=HELIX_ATL_QTY[i][1]).font = SMALL_FONT

    # ATL Subtotal
    atl_end = atl_start + len(ATL_ACTIVITIES) - 1
    r = atl_end + 1
    ws.cell(row=r, column=2, value='ATL SUBTOTAL').font = BOLD_FONT
    ws.cell(row=r, column=2).border = THIN_BORDER
    ws.cell(row=r, column=2).fill = fill('total_gray')
    for c in [5, 6, 7, 8, 9]:
        style_cell(ws, r, c, 'total_gray', BOLD_FONT, '#,##0')
        col_l = get_column_letter(c)
        ws.cell(row=r, column=c).value = f'=SUM({col_l}{atl_start}:{col_l}{atl_end})'
    for c in [1, 3, 4, 10]:
        ws.cell(row=r, column=c).fill = fill('total_gray')
        ws.cell(row=r, column=c).border = THIN_BORDER
    atl_subtotal_row = r
    refs['helix_atl_total'] = f'E{r}'
    r += 1

    # BTL Section
    write_section_header(ws, r, "BTL ACTIVITIES", max_col, 'helix_blue')
    ws.cell(row=r, column=1).font = Font(name='Calibri', bold=True, size=11, color='1F4E79')
    r += 1
    btl_start = r
    btl_cost_ref = refs['btl_cost_start']

    for i in range(len(BTL_ACTIVITIES)):
        row = btl_start + i
        cost_row = btl_cost_ref + i
        ws.cell(row=row, column=1, value=len(ATL_ACTIVITIES) + i + 1).font = NORMAL_FONT
        ws.cell(row=row, column=1).border = THIN_BORDER
        ws.cell(row=row, column=2, value=BTL_ACTIVITIES[i][0]).font = NORMAL_FONT
        ws.cell(row=row, column=2).border = THIN_BORDER

        # Quantity
        qty_val = HELIX_BTL_QTY[i][0]
        style_cell(ws, row, 3, 'calc_green', NORMAL_FONT, '#,##0')
        ws.cell(row=row, column=3).value = qty_val

        # Unit Cost
        style_cell(ws, row, 4, 'output_blue', NORMAL_FONT, '#,##0')
        ws.cell(row=row, column=4).value = f'={TAB4}C{cost_row}'

        # Annual Cost
        style_cell(ws, row, 5, 'calc_green', BOLD_FONT, '#,##0')
        ws.cell(row=row, column=5).value = f'=C{row}*D{row}'

        # Quarterly
        for qi, q_ref in enumerate(['q1_pct', 'q2_pct', 'q3_pct', 'q4_pct']):
            col = 6 + qi
            style_cell(ws, row, col, 'calc_green', NORMAL_FONT, '#,##0')
            ws.cell(row=row, column=col).value = f'=E{row}*{TAB4}{refs[q_ref]}'

        ws.cell(row=row, column=10, value=HELIX_BTL_QTY[i][1]).font = SMALL_FONT

    # BTL Subtotal
    btl_end = btl_start + len(BTL_ACTIVITIES) - 1
    r = btl_end + 1
    ws.cell(row=r, column=2, value='BTL SUBTOTAL').font = BOLD_FONT
    ws.cell(row=r, column=2).border = THIN_BORDER
    ws.cell(row=r, column=2).fill = fill('total_gray')
    for c in [5, 6, 7, 8, 9]:
        style_cell(ws, r, c, 'total_gray', BOLD_FONT, '#,##0')
        col_l = get_column_letter(c)
        ws.cell(row=r, column=c).value = f'=SUM({col_l}{btl_start}:{col_l}{btl_end})'
    for c in [1, 3, 4, 10]:
        ws.cell(row=r, column=c).fill = fill('total_gray')
        ws.cell(row=r, column=c).border = THIN_BORDER
    btl_subtotal_row = r
    refs['helix_btl_total'] = f'E{r}'
    r += 1

    # GRAND TOTAL
    ws.cell(row=r, column=2, value='GRAND TOTAL').font = Font(name='Calibri', bold=True, size=12, color='FFFFFF')
    ws.cell(row=r, column=2).fill = fill('header_dark')
    ws.cell(row=r, column=2).border = THIN_BORDER
    for c in [5, 6, 7, 8, 9]:
        style_cell(ws, r, c, 'header_dark', Font(name='Calibri', bold=True, color='FFFFFF'), '#,##0')
        col_l = get_column_letter(c)
        ws.cell(row=r, column=c).value = f'={col_l}{atl_subtotal_row}+{col_l}{btl_subtotal_row}'
    for c in [1, 3, 4, 10]:
        ws.cell(row=r, column=c).fill = fill('header_dark')
        ws.cell(row=r, column=c).border = THIN_BORDER
    grand_total_row = r
    refs['helix_grand_total'] = f'E{r}'
    refs['helix_q1'] = f'F{r}'
    refs['helix_q2'] = f'G{r}'
    refs['helix_q3'] = f'H{r}'
    refs['helix_q4'] = f'I{r}'
    r += 2

    # Budget Variance
    ws.cell(row=r, column=2, value='Allocated Budget').font = BOLD_FONT
    ws.cell(row=r, column=2).border = THIN_BORDER
    style_cell(ws, r, 5, 'output_blue', BOLD_FONT, '#,##0')
    ws.cell(row=r, column=5).value = f'={TAB4}{refs["helix_budget"]}'
    refs['helix_allocated_row'] = r
    r += 1

    ws.cell(row=r, column=2, value='Total Planned Spend').font = BOLD_FONT
    ws.cell(row=r, column=2).border = THIN_BORDER
    style_cell(ws, r, 5, 'calc_green', BOLD_FONT, '#,##0')
    ws.cell(row=r, column=5).value = f'=E{grand_total_row}'
    r += 1

    ws.cell(row=r, column=2, value='Variance (Budget - Spend)').font = BOLD_FONT
    ws.cell(row=r, column=2).border = THIN_BORDER
    style_cell(ws, r, 5, 'calc_green', BOLD_FONT, '#,##0')
    ws.cell(row=r, column=5).value = f'=E{r-2}-E{r-1}'
    ws.conditional_formatting.add(f'E{r}',
        CellIsRule(operator='lessThan', formula=['0'],
                   fill=PatternFill(fgColor='FFC7CE'), font=Font(color='9C0006', bold=True)))
    ws.conditional_formatting.add(f'E{r}',
        CellIsRule(operator='greaterThanOrEqual', formula=['0'],
                   fill=PatternFill(fgColor='C6EFCE'), font=Font(color='006100', bold=True)))
    r += 1

    ws.cell(row=r, column=2, value='Utilization %').font = NORMAL_FONT
    ws.cell(row=r, column=2).border = THIN_BORDER
    style_cell(ws, r, 5, 'calc_green', BOLD_FONT, '0.0%')
    ws.cell(row=r, column=5).value = f'=IF(E{r-3}<>0,E{r-2}/E{r-3},0)'
    r += 2

    # ---- SECTION 3: VOLUME & EFFICIENCY ----
    write_section_header(ws, r, "VOLUME & EFFICIENCY METRICS", max_col, 'header_med')
    r += 1
    write_row(ws, r, ['', 'Metric', '', '', 'Annual', 'Q1', 'Q2', 'Q3', 'Q4'],
              font=HEADER_FONT, color_key='header_light')
    r += 1

    # Helix volume (launch ramp: 10/15/30/45)
    ws.cell(row=r, column=2, value='Volume Target (MT)').font = BOLD_FONT
    ws.cell(row=r, column=2).border = THIN_BORDER
    style_cell(ws, r, 5, 'calc_green', BOLD_FONT, '#,##0')
    ws.cell(row=r, column=5).value = f'={TAB4}{refs["helix_vol"]}'
    ramp = [0.10, 0.15, 0.30, 0.45]
    for qi in range(4):
        style_cell(ws, r, 6 + qi, 'calc_green', NORMAL_FONT, '#,##0')
        ws.cell(row=r, column=6 + qi).value = f'=E{r}*{ramp[qi]}'
    vol_row = r
    r += 1

    # Marketing Spend
    ws.cell(row=r, column=2, value='Marketing Spend (Rs.)').font = BOLD_FONT
    ws.cell(row=r, column=2).border = THIN_BORDER
    style_cell(ws, r, 5, 'calc_green', BOLD_FONT, '#,##0')
    ws.cell(row=r, column=5).value = f'=E{grand_total_row}'
    for qi in range(4):
        col = 6 + qi
        style_cell(ws, r, col, 'calc_green', NORMAL_FONT, '#,##0')
        ws.cell(row=r, column=col).value = f'={get_column_letter(col)}{grand_total_row}'
    spend_row = r
    r += 1

    # Cost per MT
    ws.cell(row=r, column=2, value='Cost per MT (Rs.)').font = BOLD_FONT
    ws.cell(row=r, column=2).border = THIN_BORDER
    style_cell(ws, r, 5, 'calc_green', BOLD_FONT, '#,##0')
    ws.cell(row=r, column=5).value = f'=IF(E{vol_row}>0,E{spend_row}/E{vol_row},0)'
    for qi in range(4):
        col = 6 + qi
        cl = get_column_letter(col)
        style_cell(ws, r, col, 'calc_green', NORMAL_FONT, '#,##0')
        ws.cell(row=r, column=col).value = f'=IF({cl}{vol_row}>0,{cl}{spend_row}/{cl}{vol_row},0)'
    refs['helix_cost_mt'] = f'E{r}'
    r += 1

    # Revenue
    ws.cell(row=r, column=2, value='Revenue (Rs.)').font = BOLD_FONT
    ws.cell(row=r, column=2).border = THIN_BORDER
    style_cell(ws, r, 5, 'calc_green', BOLD_FONT, '#,##0')
    ws.cell(row=r, column=5).value = f'=E{vol_row}*{TAB4}{refs["helix_price"]}'
    for qi in range(4):
        col = 6 + qi
        cl = get_column_letter(col)
        style_cell(ws, r, col, 'calc_green', NORMAL_FONT, '#,##0')
        ws.cell(row=r, column=col).value = f'={cl}{vol_row}*{TAB4}{refs["helix_price"]}'
    rev_row = r
    r += 1

    # A/S Ratio
    ws.cell(row=r, column=2, value='A/S Ratio (Ad Spend / Revenue)').font = BOLD_FONT
    ws.cell(row=r, column=2).border = THIN_BORDER
    style_cell(ws, r, 5, 'calc_green', BOLD_FONT, '0.00%')
    ws.cell(row=r, column=5).value = f'=IF(E{rev_row}>0,E{spend_row}/E{rev_row},0)'
    for qi in range(4):
        col = 6 + qi
        cl = get_column_letter(col)
        style_cell(ws, r, col, 'calc_green', NORMAL_FONT, '0.00%')
        ws.cell(row=r, column=col).value = f'=IF({cl}{rev_row}>0,{cl}{spend_row}/{cl}{rev_row},0)'
    refs['helix_as_ratio'] = f'E{r}'
    r += 1

    # Cost per retailer
    ws.cell(row=r, column=2, value='Cost per Retailer Activated').font = NORMAL_FONT
    ws.cell(row=r, column=2).border = THIN_BORDER
    style_cell(ws, r, 5, 'calc_green', NORMAL_FONT, '#,##0')
    ws.cell(row=r, column=5).value = f'=IF({TAB4}{refs["helix_retailers"]}>0,E{spend_row}/{TAB4}{refs["helix_retailers"]},0)'
    refs['helix_cost_retailer'] = f'E{r}'
    r += 1

    # Cost per district
    ws.cell(row=r, column=2, value='Cost per Priority District').font = NORMAL_FONT
    ws.cell(row=r, column=2).border = THIN_BORDER
    style_cell(ws, r, 5, 'calc_green', NORMAL_FONT, '#,##0')
    ws.cell(row=r, column=5).value = f'=IF({TAB4}{refs["helix_districts"]}>0,E{spend_row}/{TAB4}{refs["helix_districts"]},0)'
    refs['helix_cost_district'] = f'E{r}'

    return ws, refs


# ============================================================
# TAB 3: OBJECTIVE 2 - JSW ONE TMT GROWTH
# ============================================================
def build_tab3(wb, refs):
    """Complete JSW ONE story: East + UP/HR side by side."""
    ws = wb.create_sheet("Obj2 - JSW ONE Growth")
    ws.sheet_properties.tabColor = "70AD47"
    max_col = 12
    set_col_widths(ws, [4, 30, 14, 14, 14, 14, 14, 14, 14, 14, 14, 28])
    # Cols: A=# B=Activity C=2A Qty D=2A Cost E=2B Qty F=2B Cost G=Total H=Q1 I=Q2 J=Q3 K=Q4 L=Remarks

    # ---- TITLE ----
    ws.merge_cells('A1:L1')
    ws.cell(row=1, column=1, value="OBJECTIVE 2: JSW ONE TMT - GROWTH STRATEGY").font = HEADER_FONT_LG
    ws.cell(row=1, column=1).fill = fill('header_dark')
    for c in range(2, max_col + 1):
        ws.cell(row=1, column=c).fill = fill('header_dark')

    # ---- SECTION 1: BRAND OVERVIEW ----
    r = 3
    write_section_header(ws, r, "BRAND OVERVIEW", max_col, 'header_med')
    r += 1

    overview_items = [
        ('Brand', 'JSW ONE TMT'),
        ('Total Obj2 Budget', None),
        ('Network', None),
        ('Total Volume Target', None),
        ('Selling Price', None),
    ]
    for i, (label, val) in enumerate(overview_items):
        row = r + i
        ws.cell(row=row, column=2, value=label).font = BOLD_FONT
        ws.cell(row=row, column=2).border = THIN_BORDER
        ws.cell(row=row, column=2).fill = fill('jsw_green')
        cell_v = ws.cell(row=row, column=3)
        cell_v.border = THIN_BORDER
        cell_v.fill = fill('calc_green')
        ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=7)
        for mc in range(4, 8):
            ws.cell(row=row, column=mc).fill = fill('calc_green')
            ws.cell(row=row, column=mc).border = THIN_BORDER

        if i == 0:
            cell_v.value = val
            cell_v.font = BOLD_FONT
        elif i == 1:
            cell_v.value = f'="Rs. " & TEXT({TAB4}{refs["jswone_budget"]},"##,##,##0") & " (" & TEXT({TAB4}{refs["jswone_pct"]},"0%") & " of total)"'
            cell_v.font = BOLD_FONT
        elif i == 2:
            cell_v.value = f'={TAB4}{refs["jswone_dist"]} & " Distributors, " & {TAB4}{refs["jswone_retailers"]} & " Active Retailers"'
            cell_v.font = NORMAL_FONT
        elif i == 3:
            cell_v.value = f'=TEXT({TAB4}{refs["jswone_vol"]},"##,##0") & " MT (Annual)"'
            cell_v.font = BOLD_FONT
        elif i == 4:
            cell_v.value = f'="Rs. " & TEXT({TAB4}{refs["jswone_price"]},"##,##0") & " per MT"'
            cell_v.font = NORMAL_FONT

    r = r + len(overview_items) + 1

    # Sub-objective cards
    write_section_header(ws, r, "SUB-OBJECTIVES", max_col, 'header_light')
    r += 1
    # Sub-2A
    ws.cell(row=r, column=2, value='Sub-2A: EAST INDIA GROWTH').font = BOLD_FONT
    ws.cell(row=r, column=2).border = THIN_BORDER
    ws.cell(row=r, column=2).fill = fill('jsw_green')
    ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=7)
    ws.cell(row=r, column=3).value = f'="Bihar, WB, Jharkhand, Odisha | Budget: Rs. " & TEXT({TAB4}{refs["east_budget"]},"##,##,##0") & " | 120 Districts"'
    ws.cell(row=r, column=3).font = NORMAL_FONT
    ws.cell(row=r, column=3).border = THIN_BORDER
    ws.cell(row=r, column=3).fill = fill('calc_green')
    for mc in range(4, 8):
        ws.cell(row=r, column=mc).fill = fill('calc_green')
        ws.cell(row=r, column=mc).border = THIN_BORDER
    r += 1

    # Sub-2B
    ws.cell(row=r, column=2, value='Sub-2B: UP & HARYANA 2x GROWTH').font = BOLD_FONT
    ws.cell(row=r, column=2).border = THIN_BORDER
    ws.cell(row=r, column=2).fill = fill('jsw_orange')
    ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=7)
    ws.cell(row=r, column=3).value = f'="UP & Haryana | Budget: Rs. " & TEXT({TAB4}{refs["uphr_budget"]},"##,##,##0") & " | 74 Districts | 2x Target"'
    ws.cell(row=r, column=3).font = NORMAL_FONT
    ws.cell(row=r, column=3).border = THIN_BORDER
    ws.cell(row=r, column=3).fill = fill('calc_green')
    for mc in range(4, 8):
        ws.cell(row=r, column=mc).fill = fill('calc_green')
        ws.cell(row=r, column=mc).border = THIN_BORDER
    r += 2

    # ---- SECTION 2: ACTIVITY PLAN (Side-by-side) ----
    write_section_header(ws, r, "ACTIVITY PLAN - SUB-2A (EAST) vs SUB-2B (UP+HR)", max_col, 'header_med')
    r += 1
    write_row(ws, r, ['#', 'Activity', 'Sub-2A Qty', 'Sub-2A Cost', 'Sub-2B Qty', 'Sub-2B Cost',
                       'Total Cost', 'Q1', 'Q2', 'Q3', 'Q4', 'Basis'],
              font=HEADER_FONT, color_key='header_light')
    r += 1

    # ATL
    write_section_header(ws, r, "ATL ACTIVITIES", max_col, 'jsw_green')
    ws.cell(row=r, column=1).font = Font(name='Calibri', bold=True, size=11, color='1F4E79')
    r += 1
    atl_start = r
    atl_cost_ref = refs['atl_cost_start']

    for i in range(len(ATL_ACTIVITIES)):
        row = atl_start + i
        cost_row = atl_cost_ref + i
        ws.cell(row=row, column=1, value=i + 1).font = NORMAL_FONT
        ws.cell(row=row, column=1).border = THIN_BORDER
        ws.cell(row=row, column=2, value=ATL_ACTIVITIES[i][0]).font = NORMAL_FONT
        ws.cell(row=row, column=2).border = THIN_BORDER

        # Sub-2A Qty
        style_cell(ws, row, 3, 'jsw_green', NORMAL_FONT, '#,##0')
        ws.cell(row=row, column=3).value = EAST_ATL_QTY[i][0]
        # Sub-2A Cost
        style_cell(ws, row, 4, 'calc_green', NORMAL_FONT, '#,##0')
        ws.cell(row=row, column=4).value = f'=C{row}*{TAB4}C{cost_row}'

        # Sub-2B Qty
        style_cell(ws, row, 5, 'jsw_orange', NORMAL_FONT, '#,##0')
        ws.cell(row=row, column=5).value = UPHR_ATL_QTY[i][0]
        # Sub-2B Cost
        style_cell(ws, row, 6, 'calc_green', NORMAL_FONT, '#,##0')
        ws.cell(row=row, column=6).value = f'=E{row}*{TAB4}C{cost_row}'

        # Total
        style_cell(ws, row, 7, 'calc_green', BOLD_FONT, '#,##0')
        ws.cell(row=row, column=7).value = f'=D{row}+F{row}'

        # Quarterly (of total)
        for qi, q_ref in enumerate(['q1_pct', 'q2_pct', 'q3_pct', 'q4_pct']):
            col = 8 + qi
            style_cell(ws, row, col, 'calc_green', NORMAL_FONT, '#,##0')
            ws.cell(row=row, column=col).value = f'=G{row}*{TAB4}{refs[q_ref]}'

        ws.cell(row=row, column=12, value=EAST_ATL_QTY[i][1]).font = SMALL_FONT

    # ATL Subtotal
    atl_end = atl_start + len(ATL_ACTIVITIES) - 1
    r = atl_end + 1
    ws.cell(row=r, column=2, value='ATL SUBTOTAL').font = BOLD_FONT
    ws.cell(row=r, column=2).border = THIN_BORDER
    ws.cell(row=r, column=2).fill = fill('total_gray')
    for c in [4, 6, 7, 8, 9, 10, 11]:
        style_cell(ws, r, c, 'total_gray', BOLD_FONT, '#,##0')
        cl = get_column_letter(c)
        ws.cell(row=r, column=c).value = f'=SUM({cl}{atl_start}:{cl}{atl_end})'
    for c in [1, 3, 5, 12]:
        ws.cell(row=r, column=c).fill = fill('total_gray')
        ws.cell(row=r, column=c).border = THIN_BORDER
    atl_sub_row = r
    refs['jswone_atl_total'] = f'G{r}'
    r += 1

    # BTL
    write_section_header(ws, r, "BTL ACTIVITIES", max_col, 'jsw_green')
    ws.cell(row=r, column=1).font = Font(name='Calibri', bold=True, size=11, color='1F4E79')
    r += 1
    btl_start = r
    btl_cost_ref = refs['btl_cost_start']

    for i in range(len(BTL_ACTIVITIES)):
        row = btl_start + i
        cost_row = btl_cost_ref + i
        ws.cell(row=row, column=1, value=len(ATL_ACTIVITIES) + i + 1).font = NORMAL_FONT
        ws.cell(row=row, column=1).border = THIN_BORDER
        ws.cell(row=row, column=2, value=BTL_ACTIVITIES[i][0]).font = NORMAL_FONT
        ws.cell(row=row, column=2).border = THIN_BORDER

        # Sub-2A Qty
        style_cell(ws, row, 3, 'jsw_green', NORMAL_FONT, '#,##0')
        ws.cell(row=row, column=3).value = EAST_BTL_QTY[i][0]
        # Sub-2A Cost
        style_cell(ws, row, 4, 'calc_green', NORMAL_FONT, '#,##0')
        ws.cell(row=row, column=4).value = f'=C{row}*{TAB4}C{cost_row}'

        # Sub-2B Qty
        style_cell(ws, row, 5, 'jsw_orange', NORMAL_FONT, '#,##0')
        ws.cell(row=row, column=5).value = UPHR_BTL_QTY[i][0]
        # Sub-2B Cost
        style_cell(ws, row, 6, 'calc_green', NORMAL_FONT, '#,##0')
        ws.cell(row=row, column=6).value = f'=E{row}*{TAB4}C{cost_row}'

        # Total
        style_cell(ws, row, 7, 'calc_green', BOLD_FONT, '#,##0')
        ws.cell(row=row, column=7).value = f'=D{row}+F{row}'

        # Quarterly
        for qi, q_ref in enumerate(['q1_pct', 'q2_pct', 'q3_pct', 'q4_pct']):
            col = 8 + qi
            style_cell(ws, row, col, 'calc_green', NORMAL_FONT, '#,##0')
            ws.cell(row=row, column=col).value = f'=G{row}*{TAB4}{refs[q_ref]}'

        ws.cell(row=row, column=12, value=EAST_BTL_QTY[i][1]).font = SMALL_FONT

    # BTL Subtotal
    btl_end = btl_start + len(BTL_ACTIVITIES) - 1
    r = btl_end + 1
    ws.cell(row=r, column=2, value='BTL SUBTOTAL').font = BOLD_FONT
    ws.cell(row=r, column=2).border = THIN_BORDER
    ws.cell(row=r, column=2).fill = fill('total_gray')
    for c in [4, 6, 7, 8, 9, 10, 11]:
        style_cell(ws, r, c, 'total_gray', BOLD_FONT, '#,##0')
        cl = get_column_letter(c)
        ws.cell(row=r, column=c).value = f'=SUM({cl}{btl_start}:{cl}{btl_end})'
    for c in [1, 3, 5, 12]:
        ws.cell(row=r, column=c).fill = fill('total_gray')
        ws.cell(row=r, column=c).border = THIN_BORDER
    btl_sub_row = r
    refs['jswone_btl_total'] = f'G{r}'
    r += 1

    # GRAND TOTAL
    ws.cell(row=r, column=2, value='GRAND TOTAL').font = Font(name='Calibri', bold=True, size=12, color='FFFFFF')
    ws.cell(row=r, column=2).fill = fill('header_dark')
    ws.cell(row=r, column=2).border = THIN_BORDER
    for c in [4, 6, 7, 8, 9, 10, 11]:
        style_cell(ws, r, c, 'header_dark', Font(name='Calibri', bold=True, color='FFFFFF'), '#,##0')
        cl = get_column_letter(c)
        ws.cell(row=r, column=c).value = f'={cl}{atl_sub_row}+{cl}{btl_sub_row}'
    for c in [1, 3, 5, 12]:
        ws.cell(row=r, column=c).fill = fill('header_dark')
        ws.cell(row=r, column=c).border = THIN_BORDER
    grand_total_row = r
    refs['jswone_grand_total'] = f'G{r}'
    refs['jswone_q1'] = f'H{r}'
    refs['jswone_q2'] = f'I{r}'
    refs['jswone_q3'] = f'J{r}'
    refs['jswone_q4'] = f'K{r}'
    refs['east_total'] = f'D{r}'
    refs['uphr_total'] = f'F{r}'
    r += 2

    # Budget Variance
    ws.cell(row=r, column=2, value='Allocated Budget (Obj2 Total)').font = BOLD_FONT
    ws.cell(row=r, column=2).border = THIN_BORDER
    style_cell(ws, r, 7, 'output_blue', BOLD_FONT, '#,##0')
    ws.cell(row=r, column=7).value = f'={TAB4}{refs["jswone_budget"]}'
    r += 1

    ws.cell(row=r, column=2, value='  Sub-2A East Allocated').font = NORMAL_FONT
    ws.cell(row=r, column=2).border = THIN_BORDER
    style_cell(ws, r, 4, 'output_blue', NORMAL_FONT, '#,##0')
    ws.cell(row=r, column=4).value = f'={TAB4}{refs["east_budget"]}'
    r += 1

    ws.cell(row=r, column=2, value='  Sub-2B UP+HR Allocated').font = NORMAL_FONT
    ws.cell(row=r, column=2).border = THIN_BORDER
    style_cell(ws, r, 6, 'output_blue', NORMAL_FONT, '#,##0')
    ws.cell(row=r, column=6).value = f'={TAB4}{refs["uphr_budget"]}'
    r += 1

    ws.cell(row=r, column=2, value='Total Planned Spend').font = BOLD_FONT
    ws.cell(row=r, column=2).border = THIN_BORDER
    style_cell(ws, r, 7, 'calc_green', BOLD_FONT, '#,##0')
    ws.cell(row=r, column=7).value = f'=G{grand_total_row}'
    r += 1

    ws.cell(row=r, column=2, value='Variance (Budget - Spend)').font = BOLD_FONT
    ws.cell(row=r, column=2).border = THIN_BORDER
    style_cell(ws, r, 7, 'calc_green', BOLD_FONT, '#,##0')
    ws.cell(row=r, column=7).value = f'=G{r-4}-G{r-1}'
    ws.conditional_formatting.add(f'G{r}',
        CellIsRule(operator='lessThan', formula=['0'],
                   fill=PatternFill(fgColor='FFC7CE'), font=Font(color='9C0006', bold=True)))
    ws.conditional_formatting.add(f'G{r}',
        CellIsRule(operator='greaterThanOrEqual', formula=['0'],
                   fill=PatternFill(fgColor='C6EFCE'), font=Font(color='006100', bold=True)))
    r += 2

    # ---- SECTION 3: REGIONAL VOLUME & EFFICIENCY ----
    write_section_header(ws, r, "REGIONAL VOLUME & EFFICIENCY", max_col, 'header_med')
    r += 1
    write_row(ws, r, ['', 'Metric', '', '', '', 'Annual', 'Q1', 'Q2', 'Q3', 'Q4'],
              font=HEADER_FONT, color_key='header_light')
    r += 1

    # East volume
    ws.cell(row=r, column=2, value='East Volume Target (MT)').font = BOLD_FONT
    ws.cell(row=r, column=2).border = THIN_BORDER
    style_cell(ws, r, 6, 'calc_green', BOLD_FONT, '#,##0')
    ws.cell(row=r, column=6).value = f'={TAB4}{refs["jswone_vol"]}*{TAB4}{refs["east_vol_split"]}'
    seasonal = [0.25, 0.19, 0.24, 0.32]
    for qi in range(4):
        style_cell(ws, r, 7 + qi, 'calc_green', NORMAL_FONT, '#,##0')
        ws.cell(row=r, column=7 + qi).value = f'=F{r}*{seasonal[qi]}'
    east_vol_row = r
    r += 1

    # UP+HR volume
    ws.cell(row=r, column=2, value='UP+HR Volume Target (MT)').font = BOLD_FONT
    ws.cell(row=r, column=2).border = THIN_BORDER
    style_cell(ws, r, 6, 'calc_green', BOLD_FONT, '#,##0')
    ws.cell(row=r, column=6).value = f'={TAB4}{refs["jswone_vol"]}*{TAB4}{refs["uphr_vol_split"]}'
    for qi in range(4):
        style_cell(ws, r, 7 + qi, 'calc_green', NORMAL_FONT, '#,##0')
        ws.cell(row=r, column=7 + qi).value = f'=F{r}*{seasonal[qi]}'
    uphr_vol_row = r
    r += 1

    # Combined volume
    ws.cell(row=r, column=2, value='Total JSW ONE Volume (MT)').font = BOLD_FONT
    ws.cell(row=r, column=2).border = THIN_BORDER
    ws.cell(row=r, column=2).fill = fill('total_gray')
    style_cell(ws, r, 6, 'total_gray', BOLD_FONT, '#,##0')
    ws.cell(row=r, column=6).value = f'=F{east_vol_row}+F{uphr_vol_row}'
    for qi in range(4):
        col = 7 + qi
        cl = get_column_letter(col)
        style_cell(ws, r, col, 'total_gray', BOLD_FONT, '#,##0')
        ws.cell(row=r, column=col).value = f'={cl}{east_vol_row}+{cl}{uphr_vol_row}'
    total_vol_row = r
    r += 1

    # Cost per MT
    ws.cell(row=r, column=2, value='Cost per MT - JSW ONE (Rs.)').font = BOLD_FONT
    ws.cell(row=r, column=2).border = THIN_BORDER
    style_cell(ws, r, 6, 'calc_green', BOLD_FONT, '#,##0')
    ws.cell(row=r, column=6).value = f'=IF(F{total_vol_row}>0,G{grand_total_row}/F{total_vol_row},0)'
    refs['jswone_cost_mt'] = f'F{r}'
    r += 1

    # Revenue
    ws.cell(row=r, column=2, value='Revenue (Rs.)').font = BOLD_FONT
    ws.cell(row=r, column=2).border = THIN_BORDER
    style_cell(ws, r, 6, 'calc_green', BOLD_FONT, '#,##0')
    ws.cell(row=r, column=6).value = f'=F{total_vol_row}*{TAB4}{refs["jswone_price"]}'
    rev_row = r
    r += 1

    # A/S ratio
    ws.cell(row=r, column=2, value='A/S Ratio').font = BOLD_FONT
    ws.cell(row=r, column=2).border = THIN_BORDER
    style_cell(ws, r, 6, 'calc_green', BOLD_FONT, '0.00%')
    ws.cell(row=r, column=6).value = f'=IF(F{rev_row}>0,G{grand_total_row}/F{rev_row},0)'
    refs['jswone_as_ratio'] = f'F{r}'
    r += 2

    # East state-wise allocation
    write_section_header(ws, r, "EAST STATE-WISE BUDGET ALLOCATION", max_col, 'header_light')
    r += 1
    write_row(ws, r, ['', 'State', 'Budget Share', 'Dealers', 'Budget Amount'],
              font=HEADER_FONT, color_key='header_med')
    r += 1
    east_s = refs['east_states_start']
    states_data = [
        ('Bihar', east_s, 70),
        ('West Bengal', east_s + 1, 159),
        ('Jharkhand', east_s + 2, 61),
        ('Odisha', east_s + 3, 48),
    ]
    for state, state_row, dealers in states_data:
        ws.cell(row=r, column=2, value=state).font = NORMAL_FONT
        ws.cell(row=r, column=2).border = THIN_BORDER
        style_cell(ws, r, 3, 'calc_green', NORMAL_FONT, '0.0%')
        ws.cell(row=r, column=3).value = f'={TAB4}C{state_row}'
        ws.cell(row=r, column=4, value=dealers).font = NORMAL_FONT
        ws.cell(row=r, column=4).border = THIN_BORDER
        style_cell(ws, r, 5, 'calc_green', NORMAL_FONT, '#,##0')
        ws.cell(row=r, column=5).value = f'={TAB4}{refs["east_budget"]}*{TAB4}C{state_row}'
        r += 1

    return ws, refs


# ============================================================
# TAB 1: EXECUTIVE SUMMARY (built last - references all tabs)
# ============================================================
def build_tab1(wb, refs):
    """Executive summary with KPIs, allocation table, charts."""
    ws = wb.create_sheet("Executive Summary")
    ws.sheet_properties.tabColor = "4472C4"
    max_col = 10
    set_col_widths(ws, [4, 28, 16, 16, 16, 16, 16, 16, 14, 30])

    TAB2 = "'Obj1 - Helix Launch'!"
    TAB3 = "'Obj2 - JSW ONE Growth'!"

    # ---- TITLE ----
    ws.merge_cells('A1:J1')
    ws.cell(row=1, column=1, value="JSW STEEL TMT - MARKETING BUDGET FY 2026-27").font = Font(name='Calibri', bold=True, color='FFFFFF', size=16)
    ws.cell(row=1, column=1).fill = fill('header_dark')
    ws.cell(row=1, column=1).alignment = Alignment(horizontal='center', vertical='center')
    for c in range(2, max_col + 1):
        ws.cell(row=1, column=c).fill = fill('header_dark')

    ws.merge_cells('A2:J2')
    ws.cell(row=2, column=1, value="Executive Summary - Management Review").font = Font(name='Calibri', size=11, italic=True, color='FFFFFF')
    ws.cell(row=2, column=1).fill = fill('header_med')
    ws.cell(row=2, column=1).alignment = Alignment(horizontal='center')
    for c in range(2, max_col + 1):
        ws.cell(row=2, column=c).fill = fill('header_med')

    # ---- KPI CARDS ----
    r = 4
    write_section_header(ws, r, "KEY PERFORMANCE INDICATORS", max_col, 'header_dark')
    r += 1

    kpi_data = [
        ('Total Budget', f'={TAB4}{refs["total_budget"]}', '#,##0', 'B'),
        ('Helix Budget', f'={TAB4}{refs["helix_budget"]}', '#,##0', 'D'),
        ('JSW ONE Budget', f'={TAB4}{refs["jswone_budget"]}', '#,##0', 'F'),
        ('Total Volume', f'={TAB4}{refs["helix_vol"]}+{TAB4}{refs["jswone_vol"]}', '#,##0', 'H'),
    ]

    for i, (label, formula, fmt, start_col_letter) in enumerate(kpi_data):
        col = 2 + i * 2
        # Label
        ws.cell(row=r, column=col, value=label).font = KPI_LABEL_FONT
        ws.cell(row=r, column=col).fill = fill('kpi_bg')
        ws.cell(row=r, column=col).border = THIN_BORDER
        ws.cell(row=r, column=col).alignment = Alignment(horizontal='center')
        ws.cell(row=r, column=col + 1).fill = fill('kpi_bg')
        ws.cell(row=r, column=col + 1).border = THIN_BORDER
        ws.merge_cells(start_row=r, start_column=col, end_row=r, end_column=col + 1)

    r += 1
    for i, (label, formula, fmt, start_col_letter) in enumerate(kpi_data):
        col = 2 + i * 2
        cell = ws.cell(row=r, column=col, value=formula)
        cell.font = KPI_FONT
        cell.number_format = fmt
        cell.fill = fill('kpi_bg')
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal='center')
        ws.cell(row=r, column=col + 1).fill = fill('kpi_bg')
        ws.cell(row=r, column=col + 1).border = THIN_BORDER
        ws.merge_cells(start_row=r, start_column=col, end_row=r, end_column=col + 1)

    r += 2

    # Additional KPIs row
    kpi2 = [
        ('Blended Cost/MT', f'=IF(({TAB4}{refs["helix_vol"]}+{TAB4}{refs["jswone_vol"]})>0,({TAB4}{refs["helix_budget"]}+{TAB4}{refs["jswone_budget"]})/({TAB4}{refs["helix_vol"]}+{TAB4}{refs["jswone_vol"]}),0)', '#,##0'),
        ('Contingency', f'={TAB4}{refs["contingency_amt"]}', '#,##0'),
        ('Helix A/S Ratio', f'={TAB2}{refs["helix_as_ratio"]}', '0.00%'),
        ('JSW ONE A/S Ratio', f'={TAB3}{refs["jswone_as_ratio"]}', '0.00%'),
    ]
    for i, (label, formula, fmt) in enumerate(kpi2):
        col = 2 + i * 2
        ws.cell(row=r, column=col, value=label).font = KPI_LABEL_FONT
        ws.cell(row=r, column=col).fill = fill('kpi_bg')
        ws.cell(row=r, column=col).border = THIN_BORDER
        ws.cell(row=r, column=col).alignment = Alignment(horizontal='center')
        ws.cell(row=r, column=col + 1).fill = fill('kpi_bg')
        ws.cell(row=r, column=col + 1).border = THIN_BORDER
        ws.merge_cells(start_row=r, start_column=col, end_row=r, end_column=col + 1)
    r += 1
    for i, (label, formula, fmt) in enumerate(kpi2):
        col = 2 + i * 2
        cell = ws.cell(row=r, column=col, value=formula)
        cell.font = Font(name='Calibri', bold=True, size=13, color='1F4E79')
        cell.number_format = fmt
        cell.fill = fill('kpi_bg')
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal='center')
        ws.cell(row=r, column=col + 1).fill = fill('kpi_bg')
        ws.cell(row=r, column=col + 1).border = THIN_BORDER
        ws.merge_cells(start_row=r, start_column=col, end_row=r, end_column=col + 1)

    r += 2

    # ---- STRATEGIC OBJECTIVES TABLE ----
    write_section_header(ws, r, "STRATEGIC OBJECTIVES OVERVIEW", max_col, 'header_dark')
    r += 1
    write_row(ws, r, ['', 'Objective', 'Brand', 'Geography', 'Budget (Rs.)', 'Volume (MT)',
                       'Cost/MT', 'ATL/BTL', '', 'Key Focus'],
              font=HEADER_FONT, color_key='header_med')
    r += 1

    # Obj1
    ws.cell(row=r, column=2, value='Obj1: National Launch').font = BOLD_FONT
    ws.cell(row=r, column=2).border = THIN_BORDER
    ws.cell(row=r, column=2).fill = fill('helix_blue')
    ws.cell(row=r, column=3, value='One Helix TMT').font = NORMAL_FONT
    ws.cell(row=r, column=3).border = THIN_BORDER
    ws.cell(row=r, column=4, value='Pan-India (250 dist)').font = NORMAL_FONT
    ws.cell(row=r, column=4).border = THIN_BORDER
    cell_b = style_cell(ws, r, 5, 'calc_green', BOLD_FONT, '#,##0')
    cell_b.value = f'={TAB4}{refs["helix_budget"]}'
    cell_v = style_cell(ws, r, 6, 'calc_green', NORMAL_FONT, '#,##0')
    cell_v.value = f'={TAB4}{refs["helix_vol"]}'
    cell_c = style_cell(ws, r, 7, 'calc_green', NORMAL_FONT, '#,##0')
    cell_c.value = f'={TAB2}{refs["helix_cost_mt"]}'
    ws.cell(row=r, column=8, value='40:60').font = NORMAL_FONT
    ws.cell(row=r, column=8).border = THIN_BORDER
    ws.cell(row=r, column=10, value='Brand awareness + dealer activation').font = SMALL_FONT
    obj1_row = r
    r += 1

    # Obj2A
    ws.cell(row=r, column=2, value='Obj2A: East Growth').font = BOLD_FONT
    ws.cell(row=r, column=2).border = THIN_BORDER
    ws.cell(row=r, column=2).fill = fill('jsw_green')
    ws.cell(row=r, column=3, value='JSW ONE TMT').font = NORMAL_FONT
    ws.cell(row=r, column=3).border = THIN_BORDER
    ws.cell(row=r, column=4, value='Bihar, WB, JH, OD').font = NORMAL_FONT
    ws.cell(row=r, column=4).border = THIN_BORDER
    cell_b = style_cell(ws, r, 5, 'calc_green', BOLD_FONT, '#,##0')
    cell_b.value = f'={TAB4}{refs["east_budget"]}'
    cell_v = style_cell(ws, r, 6, 'calc_green', NORMAL_FONT, '#,##0')
    cell_v.value = f'={TAB4}{refs["jswone_vol"]}*{TAB4}{refs["east_vol_split"]}'
    cell_c = style_cell(ws, r, 7, 'calc_green', NORMAL_FONT, '#,##0')
    cell_c.value = f'=IF(F{r}>0,E{r}/F{r},0)'
    ws.cell(row=r, column=8, value='BTL heavy').font = NORMAL_FONT
    ws.cell(row=r, column=8).border = THIN_BORDER
    ws.cell(row=r, column=10, value='Market penetration + retailer network').font = SMALL_FONT
    r += 1

    # Obj2B
    ws.cell(row=r, column=2, value='Obj2B: UP+HR 2x Growth').font = BOLD_FONT
    ws.cell(row=r, column=2).border = THIN_BORDER
    ws.cell(row=r, column=2).fill = fill('jsw_orange')
    ws.cell(row=r, column=3, value='JSW ONE TMT').font = NORMAL_FONT
    ws.cell(row=r, column=3).border = THIN_BORDER
    ws.cell(row=r, column=4, value='UP & Haryana').font = NORMAL_FONT
    ws.cell(row=r, column=4).border = THIN_BORDER
    cell_b = style_cell(ws, r, 5, 'calc_green', BOLD_FONT, '#,##0')
    cell_b.value = f'={TAB4}{refs["uphr_budget"]}'
    cell_v = style_cell(ws, r, 6, 'calc_green', NORMAL_FONT, '#,##0')
    cell_v.value = f'={TAB4}{refs["jswone_vol"]}*{TAB4}{refs["uphr_vol_split"]}'
    cell_c = style_cell(ws, r, 7, 'calc_green', NORMAL_FONT, '#,##0')
    cell_c.value = f'=IF(F{r}>0,E{r}/F{r},0)'
    ws.cell(row=r, column=8, value='BTL heavy').font = NORMAL_FONT
    ws.cell(row=r, column=8).border = THIN_BORDER
    ws.cell(row=r, column=10, value='Aggressive volume push + conversion').font = SMALL_FONT
    r += 1

    # Contingency
    ws.cell(row=r, column=2, value='Contingency Reserve').font = BOLD_FONT
    ws.cell(row=r, column=2).border = THIN_BORDER
    ws.cell(row=r, column=2).fill = fill('total_gray')
    for c in [3, 4, 6, 7, 8]:
        ws.cell(row=r, column=c).fill = fill('total_gray')
        ws.cell(row=r, column=c).border = THIN_BORDER
    cell_b = style_cell(ws, r, 5, 'total_gray', BOLD_FONT, '#,##0')
    cell_b.value = f'={TAB4}{refs["contingency_amt"]}'
    r += 1

    # Total row
    ws.cell(row=r, column=2, value='TOTAL').font = Font(name='Calibri', bold=True, size=11, color='FFFFFF')
    ws.cell(row=r, column=2).fill = fill('header_dark')
    ws.cell(row=r, column=2).border = THIN_BORDER
    for c in [3, 4, 6, 7, 8, 10]:
        ws.cell(row=r, column=c).fill = fill('header_dark')
        ws.cell(row=r, column=c).border = THIN_BORDER
    cell_b = style_cell(ws, r, 5, 'header_dark', Font(name='Calibri', bold=True, color='FFFFFF'), '#,##0')
    cell_b.value = f'={TAB4}{refs["total_budget"]}'
    cell_v = style_cell(ws, r, 6, 'header_dark', Font(name='Calibri', bold=True, color='FFFFFF'), '#,##0')
    cell_v.value = f'={TAB4}{refs["helix_vol"]}+{TAB4}{refs["jswone_vol"]}'
    r += 2

    # ---- BUDGET ALLOCATION BREAKDOWN ----
    write_section_header(ws, r, "BUDGET ALLOCATION BREAKDOWN", max_col, 'header_dark')
    r += 1
    write_row(ws, r, ['', 'Category', 'Obj1: Helix', 'Obj2A: East', 'Obj2B: UP+HR', 'Total Obj2', 'Grand Total', '% of Budget'],
              font=HEADER_FONT, color_key='header_med')
    r += 1

    alloc_rows = [
        ('Allocated Budget',
         f'={TAB4}{refs["helix_budget"]}',
         f'={TAB4}{refs["east_budget"]}',
         f'={TAB4}{refs["uphr_budget"]}'),
        ('ATL Spend',
         f'={TAB2}{refs["helix_atl_total"]}',
         f'={TAB3}D{refs["jswone_atl_total"][1:]}',  # Sub-2A ATL col D
         f'={TAB3}F{refs["jswone_atl_total"][1:]}'),  # Sub-2B ATL col F
        ('BTL Spend',
         f'={TAB2}{refs["helix_btl_total"]}',
         f'={TAB3}D{refs["jswone_btl_total"][1:]}',
         f'={TAB3}F{refs["jswone_btl_total"][1:]}'),
        ('Total Spend',
         f'={TAB2}{refs["helix_grand_total"]}',
         f'={TAB3}{refs["east_total"]}',
         f'={TAB3}{refs["uphr_total"]}'),
    ]

    alloc_start = r
    for i, (label, h_formula, e_formula, u_formula) in enumerate(alloc_rows):
        row = alloc_start + i
        is_total = (label == 'Total Spend')
        fnt = BOLD_FONT if is_total else NORMAL_FONT
        clr = 'total_gray' if is_total else 'calc_green'

        ws.cell(row=row, column=2, value=label).font = fnt
        ws.cell(row=row, column=2).border = THIN_BORDER
        # Helix
        style_cell(ws, row, 3, clr, fnt, '#,##0')
        ws.cell(row=row, column=3).value = h_formula
        # East
        style_cell(ws, row, 4, clr, fnt, '#,##0')
        ws.cell(row=row, column=4).value = e_formula
        # UP+HR
        style_cell(ws, row, 5, clr, fnt, '#,##0')
        ws.cell(row=row, column=5).value = u_formula
        # Total Obj2
        style_cell(ws, row, 6, clr, fnt, '#,##0')
        ws.cell(row=row, column=6).value = f'=D{row}+E{row}'
        # Grand Total
        style_cell(ws, row, 7, clr, fnt, '#,##0')
        ws.cell(row=row, column=7).value = f'=C{row}+F{row}'
        # % of Budget
        style_cell(ws, row, 8, clr, fnt, '0.0%')
        ws.cell(row=row, column=8).value = f'=IF({TAB4}{refs["total_budget"]}<>0,G{row}/{TAB4}{refs["total_budget"]},0)'

    r = alloc_start + len(alloc_rows) + 2

    # ---- QUARTERLY SPEND PLAN ----
    write_section_header(ws, r, "QUARTERLY SPEND PLAN", max_col, 'header_dark')
    r += 1
    write_row(ws, r, ['', 'Quarter', 'Obj1: Helix', 'Obj2: JSW ONE', 'Total', '% of Annual'],
              font=HEADER_FONT, color_key='header_med')
    r += 1

    q_start = r
    for qi in range(4):
        row = q_start + qi
        ws.cell(row=row, column=2, value=QUARTERS[qi]).font = NORMAL_FONT
        ws.cell(row=row, column=2).border = THIN_BORDER
        # Helix quarterly
        style_cell(ws, row, 3, 'helix_blue', NORMAL_FONT, '#,##0')
        h_q_ref = refs[f'helix_q{qi+1}']
        ws.cell(row=row, column=3).value = f'={TAB2}{h_q_ref}'
        # JSW ONE quarterly
        style_cell(ws, row, 4, 'jsw_green', NORMAL_FONT, '#,##0')
        j_q_ref = refs[f'jswone_q{qi+1}']
        ws.cell(row=row, column=4).value = f'={TAB3}{j_q_ref}'
        # Total
        style_cell(ws, row, 5, 'calc_green', BOLD_FONT, '#,##0')
        ws.cell(row=row, column=5).value = f'=C{row}+D{row}'
        # %
        style_cell(ws, row, 6, 'calc_green', NORMAL_FONT, '0.0%')
        ws.cell(row=row, column=6).value = f'=IF(E{q_start + 4}<>0,E{row}/E{q_start + 4},0)'

    # Annual total
    row = q_start + 4
    ws.cell(row=row, column=2, value='ANNUAL TOTAL').font = BOLD_FONT
    ws.cell(row=row, column=2).border = THIN_BORDER
    ws.cell(row=row, column=2).fill = fill('total_gray')
    for c in [3, 4, 5]:
        style_cell(ws, row, c, 'total_gray', BOLD_FONT, '#,##0')
        cl = get_column_letter(c)
        ws.cell(row=row, column=c).value = f'=SUM({cl}{q_start}:{cl}{q_start+3})'
    style_cell(ws, row, 6, 'total_gray', BOLD_FONT, '0.0%')
    ws.cell(row=row, column=6).value = 1.0

    r = row + 2

    # ---- CHART: Quarterly Spend ----
    chart = BarChart()
    chart.type = "col"
    chart.title = "Quarterly Spend Distribution"
    chart.y_axis.title = "Amount (Rs.)"
    chart.x_axis.title = "Quarter"
    chart.style = 10
    chart.width = 22
    chart.height = 12

    # Data
    data_ref = Reference(ws, min_col=3, max_col=4, min_row=q_start - 1, max_row=q_start + 3)
    cats_ref = Reference(ws, min_col=2, min_row=q_start, max_row=q_start + 3)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    chart.shape = 4

    ws.add_chart(chart, f"B{r}")

    return ws


# ============================================================
# MAIN BUILD
# ============================================================
def main():
    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    print("Building Tab 4: Budget Control & Review...")
    ws4, refs = build_tab4(wb)

    print("Building Tab 2: Obj1 - One Helix Launch...")
    ws2, refs = build_tab2(wb, refs)

    print("Building Tab 3: Obj2 - JSW ONE Growth...")
    ws3, refs = build_tab3(wb, refs)

    print("Building Tab 1: Executive Summary...")
    ws1 = build_tab1(wb, refs)

    # Reorder sheets: Executive Summary first
    wb.move_sheet("Executive Summary", offset=-3)

    # Save
    wb.save(OUTPUT_PATH)
    print(f"\nModel saved to: {OUTPUT_PATH}")
    print("Tabs: Executive Summary | Obj1 - Helix Launch | Obj2 - JSW ONE Growth | Budget Control")


if __name__ == '__main__':
    main()
