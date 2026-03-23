"""
JSW Steel TMT Marketing Budget Model Builder
Generates a 10-tab Excel workbook for FY2026-27 marketing budget planning.
"""

import os
from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Border, Side, Alignment, Protection, numbers
)
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.series import DataPoint
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

# === OUTPUT PATH ===
OUTPUT_DIR = os.path.join(os.path.dirname(os.path.dirname(__file__)), "Model")
OUTPUT_PATH = os.path.join(OUTPUT_DIR, "JSW_TMT_Marketing_Budget_Model.xlsx")

# === STYLE CONSTANTS ===
COLORS = {
    'header_dark': '1F4E79',
    'header_med': '2E75B6',
    'header_light': '9DC3E6',
    'input_yellow': 'FFF2CC',
    'calc_green': 'E2EFDA',
    'output_blue': 'D6E4F0',
    'total_gray': 'D9D9D9',
    'alert_red': 'FFC7CE',
    'ok_green': 'C6EFCE',
    'white': 'FFFFFF',
    'obj1_color': 'DAEEF3',
    'obj2a_color': 'EBF1DE',
    'obj2b_color': 'FDE9D9',
}

THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
HEADER_FONT = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
HEADER_FONT_LG = Font(name='Calibri', bold=True, color='FFFFFF', size=14)
TITLE_FONT = Font(name='Calibri', bold=True, size=12, color='1F4E79')
BOLD_FONT = Font(name='Calibri', bold=True, size=11)
NORMAL_FONT = Font(name='Calibri', size=10)
SMALL_FONT = Font(name='Calibri', size=9, italic=True, color='808080')

# Month labels
MONTHS = ['Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec', 'Jan', 'Feb', 'Mar']
QUARTERS = ['Q1', 'Q2', 'Q3', 'Q4']


# === HELPER FUNCTIONS ===

def fill(color_key):
    return PatternFill(start_color=COLORS[color_key], end_color=COLORS[color_key], fill_type='solid')


def style_header_row(ws, row, max_col, color_key='header_dark', font=HEADER_FONT):
    for c in range(1, max_col + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = fill(color_key)
        cell.font = font
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)


def style_range(ws, min_row, max_row, min_col, max_col, color_key='white', font=NORMAL_FONT, num_fmt=None):
    for r in range(min_row, max_row + 1):
        for c in range(min_col, max_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.fill = fill(color_key)
            cell.font = font
            cell.border = THIN_BORDER
            if num_fmt:
                cell.number_format = num_fmt


def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def write_row(ws, row, values, start_col=1, font=NORMAL_FONT, color_key='white', num_fmt=None):
    for i, v in enumerate(values):
        cell = ws.cell(row=row, column=start_col + i, value=v)
        cell.font = font
        cell.fill = fill(color_key)
        cell.border = THIN_BORDER
        cell.alignment = Alignment(vertical='center', wrap_text=True)
        if num_fmt:
            cell.number_format = num_fmt


def write_section_header(ws, row, text, max_col, color_key='header_med'):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=max_col)
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = HEADER_FONT
    cell.fill = fill(color_key)
    cell.border = THIN_BORDER
    cell.alignment = Alignment(horizontal='left', vertical='center')
    for c in range(2, max_col + 1):
        ws.cell(row=row, column=c).fill = fill(color_key)
        ws.cell(row=row, column=c).border = THIN_BORDER


# ============================================================
# TAB 8: ACTIVITY COST REFERENCE
# ============================================================
def build_tab8(wb):
    ws = wb.create_sheet("Activity Costs")
    ws.sheet_properties.tabColor = "FFC000"
    max_col = 5
    set_col_widths(ws, [5, 35, 15, 15, 40])

    # Title
    ws.merge_cells('A1:E1')
    ws.cell(row=1, column=1, value="ACTIVITY COST REFERENCE - MASTER PRICE LIST").font = HEADER_FONT_LG
    ws.cell(row=1, column=1).fill = fill('header_dark')
    for c in range(2, 6):
        ws.cell(row=1, column=c).fill = fill('header_dark')

    ws.cell(row=2, column=1, value="All unit costs are editable. Changes cascade to all objective tabs.").font = SMALL_FONT
    ws.merge_cells('A2:E2')

    # Column headers
    headers = ['#', 'Activity', 'Unit Cost (Rs.)', 'Unit', 'Remarks / Basis']
    write_row(ws, 3, headers, font=HEADER_FONT, color_key='header_med')

    # ATL Activities
    write_section_header(ws, 4, "ATL ACTIVITIES (Above The Line)", max_col, 'header_light')

    atl_items = [
        [1, 'Impact Wall Painting', 15, 'Rs/sq ft', '20,000 sq ft per district'],
        [2, 'FM Radio Campaign', 250000, 'Rs/campaign', 'Per district per quarter'],
        [3, 'Hoardings (Flex/Billboard)', 50000, 'Rs/hoarding', '5 per priority district'],
        [4, 'Digital Marketing Campaign', 500000, 'Rs/campaign', 'Social + YouTube + Display'],
        [5, 'Local Events / Trade Fairs', 500000, 'Rs/event', 'Zone-level events'],
        [6, 'Van Campaign (Mobile)', 7000, 'Rs/day', '2 months per state'],
        [7, 'Syndicate Launch Events', 500000, 'Rs/event', 'Per state launch'],
        [8, 'Transit Auto Branding', 700, 'Rs/auto', 'Auto rickshaw wraps'],
        [9, 'Tea Stall Branding', 4000, 'Rs/stall', '25 stalls per district'],
        [10, 'Paper Advertisement', 100000, 'Rs/insert', 'Regional dailies'],
    ]
    for i, item in enumerate(atl_items):
        r = 5 + i
        write_row(ws, r, item)
        ws.cell(row=r, column=3).number_format = '#,##0'
        ws.cell(row=r, column=3).fill = fill('input_yellow')
        ws.cell(row=r, column=3).protection = Protection(locked=False)

    # BTL Activities
    btl_start = 5 + len(atl_items)
    write_section_header(ws, btl_start, "BTL ACTIVITIES (Below The Line)", max_col, 'header_light')

    btl_items = [
        [11, 'NLB (Name/Lit Board)', 1200, 'Rs/board', '40/sqft x 30 sqft'],
        [12, 'GSB (Glow Sign Board)', 6510, 'Rs/board', '180x30 + 65x18'],
        [13, 'Counter Wall Painting', 10000, 'Rs/counter', 'Retailer counter branding'],
        [14, 'Inshop Branding', 40000, 'Rs/counter', 'Premium in-store display'],
        [15, 'Architect/Engineer Meet', 3500, 'Rs/person', '50 pax per meet'],
        [16, 'Contractor Meet', 1000, 'Rs/person', '50 pax per meet'],
        [17, 'Mason Meet', 300, 'Rs/person', '30 pax per meet'],
        [18, 'Dealer Certificates', 400, 'Rs/certificate', 'Recognition boards'],
        [19, 'Construction Site Gate Board', 1500, 'Rs/board', 'Per active site'],
        [20, 'POP Materials Kit', 3000, 'Rs/dealer', 'Standees, banners, flyers'],
        [21, 'Architect Gift/Engagement', 500, 'Rs/gift', 'Relationship building'],
        [22, 'Distributor Meet (Grand)', 100000, 'Rs/person', 'Annual convention per dist.'],
        [23, 'Retailer Loyalty Program', 2000, 'Rs/retailer/qtr', 'Volume-linked rewards'],
        [24, 'Mason Training Camp', 5000, 'Rs/camp', '20 masons per camp'],
        [25, 'Competitive Conversion Scheme', 10000, 'Rs/retailer', 'Switch-over incentive'],
    ]
    for i, item in enumerate(btl_items):
        r = btl_start + 1 + i
        write_row(ws, r, item)
        ws.cell(row=r, column=3).number_format = '#,##0'
        ws.cell(row=r, column=3).fill = fill('input_yellow')
        ws.cell(row=r, column=3).protection = Protection(locked=False)

    # Store key row references for cross-tab formulas
    # ATL: rows 5-14 (C5:C14), BTL: rows 16-30 (C16:C30)
    return ws


# ============================================================
# TAB 2: MASTER INPUTS
# ============================================================
def build_tab2(wb):
    ws = wb.create_sheet("Master Inputs")
    ws.sheet_properties.tabColor = "00B050"
    max_col = 6
    set_col_widths(ws, [5, 40, 18, 12, 8, 35])

    # Title
    ws.merge_cells('A1:F1')
    c = ws.cell(row=1, column=1, value="MASTER INPUTS - ALL VARIABLE PARAMETERS")
    c.font = HEADER_FONT_LG
    c.fill = fill('header_dark')
    for col in range(2, 7):
        ws.cell(row=1, column=col).fill = fill('header_dark')

    ws.merge_cells('A2:F2')
    ws.cell(row=2, column=1, value="Yellow cells are editable inputs. Green cells are auto-calculated. Change inputs here to update ALL tabs.").font = SMALL_FONT

    r = 4
    # ---- SECTION 1: BUDGET PARAMETERS ----
    write_section_header(ws, r, "BUDGET ALLOCATION CONTROLS", max_col); r += 1

    labels_budget = [
        ('Total Marketing Budget', 50000000, 'Rs.', '5 Crores', 'input_yellow', '#,##0'),
        ('Objective 1 - Helix Share %', 0.60, '', '60% for national launch', 'input_yellow', '0.0%'),
        ('Objective 2 - JSW ONE Share %', None, '', 'Auto = 1 - Helix%', 'calc_green', '0.0%'),
        ('  Sub-2A East Share (of Obj2) %', 0.60, '', '60% for broader East', 'input_yellow', '0.0%'),
        ('  Sub-2B UP+HR Share (of Obj2) %', None, '', 'Auto = 1 - East%', 'calc_green', '0.0%'),
        ('Contingency Reserve %', 0.05, '', '5% held back', 'input_yellow', '0.0%'),
    ]

    budget_start = r
    for i, (label, val, unit, note, clr, fmt) in enumerate(labels_budget):
        row = budget_start + i
        ws.cell(row=row, column=1, value=i+1).font = NORMAL_FONT
        ws.cell(row=row, column=2, value=label).font = BOLD_FONT if i == 0 else NORMAL_FONT
        cell_v = ws.cell(row=row, column=3)
        if val is not None:
            cell_v.value = val
        cell_v.fill = fill(clr)
        cell_v.border = THIN_BORDER
        cell_v.number_format = fmt
        if clr == 'input_yellow':
            cell_v.protection = Protection(locked=False)
        ws.cell(row=row, column=4, value=unit).font = NORMAL_FONT
        ws.cell(row=row, column=6, value=note).font = SMALL_FONT

    # Formulas for auto-calculated cells
    # Row for Obj2 share = 1 - Helix share
    ws.cell(row=budget_start + 2, column=3).value = f'=1-C{budget_start + 1}'
    # Row for Sub-2B = 1 - Sub-2A
    ws.cell(row=budget_start + 4, column=3).value = f'=1-C{budget_start + 3}'

    r = budget_start + len(labels_budget) + 1

    # ---- DERIVED BUDGETS ----
    write_section_header(ws, r, "DERIVED BUDGET AMOUNTS (Auto-Calculated)", max_col); r += 1

    derived_start = r
    derived = [
        ('Helix Budget (Rs.)', f'=C{budget_start}*C{budget_start+1}*(1-C{budget_start+5})', 'calc_green'),
        ('JSW ONE Total Budget (Rs.)', f'=C{budget_start}*C{budget_start+2}*(1-C{budget_start+5})', 'calc_green'),
        ('  Sub-2A East Budget (Rs.)', f'=C{derived_start+1}*C{budget_start+3}', 'calc_green'),
        ('  Sub-2B UP+HR Budget (Rs.)', f'=C{derived_start+1}*C{budget_start+4}', 'calc_green'),
        ('Contingency Reserve (Rs.)', f'=C{budget_start}*C{budget_start+5}', 'calc_green'),
        ('Total Allocated (Rs.)', f'=C{derived_start}+C{derived_start+1}+C{derived_start+4}', 'calc_green'),
    ]
    for i, (label, formula, clr) in enumerate(derived):
        row = derived_start + i
        ws.cell(row=row, column=2, value=label).font = BOLD_FONT
        ws.cell(row=row, column=3, value=formula).number_format = '#,##0'
        ws.cell(row=row, column=3).fill = fill(clr)
        ws.cell(row=row, column=3).border = THIN_BORDER

    r = derived_start + len(derived) + 1

    # ---- VERIFICATION CHECKS ----
    write_section_header(ws, r, "VERIFICATION CHECKS", max_col); r += 1
    checks = [
        ('Budget Allocation = 100%?', f'=IF(ABS(C{budget_start+1}+C{budget_start+2}-1)<0.001,"YES","NO")'),
        ('Sub-Obj sum = Obj2?', f'=IF(ABS(C{derived_start+2}+C{derived_start+3}-C{derived_start+1})<1,"YES","NO")'),
        ('Total = Budget?', f'=IF(ABS(C{derived_start+5}-C{budget_start})<1,"YES","NO")'),
    ]
    check_start = r
    for i, (label, formula) in enumerate(checks):
        row = check_start + i
        ws.cell(row=row, column=2, value=label).font = BOLD_FONT
        ws.cell(row=row, column=3, value=formula).font = BOLD_FONT
        ws.cell(row=row, column=3).border = THIN_BORDER

    # Conditional formatting for checks
    for i in range(len(checks)):
        row = check_start + i
        cell_ref = f'C{row}'
        ws.conditional_formatting.add(cell_ref,
            CellIsRule(operator='equal', formula=['"YES"'],
                       fill=PatternFill(fgColor='C6EFCE'), font=Font(color='006100', bold=True)))
        ws.conditional_formatting.add(cell_ref,
            CellIsRule(operator='equal', formula=['"NO"'],
                       fill=PatternFill(fgColor='FFC7CE'), font=Font(color='9C0006', bold=True)))

    r = check_start + len(checks) + 1

    # ---- ATL / BTL SPLIT ----
    write_section_header(ws, r, "HELIX ATL vs BTL SPLIT", max_col); r += 1
    atl_btl_start = r
    ws.cell(row=r, column=2, value='ATL Share % (Helix)').font = NORMAL_FONT
    ws.cell(row=r, column=3, value=0.40).number_format = '0.0%'
    ws.cell(row=r, column=3).fill = fill('input_yellow')
    ws.cell(row=r, column=3).border = THIN_BORDER
    ws.cell(row=r, column=3).protection = Protection(locked=False)
    ws.cell(row=r, column=6, value='40% ATL, 60% BTL').font = SMALL_FONT
    r += 1
    ws.cell(row=r, column=2, value='BTL Share % (Helix)').font = NORMAL_FONT
    ws.cell(row=r, column=3, value=f'=1-C{atl_btl_start}').number_format = '0.0%'
    ws.cell(row=r, column=3).fill = fill('calc_green')
    ws.cell(row=r, column=3).border = THIN_BORDER
    r += 2

    # ---- QUARTERLY ALLOCATION ----
    write_section_header(ws, r, "QUARTERLY ALLOCATION (Must sum to 100%)", max_col); r += 1
    q_start = r
    q_defaults = [0.25, 0.15, 0.35, 0.25]
    q_notes = ['Apr-Jun: Launch push', 'Jul-Sep: Monsoon dip', 'Oct-Dec: Peak construction', 'Jan-Mar: Year-end push']
    for i, (q, note) in enumerate(zip(QUARTERS, q_notes)):
        row = q_start + i
        ws.cell(row=row, column=2, value=f'{q} Allocation %').font = NORMAL_FONT
        ws.cell(row=row, column=3, value=q_defaults[i]).number_format = '0.0%'
        ws.cell(row=row, column=3).fill = fill('input_yellow')
        ws.cell(row=row, column=3).border = THIN_BORDER
        ws.cell(row=row, column=3).protection = Protection(locked=False)
        ws.cell(row=row, column=6, value=note).font = SMALL_FONT

    r = q_start + 4
    ws.cell(row=r, column=2, value='Quarterly Sum Check').font = BOLD_FONT
    ws.cell(row=r, column=3, value=f'=SUM(C{q_start}:C{q_start+3})').number_format = '0.0%'
    ws.cell(row=r, column=3).border = THIN_BORDER
    ws.conditional_formatting.add(f'C{r}',
        FormulaRule(formula=[f'ABS(C{r}-1)>0.001'],
                    fill=PatternFill(fgColor='FFC7CE'), font=Font(color='9C0006', bold=True)))
    ws.conditional_formatting.add(f'C{r}',
        FormulaRule(formula=[f'ABS(C{r}-1)<=0.001'],
                    fill=PatternFill(fgColor='C6EFCE'), font=Font(color='006100', bold=True)))
    r += 1
    ws.cell(row=r, column=2, value='Quarterly Check').font = BOLD_FONT
    ws.cell(row=r, column=3, value=f'=IF(ABS(C{r-1}-1)<0.001,"YES","NO")').font = BOLD_FONT
    ws.cell(row=r, column=3).border = THIN_BORDER
    ws.conditional_formatting.add(f'C{r}',
        CellIsRule(operator='equal', formula=['"YES"'],
                   fill=PatternFill(fgColor='C6EFCE'), font=Font(color='006100', bold=True)))
    ws.conditional_formatting.add(f'C{r}',
        CellIsRule(operator='equal', formula=['"NO"'],
                   fill=PatternFill(fgColor='FFC7CE'), font=Font(color='9C0006', bold=True)))
    r += 2

    # ---- INTRA-QUARTER MONTHLY WEIGHTS ----
    write_section_header(ws, r, "MONTHLY WEIGHTS WITHIN QUARTERS", max_col); r += 1
    mw_start = r
    monthly_weights = [
        ('Q1 Month Weights (Apr/May/Jun)', [0.30, 0.40, 0.30]),
        ('Q2 Month Weights (Jul/Aug/Sep)', [0.30, 0.35, 0.35]),
        ('Q3 Month Weights (Oct/Nov/Dec)', [0.35, 0.30, 0.35]),
        ('Q4 Month Weights (Jan/Feb/Mar)', [0.40, 0.30, 0.30]),
    ]
    for i, (label, weights) in enumerate(monthly_weights):
        row = mw_start + i
        ws.cell(row=row, column=2, value=label).font = NORMAL_FONT
        for j, w in enumerate(weights):
            ws.cell(row=row, column=3 + j, value=w).number_format = '0.0%'
            ws.cell(row=row, column=3 + j).fill = fill('input_yellow')
            ws.cell(row=row, column=3 + j).border = THIN_BORDER
            ws.cell(row=row, column=3 + j).protection = Protection(locked=False)
    r = mw_start + 4 + 1

    # ---- VOLUME TARGETS ----
    write_section_header(ws, r, "VOLUME TARGETS & ASSUMPTIONS", max_col); r += 1
    vol_start = r
    vol_items = [
        ('Helix - Distributors', 42, 'Nos', 'input_yellow'),
        ('Helix - Annual Volume Target', 121000, 'MT', 'input_yellow'),
        ('Helix - Volume Mar 27 Exit Rate', 27000, 'MT/month', 'input_yellow'),
        ('Helix - Avg Vol/Retailer/Month', 20, 'MT', 'input_yellow'),
        ('Helix - Active Retailers (Mar 27)', 2015, 'Nos', 'input_yellow'),
        ('Helix - Transacting Retailers', 1310, 'Nos', 'input_yellow'),
        ('Helix - Priority Districts', 250, 'Nos', 'input_yellow'),
        ('Helix - Selling Price/MT', 48868, 'Rs/MT', 'input_yellow'),
        ('JSW ONE - Distributors', 56, 'Nos', 'input_yellow'),
        ('JSW ONE - Annual Volume Target', 268000, 'MT', 'input_yellow'),
        ('JSW ONE - Volume Mar 27 Exit Rate', 31000, 'MT/month', 'input_yellow'),
        ('JSW ONE - Avg Vol/Retailer/Month', 30, 'MT', 'input_yellow'),
        ('JSW ONE - Active Retailers (Mar 27)', 1590, 'Nos', 'input_yellow'),
        ('JSW ONE - Transacting Retailers', 1033, 'Nos', 'input_yellow'),
        ('JSW ONE - Priority Districts (Total)', 194, 'Nos', 'input_yellow'),
        ('JSW ONE - Priority Districts (East)', 120, 'Nos', 'input_yellow'),
        ('JSW ONE - Priority Districts (UP+HR)', 74, 'Nos', 'input_yellow'),
        ('JSW ONE - Selling Price/MT', 50000, 'Rs/MT', 'input_yellow'),
        ('JSW ONE - East Volume Share %', 0.60, '', 'input_yellow'),
        ('JSW ONE - UP+HR Volume Share %', 0.40, '', 'input_yellow'),
        ('Target Growth Multiplier (Sub-2B)', 2.0, 'x', 'input_yellow'),
    ]
    for i, (label, val, unit, clr) in enumerate(vol_items):
        row = vol_start + i
        ws.cell(row=row, column=2, value=label).font = NORMAL_FONT
        cell_v = ws.cell(row=row, column=3, value=val)
        cell_v.fill = fill(clr)
        cell_v.border = THIN_BORDER
        cell_v.protection = Protection(locked=False)
        if '%' in label:
            cell_v.number_format = '0.0%'
        elif 'Price' in label or 'Budget' in label:
            cell_v.number_format = '#,##0'
        else:
            cell_v.number_format = '#,##0'
        ws.cell(row=row, column=4, value=unit).font = NORMAL_FONT

    r = vol_start + len(vol_items) + 1

    # ---- GEOGRAPHY: EAST STATES ----
    write_section_header(ws, r, "EAST REGION STATES (Sub-2A Coverage)", max_col); r += 1
    geo_start = r
    east_states = [
        ('Bihar', 0.30, 70),
        ('West Bengal', 0.30, 159),
        ('Jharkhand', 0.20, 61),
        ('Odisha', 0.20, 48),
    ]
    ws.cell(row=r, column=2, value='State').font = BOLD_FONT
    ws.cell(row=r, column=3, value='Budget Share %').font = BOLD_FONT
    ws.cell(row=r, column=4, value='Dealers').font = BOLD_FONT
    ws.cell(row=r, column=2).border = THIN_BORDER
    ws.cell(row=r, column=3).border = THIN_BORDER
    ws.cell(row=r, column=4).border = THIN_BORDER
    r += 1
    state_start = r
    for i, (state, share, dealers) in enumerate(east_states):
        row = state_start + i
        ws.cell(row=row, column=2, value=state).font = NORMAL_FONT
        ws.cell(row=row, column=2).border = THIN_BORDER
        ws.cell(row=row, column=3, value=share).number_format = '0.0%'
        ws.cell(row=row, column=3).fill = fill('input_yellow')
        ws.cell(row=row, column=3).border = THIN_BORDER
        ws.cell(row=row, column=3).protection = Protection(locked=False)
        ws.cell(row=row, column=4, value=dealers).number_format = '#,##0'
        ws.cell(row=row, column=4).fill = fill('input_yellow')
        ws.cell(row=row, column=4).border = THIN_BORDER
        ws.cell(row=row, column=4).protection = Protection(locked=False)

    r = state_start + len(east_states)
    ws.cell(row=r, column=2, value='East State Share Check').font = BOLD_FONT
    ws.cell(row=r, column=3, value=f'=SUM(C{state_start}:C{r-1})').number_format = '0.0%'
    ws.cell(row=r, column=3).border = THIN_BORDER
    r += 2

    # ---- UP / HARYANA SPLIT ----
    write_section_header(ws, r, "UP & HARYANA SPLIT (Sub-2B)", max_col); r += 1
    uphr_start = r
    ws.cell(row=r, column=2, value='Uttar Pradesh Budget Share %').font = NORMAL_FONT
    ws.cell(row=r, column=3, value=0.60).number_format = '0.0%'
    ws.cell(row=r, column=3).fill = fill('input_yellow')
    ws.cell(row=r, column=3).border = THIN_BORDER
    ws.cell(row=r, column=3).protection = Protection(locked=False)
    r += 1
    ws.cell(row=r, column=2, value='Haryana Budget Share %').font = NORMAL_FONT
    ws.cell(row=r, column=3, value=f'=1-C{uphr_start}').number_format = '0.0%'
    ws.cell(row=r, column=3).fill = fill('calc_green')
    ws.cell(row=r, column=3).border = THIN_BORDER
    r += 1
    ws.cell(row=r, column=2, value='UP Dealers (Current)').font = NORMAL_FONT
    ws.cell(row=r, column=3, value=152).number_format = '#,##0'
    ws.cell(row=r, column=3).fill = fill('input_yellow')
    ws.cell(row=r, column=3).border = THIN_BORDER
    ws.cell(row=r, column=3).protection = Protection(locked=False)
    r += 1
    ws.cell(row=r, column=2, value='Haryana Dealers (Current)').font = NORMAL_FONT
    ws.cell(row=r, column=3, value=81).number_format = '#,##0'
    ws.cell(row=r, column=3).fill = fill('input_yellow')
    ws.cell(row=r, column=3).border = THIN_BORDER
    ws.cell(row=r, column=3).protection = Protection(locked=False)

    # Data Validation
    dv_pct = DataValidation(type="decimal", operator="between",
                            formula1=0, formula2=1,
                            errorTitle="Invalid", error="Enter 0 to 1 (0%-100%)")
    ws.add_data_validation(dv_pct)
    dv_pct.add(ws.cell(row=budget_start + 1, column=3))  # Helix share
    dv_pct.add(ws.cell(row=budget_start + 3, column=3))  # Sub-2A share
    dv_pct.add(ws.cell(row=budget_start + 5, column=3))  # Contingency
    dv_pct.add(ws.cell(row=atl_btl_start, column=3))     # ATL share
    for i in range(4):
        dv_pct.add(ws.cell(row=q_start + i, column=3))

    dv_pos = DataValidation(type="whole", operator="greaterThan", formula1=0,
                            errorTitle="Invalid", error="Enter a positive number")
    ws.add_data_validation(dv_pos)
    dv_pos.add(ws.cell(row=budget_start, column=3))  # Total budget

    # Return key row references for other tabs
    return {
        'total_budget': f"'Master Inputs'!$C${budget_start}",
        'helix_pct': f"'Master Inputs'!$C${budget_start + 1}",
        'obj2_pct': f"'Master Inputs'!$C${budget_start + 2}",
        'sub2a_pct': f"'Master Inputs'!$C${budget_start + 3}",
        'sub2b_pct': f"'Master Inputs'!$C${budget_start + 4}",
        'contingency_pct': f"'Master Inputs'!$C${budget_start + 5}",
        'helix_budget': f"'Master Inputs'!$C${derived_start}",
        'obj2_budget': f"'Master Inputs'!$C${derived_start + 1}",
        'sub2a_budget': f"'Master Inputs'!$C${derived_start + 2}",
        'sub2b_budget': f"'Master Inputs'!$C${derived_start + 3}",
        'contingency_amt': f"'Master Inputs'!$C${derived_start + 4}",
        'atl_pct': f"'Master Inputs'!$C${atl_btl_start}",
        'btl_pct': f"'Master Inputs'!$C${atl_btl_start + 1}",
        'q1_pct': f"'Master Inputs'!$C${q_start}",
        'q2_pct': f"'Master Inputs'!$C${q_start + 1}",
        'q3_pct': f"'Master Inputs'!$C${q_start + 2}",
        'q4_pct': f"'Master Inputs'!$C${q_start + 3}",
        'mw_start': mw_start,  # row number for monthly weights
        'vol_start': vol_start,
        'helix_distributors': f"'Master Inputs'!$C${vol_start}",
        'helix_annual_vol': f"'Master Inputs'!$C${vol_start + 1}",
        'helix_active_ret': f"'Master Inputs'!$C${vol_start + 4}",
        'helix_transacting': f"'Master Inputs'!$C${vol_start + 5}",
        'helix_priority_dist': f"'Master Inputs'!$C${vol_start + 6}",
        'helix_price': f"'Master Inputs'!$C${vol_start + 7}",
        'jsw_distributors': f"'Master Inputs'!$C${vol_start + 8}",
        'jsw_annual_vol': f"'Master Inputs'!$C${vol_start + 9}",
        'jsw_active_ret': f"'Master Inputs'!$C${vol_start + 12}",
        'jsw_transacting': f"'Master Inputs'!$C${vol_start + 13}",
        'jsw_priority_east': f"'Master Inputs'!$C${vol_start + 15}",
        'jsw_priority_uphr': f"'Master Inputs'!$C${vol_start + 16}",
        'jsw_price': f"'Master Inputs'!$C${vol_start + 17}",
        'jsw_east_vol_pct': f"'Master Inputs'!$C${vol_start + 18}",
        'jsw_uphr_vol_pct': f"'Master Inputs'!$C${vol_start + 19}",
        'growth_mult': f"'Master Inputs'!$C${vol_start + 20}",
        'state_start': state_start,
        'uphr_start': uphr_start,
        'up_pct': f"'Master Inputs'!$C${uphr_start}",
        'hr_pct': f"'Master Inputs'!$C${uphr_start + 1}",
        'up_dealers': f"'Master Inputs'!$C${uphr_start + 2}",
        'hr_dealers': f"'Master Inputs'!$C${uphr_start + 3}",
        'q_start': q_start,
        'derived_start': derived_start,
        'budget_start': budget_start,
        'atl_btl_start': atl_btl_start,
    }


# ============================================================
# OBJECTIVE TAB BUILDER (shared logic for Tabs 3, 4, 5)
# ============================================================
def build_objective_tab(wb, sheet_name, tab_color, title, refs, budget_ref,
                        atl_activities, btl_activities, is_helix=False):
    """
    Generic builder for objective tabs (Helix, East, UP/HR).
    Each activity: (name, annual_formula, remarks)
    """
    ws = wb.create_sheet(sheet_name)
    ws.sheet_properties.tabColor = tab_color

    # Columns: A=Activity, B=Remarks, C=Year Total, D=%ofTotal,
    # E-G=Q1 months (Apr,May,Jun), H=Q1 Total,
    # I-K=Q2 months, L=Q2 Total,
    # M-O=Q3 months, P=Q3 Total,
    # Q-S=Q4 months, T=Q4 Total
    max_col = 20
    set_col_widths(ws, [35, 20, 15, 8, 12, 12, 12, 14, 12, 12, 12, 14, 12, 12, 12, 14, 12, 12, 12, 14])

    # Title row
    ws.merge_cells(f'A1:{get_column_letter(max_col)}1')
    ws.cell(row=1, column=1, value=title).font = HEADER_FONT_LG
    ws.cell(row=1, column=1).fill = fill('header_dark')
    for c in range(2, max_col + 1):
        ws.cell(row=1, column=c).fill = fill('header_dark')

    # Budget reference row
    ws.cell(row=2, column=1, value='Allocated Budget:').font = BOLD_FONT
    ws.cell(row=2, column=2, value=f'={budget_ref}').number_format = '#,##0'
    ws.cell(row=2, column=2).fill = fill('calc_green')
    ws.cell(row=2, column=2).border = THIN_BORDER

    if is_helix:
        ws.cell(row=2, column=4, value='ATL Budget:').font = BOLD_FONT
        ws.cell(row=2, column=5, value=f'={budget_ref}*{refs["atl_pct"]}').number_format = '#,##0'
        ws.cell(row=2, column=5).fill = fill('calc_green')
        ws.cell(row=2, column=5).border = THIN_BORDER
        ws.cell(row=2, column=7, value='BTL Budget:').font = BOLD_FONT
        ws.cell(row=2, column=8, value=f'={budget_ref}*{refs["btl_pct"]}').number_format = '#,##0'
        ws.cell(row=2, column=8).fill = fill('calc_green')
        ws.cell(row=2, column=8).border = THIN_BORDER

    # Column headers row 4
    r = 4
    col_headers = ['Activity', 'Remarks', 'Year Total', '% of Total']
    for qi in range(4):
        m_start = qi * 3
        col_headers += [MONTHS[m_start], MONTHS[m_start + 1], MONTHS[m_start + 2], f'{QUARTERS[qi]} Total']
    write_row(ws, r, col_headers, font=HEADER_FONT, color_key='header_med')

    # Monthly weight references from Master Inputs
    mw = refs['mw_start']
    q_refs = [refs['q1_pct'], refs['q2_pct'], refs['q3_pct'], refs['q4_pct']]

    def write_activity_section(start_row, activities, section_name, color_key):
        """Write a section of activities and return (section_start_row, section_end_row, total_row)"""
        # Section header
        write_section_header(ws, start_row, section_name, max_col, 'header_light')
        current = start_row + 1

        for act_name, annual_formula, remarks in activities:
            ws.cell(row=current, column=1, value=act_name).font = NORMAL_FONT
            ws.cell(row=current, column=1).fill = fill(color_key)
            ws.cell(row=current, column=1).border = THIN_BORDER

            ws.cell(row=current, column=2, value=remarks).font = SMALL_FONT
            ws.cell(row=current, column=2).fill = fill(color_key)
            ws.cell(row=current, column=2).border = THIN_BORDER

            # Year total (column C = 3)
            ws.cell(row=current, column=3, value=f'={annual_formula}').number_format = '#,##0'
            ws.cell(row=current, column=3).fill = fill(color_key)
            ws.cell(row=current, column=3).border = THIN_BORDER

            # % of total (will fill after grand total is known)
            ws.cell(row=current, column=4).fill = fill(color_key)
            ws.cell(row=current, column=4).border = THIN_BORDER

            # Monthly breakdown: for each quarter
            col = 5
            for qi in range(4):
                q_pct = q_refs[qi]
                for mi in range(3):
                    mw_col = get_column_letter(3 + mi)
                    mw_row = mw + qi
                    formula = f'=$C${current}*{q_pct}*\'Master Inputs\'!${mw_col}${mw_row}'
                    ws.cell(row=current, column=col, value=formula).number_format = '#,##0'
                    ws.cell(row=current, column=col).fill = fill(color_key)
                    ws.cell(row=current, column=col).border = THIN_BORDER
                    col += 1
                # Quarterly total
                q_start_col = col - 3
                q_end_col = col - 1
                formula = f'=SUM({get_column_letter(q_start_col)}{current}:{get_column_letter(q_end_col)}{current})'
                ws.cell(row=current, column=col, value=formula).number_format = '#,##0'
                ws.cell(row=current, column=col).fill = fill('total_gray')
                ws.cell(row=current, column=col).font = BOLD_FONT
                ws.cell(row=current, column=col).border = THIN_BORDER
                col += 1

            current += 1

        # Section total row
        total_row = current
        ws.cell(row=total_row, column=1, value=f'{section_name} TOTAL').font = BOLD_FONT
        ws.cell(row=total_row, column=1).fill = fill('total_gray')
        ws.cell(row=total_row, column=1).border = THIN_BORDER

        first_act = start_row + 1
        last_act = current - 1
        for c in range(2, max_col + 1):
            ws.cell(row=total_row, column=c).fill = fill('total_gray')
            ws.cell(row=total_row, column=c).border = THIN_BORDER
            ws.cell(row=total_row, column=c).font = BOLD_FONT
            if c >= 3:
                col_l = get_column_letter(c)
                ws.cell(row=total_row, column=c,
                        value=f'=SUM({col_l}{first_act}:{col_l}{last_act})').number_format = '#,##0'

        return start_row + 1, last_act, total_row

    # ATL Section
    atl_first, atl_last, atl_total = write_activity_section(5, atl_activities, 'ATL ACTIVITIES', 'output_blue')

    # BTL Section
    btl_start_row = atl_total + 2
    btl_first, btl_last, btl_total = write_activity_section(btl_start_row, btl_activities, 'BTL ACTIVITIES', 'calc_green')

    # Grand Total
    gt_row = btl_total + 2
    ws.cell(row=gt_row, column=1, value='GRAND TOTAL').font = Font(name='Calibri', bold=True, size=12)
    ws.cell(row=gt_row, column=1).fill = fill('header_dark')
    ws.cell(row=gt_row, column=1).font = Font(name='Calibri', bold=True, size=12, color='FFFFFF')
    for c in range(2, max_col + 1):
        ws.cell(row=gt_row, column=c).fill = fill('header_dark')
        ws.cell(row=gt_row, column=c).font = Font(name='Calibri', bold=True, color='FFFFFF')
        ws.cell(row=gt_row, column=c).border = THIN_BORDER
        if c >= 3:
            col_l = get_column_letter(c)
            ws.cell(row=gt_row, column=c,
                    value=f'={col_l}{atl_total}+{col_l}{btl_total}').number_format = '#,##0'

    # Fill % of total column (D) for each activity
    for act_row in range(atl_first, atl_last + 1):
        ws.cell(row=act_row, column=4, value=f'=IF(C${gt_row}=0,0,C{act_row}/C${gt_row})').number_format = '0.0%'
    for act_row in range(btl_first, btl_last + 1):
        ws.cell(row=act_row, column=4, value=f'=IF(C${gt_row}=0,0,C{act_row}/C${gt_row})').number_format = '0.0%'
    ws.cell(row=atl_total, column=4, value=f'=IF(C${gt_row}=0,0,C{atl_total}/C${gt_row})').number_format = '0.0%'
    ws.cell(row=btl_total, column=4, value=f'=IF(C${gt_row}=0,0,C{btl_total}/C${gt_row})').number_format = '0.0%'
    ws.cell(row=gt_row, column=4, value='100.0%').number_format = '0.0%'

    # Budget variance row
    var_row = gt_row + 1
    ws.cell(row=var_row, column=1, value='Allocated Budget').font = BOLD_FONT
    ws.cell(row=var_row, column=3, value=f'={budget_ref}').number_format = '#,##0'
    ws.cell(row=var_row, column=3).border = THIN_BORDER
    var_row += 1
    ws.cell(row=var_row, column=1, value='Variance (Budget - Spend)').font = BOLD_FONT
    ws.cell(row=var_row, column=3, value=f'=C{var_row-1}-C{gt_row}').number_format = '#,##0'
    ws.cell(row=var_row, column=3).border = THIN_BORDER
    ws.conditional_formatting.add(f'C{var_row}',
        CellIsRule(operator='lessThan', formula=['0'],
                   fill=PatternFill(fgColor='FFC7CE'), font=Font(color='9C0006', bold=True)))
    ws.conditional_formatting.add(f'C{var_row}',
        CellIsRule(operator='greaterThanOrEqual', formula=['0'],
                   fill=PatternFill(fgColor='C6EFCE'), font=Font(color='006100', bold=True)))
    var_row += 1
    ws.cell(row=var_row, column=1, value='Utilization %').font = BOLD_FONT
    ws.cell(row=var_row, column=3, value=f'=IF(C{var_row-2}=0,0,C{gt_row}/C{var_row-2})').number_format = '0.0%'
    ws.cell(row=var_row, column=3).border = THIN_BORDER

    return {'gt_row': gt_row, 'atl_total': atl_total, 'btl_total': btl_total,
            'sheet_name': sheet_name}


# ============================================================
# TAB 3: OBJECTIVE 1 - HELIX LAUNCH
# ============================================================
def build_tab3(wb, refs):
    # ATL activities: (name, annual_formula, remarks)
    # Formulas reference Activity Costs tab (C column = unit cost) and Master Inputs
    atl = [
        ('Impact Wall Painting',
         f"'Activity Costs'!C5*{refs['helix_priority_dist']}*20000",
         '20K sqft x districts'),
        ('FM Radio Campaign',
         f"'Activity Costs'!C6*{refs['helix_priority_dist']}*0.05*4",
         '5% districts x 4 qtrs'),
        ('Hoardings (Flex/Billboard)',
         f"'Activity Costs'!C7*{refs['helix_priority_dist']}*2",
         '2 per district'),
        ('Digital Marketing Campaign',
         f"'Activity Costs'!C8*4",
         '4 campaigns/year'),
        ('Local Events / Trade Fairs',
         f"'Activity Costs'!C9*8",
         '8 zonal events'),
        ('Van Campaign (Mobile)',
         f"'Activity Costs'!C10*10*60",
         '10 vans x 60 days'),
        ('Syndicate Launch Events',
         f"'Activity Costs'!C11*15",
         '15 state launches'),
        ('Transit Auto Branding',
         f"'Activity Costs'!C12*500",
         '500 autos nationally'),
        ('Tea Stall Branding',
         f"'Activity Costs'!C13*{refs['helix_priority_dist']}*0.1*25",
         '10% districts x 25 stalls'),
        ('Paper Advertisement',
         f"'Activity Costs'!C14*12",
         '12 insertions/year'),
    ]

    btl = [
        ('NLB (Name/Lit Board)',
         f"'Activity Costs'!C16*{refs['helix_active_ret']}*0.5",
         '50% active retailers'),
        ('GSB (Glow Sign Board)',
         f"'Activity Costs'!C17*{refs['helix_active_ret']}*0.25",
         '25% active retailers'),
        ('Counter Wall Painting',
         f"'Activity Costs'!C18*{refs['helix_active_ret']}*0.2",
         '20% active retailers'),
        ('Inshop Branding',
         f"'Activity Costs'!C19*{refs['helix_active_ret']}*0.1",
         '10% active retailers'),
        ('Architect/Engineer Meets',
         f"'Activity Costs'!C20*50*{refs['helix_distributors']}*0.5*4",
         '50 pax x 50% dist x 4 qtrs'),
        ('Contractor Meets',
         f"'Activity Costs'!C21*50*{refs['helix_distributors']}*4",
         '50 pax x all dist x 4 qtrs'),
        ('Mason Meets',
         f"'Activity Costs'!C22*30*{refs['helix_distributors']}*4",
         '30 pax x all dist x 4 qtrs'),
        ('Dealer Certificates',
         f"'Activity Costs'!C23*{refs['helix_active_ret']}",
         'All active retailers'),
        ('POP Materials Kit',
         f"'Activity Costs'!C25*{refs['helix_distributors']}",
         'All distributors'),
        ('Distributor Meet (Grand)',
         f"'Activity Costs'!C27*{refs['helix_distributors']}",
         'Annual convention'),
    ]

    return build_objective_tab(wb, 'Obj1 Helix', '0070C0',
                               'OBJECTIVE 1: ONE HELIX INDIA LAUNCH - ATL + BTL PLAN (FY 2026-27)',
                               refs, refs['helix_budget'], atl, btl, is_helix=True)


# ============================================================
# TAB 4: OBJECTIVE 2A - JSW ONE EAST GROWTH
# ============================================================
def build_tab4(wb, refs):
    atl = [
        ('Impact Wall Painting',
         f"'Activity Costs'!C5*{refs['jsw_priority_east']}*15000",
         '15K sqft x East districts'),
        ('Digital Marketing (Regional)',
         f"'Activity Costs'!C8*2",
         '2 East-focused campaigns'),
        ('Local Market Events',
         f"'Activity Costs'!C9*4",
         '4 events across East'),
    ]

    btl = [
        ('NLB (Name/Lit Board)',
         f"'Activity Costs'!C16*{refs['jsw_active_ret']}*{refs['jsw_east_vol_pct']}*0.4",
         '40% of East retailers'),
        ('GSB (Glow Sign Board)',
         f"'Activity Costs'!C17*{refs['jsw_active_ret']}*{refs['jsw_east_vol_pct']}*0.2",
         '20% of East retailers'),
        ('Counter Wall Painting',
         f"'Activity Costs'!C18*{refs['jsw_active_ret']}*{refs['jsw_east_vol_pct']}*0.15",
         '15% of East retailers'),
        ('Architect/Engineer Meets',
         f"'Activity Costs'!C20*50*{refs['jsw_distributors']}*{refs['jsw_east_vol_pct']}*4",
         '50 pax x East dist x 4 qtrs'),
        ('Contractor Meets',
         f"'Activity Costs'!C21*50*{refs['jsw_distributors']}*{refs['jsw_east_vol_pct']}*4",
         '50 pax x East dist x 4 qtrs'),
        ('Mason Meets',
         f"'Activity Costs'!C22*30*{refs['jsw_distributors']}*{refs['jsw_east_vol_pct']}*4",
         '30 pax x East dist x 4 qtrs'),
        ('Construction Site Gate Boards',
         f"'Activity Costs'!C24*{refs['jsw_priority_east']}*5",
         '5 boards per district'),
        ('POP Materials Kit',
         f"'Activity Costs'!C25*{refs['jsw_distributors']}*{refs['jsw_east_vol_pct']}",
         'East distributors'),
        ('Retailer Loyalty Program',
         f"'Activity Costs'!C28*{refs['jsw_transacting']}*{refs['jsw_east_vol_pct']}*4",
         'Per transacting retailer x 4 qtrs'),
        ('Dealer Certificates',
         f"'Activity Costs'!C23*{refs['jsw_active_ret']}*{refs['jsw_east_vol_pct']}",
         'East active retailers'),
    ]

    return build_objective_tab(wb, 'Obj2A East', '92D050',
                               'OBJECTIVE 2A: JSW ONE TMT - EAST INDIA GROWTH (BTL-Focused)',
                               refs, refs['sub2a_budget'], atl, btl)


# ============================================================
# TAB 5: OBJECTIVE 2B - JSW ONE UP & HARYANA
# ============================================================
def build_tab5(wb, refs):
    atl = [
        ('Impact Wall Painting',
         f"'Activity Costs'!C5*{refs['jsw_priority_uphr']}*15000",
         '15K sqft x UP+HR districts'),
        ('Transit Auto Branding',
         f"'Activity Costs'!C12*200",
         '200 autos in UP+HR'),
    ]

    btl = [
        ('NLB (Name/Lit Board)',
         f"'Activity Costs'!C16*{refs['jsw_active_ret']}*{refs['jsw_uphr_vol_pct']}*0.5",
         '50% UP+HR retailers'),
        ('GSB (Glow Sign Board)',
         f"'Activity Costs'!C17*{refs['jsw_active_ret']}*{refs['jsw_uphr_vol_pct']}*0.3",
         '30% UP+HR retailers'),
        ('Counter Wall Painting',
         f"'Activity Costs'!C18*{refs['jsw_active_ret']}*{refs['jsw_uphr_vol_pct']}*0.2",
         '20% UP+HR retailers'),
        ('Inshop Branding',
         f"'Activity Costs'!C19*{refs['jsw_active_ret']}*{refs['jsw_uphr_vol_pct']}*0.1",
         '10% premium retailers'),
        ('Monthly Dealer Conclaves',
         f"'Activity Costs'!C20*50*({refs['up_dealers']}+{refs['hr_dealers']})*0.3*12",
         '30% dealers x 50 pax x 12 months'),
        ('Contractor Meets (Bi-weekly)',
         f"'Activity Costs'!C21*30*({refs['up_dealers']}+{refs['hr_dealers']})*0.5*24",
         '50% dealers x 30 pax x 24 meets'),
        ('Mason Training Camps',
         f"'Activity Costs'!C29*{refs['jsw_priority_uphr']}*4",
         '4 camps per district/year'),
        ('Construction Site Gate Boards',
         f"'Activity Costs'!C24*{refs['jsw_priority_uphr']}*8",
         '8 boards per district'),
        ('POP Materials Kit',
         f"'Activity Costs'!C25*({refs['up_dealers']}+{refs['hr_dealers']})",
         'All UP+HR dealers'),
        ('Retailer Loyalty Program',
         f"'Activity Costs'!C28*{refs['jsw_transacting']}*{refs['jsw_uphr_vol_pct']}*4",
         'Per transacting x 4 qtrs'),
        ('Competitive Conversion Scheme',
         f"'Activity Costs'!C30*{refs['jsw_active_ret']}*{refs['jsw_uphr_vol_pct']}*0.15",
         '15% target conversion'),
        ('Dealer Certificates',
         f"'Activity Costs'!C23*{refs['jsw_active_ret']}*{refs['jsw_uphr_vol_pct']}",
         'UP+HR active retailers'),
        ('Distributor Meet',
         f"'Activity Costs'!C27*({refs['up_dealers']}+{refs['hr_dealers']})*0.1",
         '10% of dealers at annual meet'),
    ]

    return build_objective_tab(wb, 'Obj2B UPHR', 'FF6600',
                               'OBJECTIVE 2B: JSW ONE TMT - UP & HARYANA INTENSIVE (2x Volume Target)',
                               refs, refs['sub2b_budget'], atl, btl)


# ============================================================
# TAB 6: CONSOLIDATED BUDGET
# ============================================================
def build_tab6(wb, refs, obj1_info, obj2a_info, obj2b_info):
    ws = wb.create_sheet("Consolidated")
    ws.sheet_properties.tabColor = "7030A0"
    max_col = 8
    set_col_widths(ws, [30, 18, 18, 18, 18, 18, 18, 12])

    # Title
    ws.merge_cells(f'A1:{get_column_letter(max_col)}1')
    ws.cell(row=1, column=1, value="CONSOLIDATED BUDGET - ALL OBJECTIVES (FY 2026-27)").font = HEADER_FONT_LG
    ws.cell(row=1, column=1).fill = fill('header_dark')
    for c in range(2, max_col + 1):
        ws.cell(row=1, column=c).fill = fill('header_dark')

    # Headers
    headers = ['Category', 'Obj 1: Helix\n(India Launch)',
               'Obj 2A: JSW ONE\n(East Growth)', 'Obj 2B: JSW ONE\n(UP+HR 2x)',
               'Obj 2 Total', 'GRAND TOTAL', 'Contingency', '% of Total']
    write_row(ws, 3, headers, font=HEADER_FONT, color_key='header_med')
    for c in range(1, max_col + 1):
        ws.cell(row=3, column=c).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    r = 5
    # Budget allocation summary
    items = [
        ('Allocated Budget', refs['helix_budget'], refs['sub2a_budget'], refs['sub2b_budget']),
        ('ATL Spend', f"'{obj1_info['sheet_name']}'!C{obj1_info['atl_total']}",
         f"'{obj2a_info['sheet_name']}'!C{obj2a_info['atl_total']}",
         f"'{obj2b_info['sheet_name']}'!C{obj2b_info['atl_total']}"),
        ('BTL Spend', f"'{obj1_info['sheet_name']}'!C{obj1_info['btl_total']}",
         f"'{obj2a_info['sheet_name']}'!C{obj2a_info['btl_total']}",
         f"'{obj2b_info['sheet_name']}'!C{obj2b_info['btl_total']}"),
        ('Total Spend', f"'{obj1_info['sheet_name']}'!C{obj1_info['gt_row']}",
         f"'{obj2a_info['sheet_name']}'!C{obj2a_info['gt_row']}",
         f"'{obj2b_info['sheet_name']}'!C{obj2b_info['gt_row']}"),
    ]

    for i, (label, f1, f2, f3) in enumerate(items):
        row = r + i
        ws.cell(row=row, column=1, value=label).font = BOLD_FONT
        ws.cell(row=row, column=1).border = THIN_BORDER
        ws.cell(row=row, column=2, value=f'={f1}').number_format = '#,##0'
        ws.cell(row=row, column=2).border = THIN_BORDER
        ws.cell(row=row, column=2).fill = fill('obj1_color')
        ws.cell(row=row, column=3, value=f'={f2}').number_format = '#,##0'
        ws.cell(row=row, column=3).border = THIN_BORDER
        ws.cell(row=row, column=3).fill = fill('obj2a_color')
        ws.cell(row=row, column=4, value=f'={f3}').number_format = '#,##0'
        ws.cell(row=row, column=4).border = THIN_BORDER
        ws.cell(row=row, column=4).fill = fill('obj2b_color')
        ws.cell(row=row, column=5, value=f'=C{row}+D{row}').number_format = '#,##0'
        ws.cell(row=row, column=5).border = THIN_BORDER
        ws.cell(row=row, column=5).fill = fill('total_gray')
        ws.cell(row=row, column=6, value=f'=B{row}+E{row}').number_format = '#,##0'
        ws.cell(row=row, column=6).border = THIN_BORDER
        ws.cell(row=row, column=6).font = BOLD_FONT

        if i == 0:  # Allocated Budget row - add contingency
            ws.cell(row=row, column=7, value=f'={refs["contingency_amt"]}').number_format = '#,##0'
            ws.cell(row=row, column=7).border = THIN_BORDER

        ws.cell(row=row, column=8, value=f'=IF(F${r+3}=0,0,F{row}/F${r+3})').number_format = '0.0%'
        ws.cell(row=row, column=8).border = THIN_BORDER

    # Highlight total spend row
    total_row = r + 3
    style_range(ws, total_row, total_row, 1, max_col, 'header_dark')
    for c in range(1, max_col + 1):
        ws.cell(row=total_row, column=c).font = Font(name='Calibri', bold=True, color='FFFFFF')

    r = total_row + 2

    # Variance
    ws.cell(row=r, column=1, value='Budget Variance').font = BOLD_FONT
    ws.cell(row=r, column=1).border = THIN_BORDER
    for c in range(2, 7):
        col_l = get_column_letter(c)
        ws.cell(row=r, column=c, value=f'={col_l}{r-6}-{col_l}{r-3}').number_format = '#,##0'
        ws.cell(row=r, column=c).border = THIN_BORDER
    variance_row = r

    # Conditional formatting on variance
    ws.conditional_formatting.add(f'B{variance_row}:F{variance_row}',
        CellIsRule(operator='lessThan', formula=['0'],
                   fill=PatternFill(fgColor='FFC7CE'), font=Font(color='9C0006', bold=True)))
    ws.conditional_formatting.add(f'B{variance_row}:F{variance_row}',
        CellIsRule(operator='greaterThanOrEqual', formula=['0'],
                   fill=PatternFill(fgColor='C6EFCE'), font=Font(color='006100', bold=True)))

    r += 1
    ws.cell(row=r, column=1, value='Utilization %').font = BOLD_FONT
    ws.cell(row=r, column=1).border = THIN_BORDER
    for c in range(2, 7):
        col_l = get_column_letter(c)
        ws.cell(row=r, column=c, value=f'=IF({col_l}{r-7}=0,0,{col_l}{r-4}/{col_l}{r-7})').number_format = '0.0%'
        ws.cell(row=r, column=c).border = THIN_BORDER

    r += 2

    # Quarterly Breakdown
    write_section_header(ws, r, "QUARTERLY SPEND BREAKDOWN", max_col); r += 1
    q_headers = ['Quarter', 'Obj 1: Helix', 'Obj 2A: East', 'Obj 2B: UP+HR',
                 'Obj 2 Total', 'Grand Total', '', '% of Annual']
    write_row(ws, r, q_headers, font=HEADER_FONT, color_key='header_light'); r += 1

    # Q total columns in objective tabs: H=8, L=12, P=16, T=20
    q_cols = [8, 12, 16, 20]
    q_start_row = r
    for qi in range(4):
        col_l = get_column_letter(q_cols[qi])
        row = r + qi
        ws.cell(row=row, column=1, value=QUARTERS[qi]).font = BOLD_FONT
        ws.cell(row=row, column=1).border = THIN_BORDER
        for obj_idx, info in enumerate([obj1_info, obj2a_info, obj2b_info], 2):
            gt = info['gt_row']
            ws.cell(row=row, column=obj_idx,
                    value=f"='{info['sheet_name']}'!{col_l}{gt}").number_format = '#,##0'
            ws.cell(row=row, column=obj_idx).border = THIN_BORDER
        ws.cell(row=row, column=5, value=f'=C{row}+D{row}').number_format = '#,##0'
        ws.cell(row=row, column=5).border = THIN_BORDER
        ws.cell(row=row, column=6, value=f'=B{row}+E{row}').number_format = '#,##0'
        ws.cell(row=row, column=6).border = THIN_BORDER
        ws.cell(row=row, column=6).font = BOLD_FONT
        annual_total = f'F{total_row}'
        ws.cell(row=row, column=8, value=f'=IF({annual_total}=0,0,F{row}/{annual_total})').number_format = '0.0%'
        ws.cell(row=row, column=8).border = THIN_BORDER

    # Annual total row
    r = q_start_row + 4
    ws.cell(row=r, column=1, value='Annual Total').font = BOLD_FONT
    ws.cell(row=r, column=1).fill = fill('total_gray')
    ws.cell(row=r, column=1).border = THIN_BORDER
    for c in range(2, 7):
        col_l = get_column_letter(c)
        ws.cell(row=r, column=c,
                value=f'=SUM({col_l}{q_start_row}:{col_l}{q_start_row+3})').number_format = '#,##0'
        ws.cell(row=r, column=c).fill = fill('total_gray')
        ws.cell(row=r, column=c).border = THIN_BORDER
        ws.cell(row=r, column=c).font = BOLD_FONT

    return {'total_row': total_row, 'variance_row': variance_row,
            'q_start_row': q_start_row, 'sheet_name': 'Consolidated'}


# ============================================================
# TAB 7: VOLUME PROJECTIONS & COST EFFICIENCY
# ============================================================
def build_tab7(wb, refs, cons_info, obj1_info, obj2a_info, obj2b_info):
    ws = wb.create_sheet("Vol Projections")
    ws.sheet_properties.tabColor = "00B0F0"
    max_col = 8
    set_col_widths(ws, [30, 15, 15, 15, 15, 15, 15, 15])

    # Title
    ws.merge_cells(f'A1:{get_column_letter(max_col)}1')
    ws.cell(row=1, column=1, value="VOLUME PROJECTIONS & COST EFFICIENCY (FY 2026-27)").font = HEADER_FONT_LG
    ws.cell(row=1, column=1).fill = fill('header_dark')
    for c in range(2, max_col + 1):
        ws.cell(row=1, column=c).fill = fill('header_dark')

    r = 3
    # Volume Summary
    write_section_header(ws, r, "VOLUME SUMMARY (MT)", max_col); r += 1
    vol_headers = ['Brand', 'Q1', 'Q2', 'Q3', 'Q4', 'Annual Total', '', '']
    write_row(ws, r, vol_headers, font=HEADER_FONT, color_key='header_light'); r += 1

    # Helix quarterly volumes (ramp-up: 10%, 15%, 30%, 45% of annual)
    helix_vol_row = r
    ws.cell(row=r, column=1, value='One Helix (New Launch)').font = BOLD_FONT
    ws.cell(row=r, column=1).border = THIN_BORDER
    helix_q_pcts = [0.10, 0.15, 0.30, 0.45]  # Launch ramp-up curve
    for qi, pct in enumerate(helix_q_pcts):
        ws.cell(row=r, column=2 + qi, value=f'={refs["helix_annual_vol"]}*{pct}').number_format = '#,##0'
        ws.cell(row=r, column=2 + qi).border = THIN_BORDER
    ws.cell(row=r, column=6, value=f'=SUM(B{r}:E{r})').number_format = '#,##0'
    ws.cell(row=r, column=6).border = THIN_BORDER
    ws.cell(row=r, column=6).font = BOLD_FONT
    r += 1

    # JSW ONE quarterly volumes
    jsw_vol_row = r
    ws.cell(row=r, column=1, value='JSW ONE TMT (Total)').font = BOLD_FONT
    ws.cell(row=r, column=1).border = THIN_BORDER
    jsw_q_pcts = [0.25, 0.19, 0.24, 0.32]  # From reference
    for qi, pct in enumerate(jsw_q_pcts):
        ws.cell(row=r, column=2 + qi, value=f'={refs["jsw_annual_vol"]}*{pct}').number_format = '#,##0'
        ws.cell(row=r, column=2 + qi).border = THIN_BORDER
    ws.cell(row=r, column=6, value=f'=SUM(B{r}:E{r})').number_format = '#,##0'
    ws.cell(row=r, column=6).border = THIN_BORDER
    ws.cell(row=r, column=6).font = BOLD_FONT
    r += 1

    # Combined
    combined_row = r
    ws.cell(row=r, column=1, value='COMBINED').font = BOLD_FONT
    ws.cell(row=r, column=1).fill = fill('total_gray')
    ws.cell(row=r, column=1).border = THIN_BORDER
    for c in range(2, 7):
        col_l = get_column_letter(c)
        ws.cell(row=r, column=c, value=f'={col_l}{helix_vol_row}+{col_l}{jsw_vol_row}').number_format = '#,##0'
        ws.cell(row=r, column=c).fill = fill('total_gray')
        ws.cell(row=r, column=c).border = THIN_BORDER
        ws.cell(row=r, column=c).font = BOLD_FONT
    r += 2

    # Cost per MT
    write_section_header(ws, r, "MARKETING COST PER MT (Rs.)", max_col); r += 1
    cost_headers = ['Brand', 'Q1', 'Q2', 'Q3', 'Q4', 'Annual Avg', '', '']
    write_row(ws, r, cost_headers, font=HEADER_FONT, color_key='header_light'); r += 1

    # Helix cost/MT
    q_start = cons_info['q_start_row']
    ws.cell(row=r, column=1, value='One Helix Cost/MT').font = BOLD_FONT
    ws.cell(row=r, column=1).border = THIN_BORDER
    for qi in range(4):
        ws.cell(row=r, column=2 + qi,
                value=f"=IF(B{helix_vol_row}=0,0,'Consolidated'!B{q_start+qi}/B{helix_vol_row})").number_format = '#,##0'
        ws.cell(row=r, column=2 + qi).border = THIN_BORDER
    ws.cell(row=r, column=6,
            value=f"=IF(F{helix_vol_row}=0,0,'Consolidated'!B{cons_info['total_row']}/F{helix_vol_row})").number_format = '#,##0'
    ws.cell(row=r, column=6).border = THIN_BORDER
    ws.cell(row=r, column=6).font = BOLD_FONT
    r += 1

    # JSW ONE cost/MT
    ws.cell(row=r, column=1, value='JSW ONE Cost/MT').font = BOLD_FONT
    ws.cell(row=r, column=1).border = THIN_BORDER
    for qi in range(4):
        ws.cell(row=r, column=2 + qi,
                value=f"=IF(B{jsw_vol_row}=0,0,('Consolidated'!C{q_start+qi}+'Consolidated'!D{q_start+qi})/B{jsw_vol_row})").number_format = '#,##0'
        ws.cell(row=r, column=2 + qi).border = THIN_BORDER
    ws.cell(row=r, column=6,
            value=f"=IF(F{jsw_vol_row}=0,0,('Consolidated'!C{cons_info['total_row']}+'Consolidated'!D{cons_info['total_row']})/F{jsw_vol_row})").number_format = '#,##0'
    ws.cell(row=r, column=6).border = THIN_BORDER
    ws.cell(row=r, column=6).font = BOLD_FONT
    r += 1

    # Blended
    ws.cell(row=r, column=1, value='Blended Cost/MT').font = BOLD_FONT
    ws.cell(row=r, column=1).fill = fill('total_gray')
    ws.cell(row=r, column=1).border = THIN_BORDER
    for qi in range(4):
        ws.cell(row=r, column=2 + qi,
                value=f"=IF(B{combined_row}=0,0,'Consolidated'!F{q_start+qi}/B{combined_row})").number_format = '#,##0'
        ws.cell(row=r, column=2 + qi).fill = fill('total_gray')
        ws.cell(row=r, column=2 + qi).border = THIN_BORDER
    ws.cell(row=r, column=6,
            value=f"=IF(F{combined_row}=0,0,'Consolidated'!F{cons_info['total_row']}/F{combined_row})").number_format = '#,##0'
    ws.cell(row=r, column=6).fill = fill('total_gray')
    ws.cell(row=r, column=6).border = THIN_BORDER
    ws.cell(row=r, column=6).font = BOLD_FONT
    r += 2

    # Revenue & A/S Ratio
    write_section_header(ws, r, "REVENUE & A/S (Advertising-to-Sales) RATIO", max_col); r += 1
    rev_headers = ['Metric', 'Helix', 'JSW ONE', 'Combined', '', '', '', '']
    write_row(ws, r, rev_headers, font=HEADER_FONT, color_key='header_light'); r += 1

    rev_start = r
    ws.cell(row=r, column=1, value='Projected Revenue (Rs.)').font = BOLD_FONT
    ws.cell(row=r, column=1).border = THIN_BORDER
    ws.cell(row=r, column=2, value=f'=F{helix_vol_row}*{refs["helix_price"]}').number_format = '#,##0'
    ws.cell(row=r, column=2).border = THIN_BORDER
    ws.cell(row=r, column=3, value=f'=F{jsw_vol_row}*{refs["jsw_price"]}').number_format = '#,##0'
    ws.cell(row=r, column=3).border = THIN_BORDER
    ws.cell(row=r, column=4, value=f'=B{r}+C{r}').number_format = '#,##0'
    ws.cell(row=r, column=4).border = THIN_BORDER
    r += 1

    ws.cell(row=r, column=1, value='Marketing Spend (Rs.)').font = BOLD_FONT
    ws.cell(row=r, column=1).border = THIN_BORDER
    ws.cell(row=r, column=2, value=f"='Consolidated'!B{cons_info['total_row']}").number_format = '#,##0'
    ws.cell(row=r, column=2).border = THIN_BORDER
    ws.cell(row=r, column=3, value=f"='Consolidated'!E{cons_info['total_row']}").number_format = '#,##0'
    ws.cell(row=r, column=3).border = THIN_BORDER
    ws.cell(row=r, column=4, value=f"='Consolidated'!F{cons_info['total_row']}").number_format = '#,##0'
    ws.cell(row=r, column=4).border = THIN_BORDER
    r += 1

    ws.cell(row=r, column=1, value='A/S Ratio (%)').font = BOLD_FONT
    ws.cell(row=r, column=1).fill = fill('total_gray')
    ws.cell(row=r, column=1).border = THIN_BORDER
    for c in range(2, 5):
        col_l = get_column_letter(c)
        ws.cell(row=r, column=c, value=f'=IF({col_l}{rev_start}=0,0,{col_l}{rev_start+1}/{col_l}{rev_start})').number_format = '0.00%'
        ws.cell(row=r, column=c).fill = fill('total_gray')
        ws.cell(row=r, column=c).border = THIN_BORDER
    r += 2

    # Retailer Activation Funnel
    write_section_header(ws, r, "RETAILER ACTIVATION FUNNEL", max_col); r += 1
    funnel_headers = ['Metric', 'Helix', 'JSW ONE', '', '', '', '', '']
    write_row(ws, r, funnel_headers, font=HEADER_FONT, color_key='header_light'); r += 1

    funnel = [
        ('Total Retailers (Mar 27)', refs['helix_active_ret'], refs['jsw_active_ret']),
        ('Transacting Retailers', refs['helix_transacting'], refs['jsw_transacting']),
        ('Conversion Rate', None, None),
    ]
    funnel_start = r
    for i, (label, h_ref, j_ref) in enumerate(funnel):
        row = funnel_start + i
        ws.cell(row=row, column=1, value=label).font = BOLD_FONT
        ws.cell(row=row, column=1).border = THIN_BORDER
        if h_ref:
            ws.cell(row=row, column=2, value=f'={h_ref}').number_format = '#,##0'
        if j_ref:
            ws.cell(row=row, column=3, value=f'={j_ref}').number_format = '#,##0'
        ws.cell(row=row, column=2).border = THIN_BORDER
        ws.cell(row=row, column=3).border = THIN_BORDER

    # Conversion rate
    conv_row = funnel_start + 2
    ws.cell(row=conv_row, column=2, value=f'=IF(B{funnel_start}=0,0,B{funnel_start+1}/B{funnel_start})').number_format = '0.0%'
    ws.cell(row=conv_row, column=3, value=f'=IF(C{funnel_start}=0,0,C{funnel_start+1}/C{funnel_start})').number_format = '0.0%'

    return {'helix_vol_row': helix_vol_row, 'jsw_vol_row': jsw_vol_row,
            'combined_row': combined_row, 'rev_start': rev_start}


# ============================================================
# TAB 9: ROI & METRICS TRACKER
# ============================================================
def build_tab9(wb, refs, cons_info, vol_info):
    ws = wb.create_sheet("ROI Metrics")
    ws.sheet_properties.tabColor = "C00000"
    max_col = 7
    set_col_widths(ws, [30, 15, 15, 15, 15, 15, 15])

    # Title
    ws.merge_cells(f'A1:{get_column_letter(max_col)}1')
    ws.cell(row=1, column=1, value="ROI & METRICS TRACKER (FY 2026-27)").font = HEADER_FONT_LG
    ws.cell(row=1, column=1).fill = fill('header_dark')
    for c in range(2, max_col + 1):
        ws.cell(row=1, column=c).fill = fill('header_dark')

    r = 3
    # Scenario Analysis
    write_section_header(ws, r, "SCENARIO ANALYSIS - BUDGET SENSITIVITY", max_col); r += 1
    sc_headers = ['Scenario', 'Budget (Rs.)', 'Helix Alloc', 'JSW ONE Alloc',
                  'Cost/MT (Helix)', 'Cost/MT (JSW ONE)', 'Blended Cost/MT']
    write_row(ws, r, sc_headers, font=HEADER_FONT, color_key='header_light'); r += 1

    scenarios = [
        ('Base Case (Current)', 1.0),
        ('-20% Budget', 0.8),
        ('-10% Budget', 0.9),
        ('+10% Budget', 1.1),
        ('+20% Budget', 1.2),
    ]
    sc_start = r
    for i, (label, mult) in enumerate(scenarios):
        row = sc_start + i
        ws.cell(row=row, column=1, value=label).font = NORMAL_FONT
        ws.cell(row=row, column=1).border = THIN_BORDER
        ws.cell(row=row, column=2, value=f'={refs["total_budget"]}*{mult}').number_format = '#,##0'
        ws.cell(row=row, column=2).border = THIN_BORDER
        ws.cell(row=row, column=3, value=f'=B{row}*{refs["helix_pct"]}*(1-{refs["contingency_pct"]})').number_format = '#,##0'
        ws.cell(row=row, column=3).border = THIN_BORDER
        ws.cell(row=row, column=4, value=f'=B{row}*{refs["obj2_pct"]}*(1-{refs["contingency_pct"]})').number_format = '#,##0'
        ws.cell(row=row, column=4).border = THIN_BORDER

        h_vol = vol_info['helix_vol_row']
        j_vol = vol_info['jsw_vol_row']
        ws.cell(row=row, column=5,
                value=f"=IF('Vol Projections'!F{h_vol}=0,0,C{row}/'Vol Projections'!F{h_vol})").number_format = '#,##0'
        ws.cell(row=row, column=5).border = THIN_BORDER
        ws.cell(row=row, column=6,
                value=f"=IF('Vol Projections'!F{j_vol}=0,0,D{row}/'Vol Projections'!F{j_vol})").number_format = '#,##0'
        ws.cell(row=row, column=6).border = THIN_BORDER
        comb = vol_info['combined_row']
        ws.cell(row=row, column=7,
                value=f"=IF('Vol Projections'!F{comb}=0,0,B{row}/'Vol Projections'!F{comb})").number_format = '#,##0'
        ws.cell(row=row, column=7).border = THIN_BORDER

    r = sc_start + len(scenarios) + 1

    # Efficiency Metrics
    write_section_header(ws, r, "EFFICIENCY METRICS", max_col); r += 1
    eff_headers = ['Metric', 'Helix', 'JSW ONE (2A)', 'JSW ONE (2B)', 'Combined', '', '']
    write_row(ws, r, eff_headers, font=HEADER_FONT, color_key='header_light'); r += 1

    eff_start = r
    eff_items = [
        'Cost per Dealer Activated',
        'Cost per Retailer Engaged',
        'Marketing Spend per Distributor',
        'ATL Spend per Priority District',
    ]
    for i, item in enumerate(eff_items):
        row = eff_start + i
        ws.cell(row=row, column=1, value=item).font = NORMAL_FONT
        ws.cell(row=row, column=1).border = THIN_BORDER
        for c in range(2, 6):
            ws.cell(row=row, column=c).border = THIN_BORDER
            ws.cell(row=row, column=c).number_format = '#,##0'

    # Cost per dealer
    ws.cell(row=eff_start, column=2,
            value=f"=IF({refs['helix_active_ret']}=0,0,'Consolidated'!B{cons_info['total_row']}/{refs['helix_active_ret']})").number_format = '#,##0'
    ws.cell(row=eff_start, column=3,
            value=f"=IF({refs['jsw_active_ret']}*{refs['jsw_east_vol_pct']}=0,0,'Consolidated'!C{cons_info['total_row']}/({refs['jsw_active_ret']}*{refs['jsw_east_vol_pct']}))").number_format = '#,##0'
    ws.cell(row=eff_start, column=4,
            value=f"=IF({refs['jsw_active_ret']}*{refs['jsw_uphr_vol_pct']}=0,0,'Consolidated'!D{cons_info['total_row']}/({refs['jsw_active_ret']}*{refs['jsw_uphr_vol_pct']}))").number_format = '#,##0'

    # Cost per retailer
    ws.cell(row=eff_start + 1, column=2,
            value=f"=IF({refs['helix_transacting']}=0,0,'Consolidated'!B{cons_info['total_row']}/{refs['helix_transacting']})").number_format = '#,##0'

    # Spend per distributor
    ws.cell(row=eff_start + 2, column=2,
            value=f"=IF({refs['helix_distributors']}=0,0,'Consolidated'!B{cons_info['total_row']}/{refs['helix_distributors']})").number_format = '#,##0'
    ws.cell(row=eff_start + 2, column=3,
            value=f"=IF({refs['jsw_distributors']}*{refs['jsw_east_vol_pct']}=0,0,'Consolidated'!C{cons_info['total_row']}/({refs['jsw_distributors']}*{refs['jsw_east_vol_pct']}))").number_format = '#,##0'

    r = eff_start + len(eff_items) + 1

    # Objective-wise ROI comparison
    write_section_header(ws, r, "OBJECTIVE-WISE COMPARISON", max_col); r += 1
    comp_headers = ['Metric', 'Obj 1: Helix', 'Obj 2A: East', 'Obj 2B: UP+HR', 'Total', '', '']
    write_row(ws, r, comp_headers, font=HEADER_FONT, color_key='header_light'); r += 1

    comp_items = [
        ('Budget Allocated (Rs.)', refs['helix_budget'], refs['sub2a_budget'], refs['sub2b_budget']),
        ('Total Spend (Rs.)',
         f"'Consolidated'!B{cons_info['total_row']}",
         f"'Consolidated'!C{cons_info['total_row']}",
         f"'Consolidated'!D{cons_info['total_row']}"),
        ('Budget Utilization %', None, None, None),
    ]
    comp_start = r
    for i, (label, *formulas) in enumerate(comp_items):
        row = comp_start + i
        ws.cell(row=row, column=1, value=label).font = BOLD_FONT
        ws.cell(row=row, column=1).border = THIN_BORDER
        if formulas[0]:
            for c, f in enumerate(formulas, 2):
                ws.cell(row=row, column=c, value=f'={f}').number_format = '#,##0'
                ws.cell(row=row, column=c).border = THIN_BORDER
            ws.cell(row=row, column=5, value=f'=B{row}+C{row}+D{row}').number_format = '#,##0'
            ws.cell(row=row, column=5).border = THIN_BORDER
        else:
            for c in range(2, 6):
                ws.cell(row=row, column=c).border = THIN_BORDER

    # Utilization %
    util_row = comp_start + 2
    for c in range(2, 6):
        col_l = get_column_letter(c)
        ws.cell(row=util_row, column=c,
                value=f'=IF({col_l}{comp_start}=0,0,{col_l}{comp_start+1}/{col_l}{comp_start})').number_format = '0.0%'

    return ws


# ============================================================
# TAB 10: QUARTERLY REVIEW TRACKER
# ============================================================
def build_tab10(wb, refs, cons_info):
    ws = wb.create_sheet("Quarterly Review")
    ws.sheet_properties.tabColor = "808080"
    max_col = 7
    set_col_widths(ws, [25, 18, 18, 18, 18, 18, 25])

    # Title
    ws.merge_cells(f'A1:{get_column_letter(max_col)}1')
    ws.cell(row=1, column=1, value="QUARTERLY REVIEW TRACKER - ACTUAL vs PLAN (FY 2026-27)").font = HEADER_FONT_LG
    ws.cell(row=1, column=1).fill = fill('header_dark')
    for c in range(2, max_col + 1):
        ws.cell(row=1, column=c).fill = fill('header_dark')

    ws.cell(row=2, column=1, value="Enter actuals in yellow cells after each quarter. Variance auto-calculates.").font = SMALL_FONT
    ws.merge_cells('A2:G2')

    r = 4
    q_start = cons_info['q_start_row']

    for qi in range(4):
        write_section_header(ws, r, f"{QUARTERS[qi]} REVIEW ({MONTHS[qi*3]}-{MONTHS[qi*3+2]} 2026-27)", max_col)
        r += 1

        review_headers = ['Metric', 'Obj1: Helix', 'Obj2A: East', 'Obj2B: UP+HR', 'Total', 'Variance', 'Remarks']
        write_row(ws, r, review_headers, font=HEADER_FONT, color_key='header_light')
        r += 1

        # Planned row
        ws.cell(row=r, column=1, value='Planned Spend (Rs.)').font = BOLD_FONT
        ws.cell(row=r, column=1).border = THIN_BORDER
        for c in range(2, 5):
            col_l = get_column_letter(c)
            ws.cell(row=r, column=c,
                    value=f"='Consolidated'!{col_l}{q_start+qi}").number_format = '#,##0'
            ws.cell(row=r, column=c).fill = fill('calc_green')
            ws.cell(row=r, column=c).border = THIN_BORDER
        ws.cell(row=r, column=5, value=f'=SUM(B{r}:D{r})').number_format = '#,##0'
        ws.cell(row=r, column=5).border = THIN_BORDER
        planned_row = r
        r += 1

        # Actual row
        ws.cell(row=r, column=1, value='Actual Spend (Rs.)').font = BOLD_FONT
        ws.cell(row=r, column=1).border = THIN_BORDER
        for c in range(2, 5):
            ws.cell(row=r, column=c, value=0).number_format = '#,##0'
            ws.cell(row=r, column=c).fill = fill('input_yellow')
            ws.cell(row=r, column=c).border = THIN_BORDER
            ws.cell(row=r, column=c).protection = Protection(locked=False)
        ws.cell(row=r, column=5, value=f'=SUM(B{r}:D{r})').number_format = '#,##0'
        ws.cell(row=r, column=5).border = THIN_BORDER
        actual_row = r
        r += 1

        # Variance
        ws.cell(row=r, column=1, value='Variance (Plan - Actual)').font = BOLD_FONT
        ws.cell(row=r, column=1).border = THIN_BORDER
        for c in range(2, 6):
            col_l = get_column_letter(c)
            ws.cell(row=r, column=c,
                    value=f'={col_l}{planned_row}-{col_l}{actual_row}').number_format = '#,##0'
            ws.cell(row=r, column=c).border = THIN_BORDER
        # Conditional formatting on variance
        ws.conditional_formatting.add(f'B{r}:E{r}',
            CellIsRule(operator='lessThan', formula=['0'],
                       fill=PatternFill(fgColor='FFC7CE'), font=Font(color='9C0006', bold=True)))
        r += 1

        # Variance %
        ws.cell(row=r, column=1, value='Variance %').font = BOLD_FONT
        ws.cell(row=r, column=1).border = THIN_BORDER
        for c in range(2, 6):
            col_l = get_column_letter(c)
            ws.cell(row=r, column=c,
                    value=f'=IF({col_l}{planned_row}=0,0,{col_l}{r-1}/{col_l}{planned_row})').number_format = '0.0%'
            ws.cell(row=r, column=c).border = THIN_BORDER
        # Flag >15% deviation
        ws.conditional_formatting.add(f'B{r}:E{r}',
            FormulaRule(formula=[f'ABS(B{r})>0.15'],
                        fill=PatternFill(fgColor='FFC7CE'), font=Font(color='9C0006', bold=True)))
        r += 1

        # Observations
        ws.cell(row=r, column=1, value='Key Observations').font = NORMAL_FONT
        ws.cell(row=r, column=1).border = THIN_BORDER
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=7)
        ws.cell(row=r, column=2).fill = fill('input_yellow')
        ws.cell(row=r, column=2).border = THIN_BORDER
        ws.cell(row=r, column=2).protection = Protection(locked=False)
        r += 2

    # Annual Summary
    write_section_header(ws, r, "ANNUAL SUMMARY", max_col); r += 1
    ws.cell(row=r, column=1, value='Total Planned').font = BOLD_FONT
    ws.cell(row=r, column=1).border = THIN_BORDER
    ws.cell(row=r, column=5, value=f"='Consolidated'!F{cons_info['total_row']}").number_format = '#,##0'
    ws.cell(row=r, column=5).border = THIN_BORDER

    return ws


# ============================================================
# TAB 1: EXECUTIVE DASHBOARD
# ============================================================
def build_tab1(wb, refs, cons_info, vol_info):
    ws = wb.create_sheet("Dashboard")
    ws.sheet_properties.tabColor = "002060"
    max_col = 12
    set_col_widths(ws, [3, 20, 15, 15, 15, 15, 3, 20, 15, 15, 15, 15])

    # Title banner
    ws.merge_cells('A1:L1')
    c = ws.cell(row=1, column=1, value="JSW STEEL TMT - MARKETING BUDGET MODEL FY 2026-27")
    c.font = Font(name='Calibri', bold=True, size=16, color='FFFFFF')
    c.fill = fill('header_dark')
    c.alignment = Alignment(horizontal='center', vertical='center')
    for col in range(2, max_col + 1):
        ws.cell(row=1, column=col).fill = fill('header_dark')
    ws.row_dimensions[1].height = 40

    ws.merge_cells('A2:L2')
    ws.cell(row=2, column=1, value="Executive Summary Dashboard | Total Budget: Rs. 5 Crore | Period: Apr 2026 - Mar 2027").font = Font(name='Calibri', size=10, italic=True, color='FFFFFF')
    ws.cell(row=2, column=1).fill = fill('header_med')
    ws.cell(row=2, column=1).alignment = Alignment(horizontal='center')
    for col in range(2, max_col + 1):
        ws.cell(row=2, column=col).fill = fill('header_med')

    r = 4
    # KPI Cards
    kpi_data = [
        ('Total Budget', f'={refs["total_budget"]}', '#,##0', 'B'),
        ('Helix Budget', f'={refs["helix_budget"]}', '#,##0', 'D'),
        ('JSW ONE Budget', f'={refs["obj2_budget"]}', '#,##0', 'F'),
        ('Blended Cost/MT', f"='Vol Projections'!F{vol_info['combined_row']-1}", '#,##0', 'H'),
        ('Total Volume (MT)', f"='Vol Projections'!F{vol_info['combined_row']}", '#,##0', 'J'),
    ]
    for i, (label, formula, fmt, col_start) in enumerate(kpi_data):
        col_idx = 2 + i * 2
        ws.cell(row=r, column=col_idx, value=label).font = Font(name='Calibri', size=9, color='1F4E79')
        ws.cell(row=r, column=col_idx).alignment = Alignment(horizontal='center')
        ws.cell(row=r, column=col_idx).fill = fill('output_blue')
        ws.cell(row=r, column=col_idx).border = THIN_BORDER
        ws.cell(row=r + 1, column=col_idx, value=formula).number_format = fmt
        ws.cell(row=r + 1, column=col_idx).font = Font(name='Calibri', bold=True, size=14, color='1F4E79')
        ws.cell(row=r + 1, column=col_idx).alignment = Alignment(horizontal='center')
        ws.cell(row=r + 1, column=col_idx).fill = fill('output_blue')
        ws.cell(row=r + 1, column=col_idx).border = THIN_BORDER

    r = 7
    # Budget Allocation Table
    write_section_header(ws, r, "BUDGET ALLOCATION SUMMARY", max_col); r += 1

    alloc_headers = ['', 'Objective', 'Budget (Rs.)', '% of Total', 'ATL Spend', 'BTL Spend', '',
                     'Metric', 'Helix', 'JSW ONE', 'Combined', '']
    write_row(ws, r, alloc_headers, font=HEADER_FONT, color_key='header_light'); r += 1

    # Left side: Budget allocation
    alloc = [
        ('Obj 1: Helix Launch', refs['helix_budget'], refs['helix_pct']),
        ('Obj 2A: JSW ONE East', refs['sub2a_budget'], f'{refs["obj2_pct"]}*{refs["sub2a_pct"]}'),
        ('Obj 2B: JSW ONE UP+HR', refs['sub2b_budget'], f'{refs["obj2_pct"]}*{refs["sub2b_pct"]}'),
        ('Contingency Reserve', refs['contingency_amt'], refs['contingency_pct']),
    ]
    alloc_start = r
    for i, (label, budget_f, pct_f) in enumerate(alloc):
        row = alloc_start + i
        ws.cell(row=row, column=1, value=i + 1).font = NORMAL_FONT
        ws.cell(row=row, column=1).border = THIN_BORDER
        ws.cell(row=row, column=2, value=label).font = BOLD_FONT
        ws.cell(row=row, column=2).border = THIN_BORDER
        ws.cell(row=row, column=3, value=f'={budget_f}').number_format = '#,##0'
        ws.cell(row=row, column=3).border = THIN_BORDER
        ws.cell(row=row, column=4, value=f'={pct_f}').number_format = '0.0%'
        ws.cell(row=row, column=4).border = THIN_BORDER

    # Right side: Key metrics
    metrics = [
        ('A/S Ratio (Helix)', f"='Vol Projections'!B{vol_info['rev_start']+2}", '0.00%'),
        ('A/S Ratio (JSW ONE)', f"='Vol Projections'!C{vol_info['rev_start']+2}", '0.00%'),
        ('Helix Volume (MT)', f"='Vol Projections'!F{vol_info['helix_vol_row']}", '#,##0'),
        ('JSW ONE Volume (MT)', f"='Vol Projections'!F{vol_info['jsw_vol_row']}", '#,##0'),
    ]
    for i, (label, formula, fmt) in enumerate(metrics):
        row = alloc_start + i
        ws.cell(row=row, column=8, value=label).font = BOLD_FONT
        ws.cell(row=row, column=8).border = THIN_BORDER
        ws.cell(row=row, column=9, value=formula).number_format = fmt
        ws.cell(row=row, column=9).border = THIN_BORDER

    r = alloc_start + len(alloc) + 1

    # Total row
    ws.cell(row=r, column=2, value='TOTAL').font = BOLD_FONT
    ws.cell(row=r, column=2).fill = fill('total_gray')
    ws.cell(row=r, column=2).border = THIN_BORDER
    ws.cell(row=r, column=3, value=f'={refs["total_budget"]}').number_format = '#,##0'
    ws.cell(row=r, column=3).fill = fill('total_gray')
    ws.cell(row=r, column=3).border = THIN_BORDER
    ws.cell(row=r, column=3).font = BOLD_FONT
    ws.cell(row=r, column=4, value='100.0%').number_format = '0.0%'
    ws.cell(row=r, column=4).fill = fill('total_gray')
    ws.cell(row=r, column=4).border = THIN_BORDER

    r += 2

    # Quarterly Spend Chart Data
    write_section_header(ws, r, "QUARTERLY SPEND TRAJECTORY", max_col); r += 1
    q_headers = ['', 'Quarter', 'Helix', 'JSW ONE East', 'JSW ONE UP+HR', 'Total', '', '', '', '', '', '']
    write_row(ws, r, q_headers, font=HEADER_FONT, color_key='header_light'); r += 1

    q_data_start = r
    q_start = cons_info['q_start_row']
    for qi in range(4):
        row = q_data_start + qi
        ws.cell(row=row, column=2, value=QUARTERS[qi]).font = BOLD_FONT
        ws.cell(row=row, column=2).border = THIN_BORDER
        ws.cell(row=row, column=3, value=f"='Consolidated'!B{q_start+qi}").number_format = '#,##0'
        ws.cell(row=row, column=3).border = THIN_BORDER
        ws.cell(row=row, column=4, value=f"='Consolidated'!C{q_start+qi}").number_format = '#,##0'
        ws.cell(row=row, column=4).border = THIN_BORDER
        ws.cell(row=row, column=5, value=f"='Consolidated'!D{q_start+qi}").number_format = '#,##0'
        ws.cell(row=row, column=5).border = THIN_BORDER
        ws.cell(row=row, column=6, value=f"='Consolidated'!F{q_start+qi}").number_format = '#,##0'
        ws.cell(row=row, column=6).border = THIN_BORDER
        ws.cell(row=row, column=6).font = BOLD_FONT

    # Create bar chart for quarterly spend
    chart = BarChart()
    chart.type = "col"
    chart.title = "Quarterly Marketing Spend by Objective"
    chart.y_axis.title = "Spend (Rs.)"
    chart.x_axis.title = "Quarter"
    chart.style = 10
    chart.width = 20
    chart.height = 12

    cats = Reference(ws, min_col=2, min_row=q_data_start, max_row=q_data_start + 3)
    for ci, label in enumerate(['Helix', 'JSW ONE East', 'JSW ONE UP+HR'], 3):
        data = Reference(ws, min_col=ci, min_row=q_data_start - 1, max_row=q_data_start + 3)
        chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.shape = 4

    r = q_data_start + 5
    ws.add_chart(chart, f'B{r}')

    r += 16

    # Instructions
    write_section_header(ws, r, "HOW TO USE THIS MODEL", max_col); r += 1
    instructions = [
        "1. Go to 'Master Inputs' tab to change budget allocation, quarterly splits, and volume targets.",
        "2. Yellow cells are editable - all other cells contain formulas that auto-update.",
        "3. Change Helix vs JSW ONE split in Budget Allocation Controls section.",
        "4. Adjust ATL/BTL mix in the Helix ATL vs BTL Split section.",
        "5. Modify quarterly distribution in the Quarterly Allocation section.",
        "6. Update unit costs in 'Activity Costs' tab to reflect market rates.",
        "7. Track actuals in 'Quarterly Review' tab after each quarter.",
        "8. Color Legend: YELLOW = Input | GREEN = Calculated | GRAY = Totals | RED = Alerts",
    ]
    for inst in instructions:
        ws.cell(row=r, column=2, value=inst).font = NORMAL_FONT
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=max_col)
        r += 1

    return ws


# ============================================================
# MAIN BUILD FUNCTION
# ============================================================
def main():
    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    print("Building Tab 8: Activity Cost Reference...")
    build_tab8(wb)

    print("Building Tab 2: Master Inputs...")
    refs = build_tab2(wb)

    print("Building Tab 3: Objective 1 - Helix Launch...")
    obj1_info = build_tab3(wb, refs)

    print("Building Tab 4: Objective 2A - JSW ONE East Growth...")
    obj2a_info = build_tab4(wb, refs)

    print("Building Tab 5: Objective 2B - JSW ONE UP & Haryana...")
    obj2b_info = build_tab5(wb, refs)

    print("Building Tab 6: Consolidated Budget...")
    cons_info = build_tab6(wb, refs, obj1_info, obj2a_info, obj2b_info)

    print("Building Tab 7: Volume Projections...")
    vol_info = build_tab7(wb, refs, cons_info, obj1_info, obj2a_info, obj2b_info)

    print("Building Tab 9: ROI & Metrics Tracker...")
    build_tab9(wb, refs, cons_info, vol_info)

    print("Building Tab 10: Quarterly Review Tracker...")
    build_tab10(wb, refs, cons_info)

    print("Building Tab 1: Executive Dashboard...")
    build_tab1(wb, refs, cons_info, vol_info)

    # Reorder sheets: Dashboard first
    sheet_order = ['Dashboard', 'Master Inputs', 'Obj1 Helix', 'Obj2A East',
                   'Obj2B UPHR', 'Consolidated', 'Vol Projections',
                   'Activity Costs', 'ROI Metrics', 'Quarterly Review']
    for i, name in enumerate(sheet_order):
        idx = wb.sheetnames.index(name)
        wb.move_sheet(name, offset=i - idx)

    # Save
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    wb.save(OUTPUT_PATH)
    print(f"\nModel saved to: {OUTPUT_PATH}")
    print("Done! Open the Excel file and verify formulas calculate correctly.")


if __name__ == '__main__':
    main()
