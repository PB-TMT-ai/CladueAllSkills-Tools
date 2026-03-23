"""
TMT Incentive Extraction - JSW Steel Annual Incentive Scheme FY 2025-26
Executive document-style Excel output for top management.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─── Configuration ───────────────────────────────────────────────────────────

OUTPUT_FILE = r"D:\MOU Slabs\TMT_Incentives_FY2025-26.xlsx"
SOURCE_NOTE = "Source: JSW Steel Annual Incentive Scheme FY 2025-26, dated 5th May 2025"
NUM_COLS = 9  # A through I

# ─── Styles ──────────────────────────────────────────────────────────────────

FONT_TITLE = Font(name="Arial", size=14, bold=True, color="FFFFFF")
FONT_SUBTITLE = Font(name="Arial", size=11, bold=True)
FONT_SECTION = Font(name="Arial", size=12, bold=True, color="FFFFFF")
FONT_SUBSECTION = Font(name="Arial", size=11, bold=True)
FONT_COL_HEADER = Font(name="Arial", size=10, bold=True)
FONT_DATA = Font(name="Arial", size=10)
FONT_DATA_CENTER = Font(name="Arial", size=10)
FONT_NA = Font(name="Arial", size=10, italic=True, color="555555")
FONT_FOOTER = Font(name="Arial", size=8, italic=True, color="888888")
FONT_TC_TITLE = Font(name="Arial", size=13, bold=True, color="FFFFFF")
FONT_TC_CLAUSE = Font(name="Arial", size=10)

FILL_DARK_BLUE = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
FILL_LIGHT_BLUE = PatternFill(start_color="B8CCE4", end_color="B8CCE4", fill_type="solid")
FILL_GRAY = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
FILL_LIGHT_GRAY = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

ALIGN_LEFT_WRAP = Alignment(horizontal="left", vertical="top", wrap_text=True)
ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_CENTER_TOP = Alignment(horizontal="center", vertical="top", wrap_text=True)
ALIGN_LEFT_CENTER = Alignment(horizontal="left", vertical="center", wrap_text=True)

BORDER_THIN = Border(
    left=Side(style="thin", color="C0C0C0"),
    right=Side(style="thin", color="C0C0C0"),
    top=Side(style="thin", color="C0C0C0"),
    bottom=Side(style="thin", color="C0C0C0"),
)

BORDER_NONE = Border()

# ─── Data ────────────────────────────────────────────────────────────────────

TMT_RETAIL_SLABS = [
    ("\u22651,200 - <1,800", 150),
    ("\u22651,800 - <2,400", 175),
    ("\u22652,400 - <3,600", 200),
    ("\u22653,600 - <4,800", 225),
    ("\u22654,800 - <6,000", 250),
    ("\u22656,000 - <9,000", 275),
    ("\u22659,000 - <12,000", 300),
    ("\u226512,000 - <18,000", 325),
    ("\u226518,000 - <27,000", 350),
    ("\u226527,000 - <36,000", 375),
    ("\u226536,000 - <48,000", 425),
    ("\u226548,000 - <60,000", 450),
    ("\u226560,000", 475),
]

TMT_OEM_SLABS = [
    ("\u22653,600 - <4,800", 125),
    ("\u22654,800 - <6,000", 150),
    ("\u22656,000 - <9,000", 175),
    ("\u22659,000 - <12,000", 200),
    ("\u226512,000 - <18,000", 225),
    ("\u226518,000 - <27,000", 250),
    ("\u226527,000 - <36,000", 275),
    ("\u226536,000 - <48,000", 300),
    ("\u226548,000 - <60,000", 325),
    ("\u226560,000", 375),
]

# Qualification texts
VOL_PAYOUT_FREQ = "Half Yearly and Annually"

VOL_QUAL = (
    "Min 20% quarterly prorate signed volume with exception of one quarter "
    "(within first 3 quarters) where customer can lift min 18%. "
    "No exception in Quarter 4. Volume Incentive applicable on original "
    "signed/enhanced quantity with rate remaining same as per original quantity."
)

VOL_HY_RULE = (
    "Min 50% of Volume compliance up to 30th Sept 2025 (H1 Period); "
    "paid 80% of volume incentive rate after completion of H1."
)

VOL_YEARLY_RULE = (
    "100% achievement of signed quantity required. "
    "Differential of Volume Incentive payable at year end after "
    "completion of 100% signed volume for product family."
)

VOL_MONTHLY = (
    "Min 20% quarterly prorate; exception of 1 quarter "
    "(first 3 Qs) where min 18% allowed; no exception in Q4."
)

VOL_CAP = "120% of signed quantity"

CONSIST_QUAL = (
    "22.5% quarterly pro-rata achievement of signed quantity for each product "
    "and 21.25% with lean period quarter. Monthly achievement shall be min 7.5% "
    "for each month and 6.25% for lean month period."
)

CONSIST_MONTHLY = (
    "Min 7.5% monthly for each month; 6.25% for lean period month."
)

CONSIST_LEAN = (
    "Max 2 non-consecutive months during the year and max 1 month during any quarter."
)

CONSIST_ENHANCEMENT = (
    "Applicable on Original quantity before enhancement (Q1 & Q2). "
    "Post enhancement, qualification as per enhanced quantity for Q3 and Q4."
)

LOYALTY_QUAL = (
    "FY 25-26 Signing \u2265 110% of FY 24-25 Signed/Amended Volume (whichever is higher); "
    "AND Volume achievement of 100% for FY 24-25."
)

LOYALTY_YEARLY = "Payout is yearly and subjected to 100% achievement of signed for FY 25-26."

MSME_QUAL = (
    "UDYAM registration (latest) certificate or Certificate of Udyog Aadhar No. "
    "(Applicable for Non-Retail Customers). Customer onboarding in MSME distribution "
    "channel (MS) is continual process, onboarding every quarter."
)

MSME_PAYOUT = (
    "Payable yearly on completion of 100% of Overall Signed Volume (not product specific). "
    "Latest UDYAM certificate to be updated before payout."
)

MSME_CAP = "Upper cap of 12,000 MT annually; sales including OE/MS as distribution channel considered."

# Terms & Conditions
TERMS_AND_CONDITIONS = [
    "The applicable Incentive Scheme shall be signed by JSW Group and its Buyers/submitted on or before 30th June 2025.",
    "The Incentive Scheme shall be signed for single product / multiple product family(s) (collectively referred as \"Products\") as stated in column no. 2 and 3 in Table-1 above.",
    "The volume based Incentive shall apply on the quantities as stated under Table-3 in multiples of 100 MT.",
    "There shall be single Incentive Scheme document for each Primary Buyer including the Buyer Group companies, which shall apply to the combined quantity, multi-locational units and/or approved sister concerns. In such case, quantities lifted by all units/approved sister concerns, Buyer Group entities would qualify the Incentive Scheme.",
    "The incentives shall be paid within 30 days, subject to fulfilment of the terms and conditions, qualification criteria applicable to the said Incentive Scheme.",
    "The incentives shall be applicable for total Prime / NCO (including non-Prime) sales and shall exclude quantities sold under auctions and material returned as well as under complaints / disputes. S1 category (non-Prime) of material from \"Electrical\" Product Family shall not qualify incentive scheme.",
    "JSW Group Company reserves the right of supplying Products from any of its work locations and/or stockyards and/or job work/Conversion-agent and/or service centers and/or Consignment agents and/or affiliated and/or joint venture and/or associate companies.",
    "Dispatch by Rail, Road, Sea or Multi-Modal transportation shall be at the option of JSW group companies.",
    "Each Primary Buyer / Buyer group shall have unique distribution channel i.e. either OEM (including industrial customers) or Retail. Distribution channel(s) i.e. OE / DE / SA / MS (MSME) shall be considered as OEM (including industrial customers) while only RE shall be considered as Retail. These distribution channels are permanent and not interchangeable during the entire period of this Incentive Scheme. Multiple Distribution channel could not be part of single MOU Group.",
    "In case of \"Coated\" Products family all product groups are separate from each other. Refer Table-1 for definition of Product Groups C1, C2, C3.1, C3.2, C3.3 and C4.",
    "The eligibility of the Primary Buyer for the incentives shall be based on Sold to party and applicable payee to payer.",
    "Maximum Payment of all incentives shall be limited to total quantity lifted, with a cap of 120%, of the applicable signed quantity for all product family.",
    "Till FY 24-25 Incentive scheme gets rolled out draft provisions shall be on the basis of FY24-25 scheme. These provisions shall be reversed post implementation of new scheme for FY 25-26.",
    "Original signed quantity can be increased, i.e., Enhancement subject to mutual formal agreement after completion of Quarter 2 i.e. in between 1st - 15th October 2025. No change in incentive rates / addition of new product family to be part of enhancement is not allowed.",
    "Maximum payout shall be limited to total quantity lifted, with a cap of 120%, of the Pro-rata signed original / enhanced quantity for all incentive unless mentioned.",
    "In case the Buyer or any Buyer Group is found indulged in corrupt, fraudulent, collusive, unethical practices or misrepresentation, misconduct, negligent, indiscipline or unruly behavior and conduct, which is detrimental to the business interest, security, safety or reputation of JSW or any JSW Group Company, then, JSW or such JSW Group Company shall immediately terminate, withdraw, rescind from the Incentive Scheme extended to the defaulting Primary Buyer or their entire Buyer Group, without any prior notice. All associated unpaid incentives shall also stand forfeited. Delay in payment may attract interest at the rate of 18% per annum.",
    "JSW Group reserves the right to withdraw, terminate, cancel any Incentive Scheme signed pursuant to this Annual Incentive Scheme, at its sole discretion, with prior written notice of 7 days to Primary Buyer. Any notice issued to Primary Buyer shall be deemed to be notice to the Buyer Group.",
    "JSW Group Company shall have a right to terminate any Incentive Scheme, in the event of any default or breach committed by the Primary Buyer or any entity of Buyer Group of the terms and conditions of the Distribution Agreement and/or the Incentive Scheme as the case maybe, including the Primary Buyer or any entity of Buyer Group being adjudged as bankrupt or a resolution/order being passed for dissolution of the firm / Commencement of proceedings under Insolvency and Bankruptcy Code.",
    "In case of termination / withdrawal of Incentive Scheme for Primary Buyer, all incentives, benefits extended to Buyer Group under the Incentive Scheme shall also stand withdrawn immediately. JSW Group Company shall be entitled to recover such amount by deducting in part or in whole from any sum payable or thereafter becoming payable to the Primary Buyer or Buyer Group by any JSW Group Company, for any set-off or counterclaim.",
    "In the event that the amount so deducted by the relevant JSW Group Company is not sufficient to cover the full amount recoverable by JSW or JSW Group Company, the Primary Buyer or Buyer Group shall on demand make immediate payment of such remaining amount. Both Primary Buyer and each entity of Buyer Group shall be jointly and severally responsible and liable for any recovery under the Incentive Scheme.",
    "This Incentive Scheme shall be interpreted, construed and governed by Laws of India.",
]


# ─── Helper Functions ────────────────────────────────────────────────────────

def merge_write(ws, row, col_start, col_end, value, font, fill=None, alignment=None, border=None, row_height=None):
    """Write a value into a merged cell range and apply styling."""
    if col_start != col_end:
        ws.merge_cells(start_row=row, start_column=col_start, end_row=row, end_column=col_end)
    cell = ws.cell(row=row, column=col_start, value=value)
    cell.font = font
    if fill:
        cell.fill = fill
    cell.alignment = alignment or ALIGN_LEFT_WRAP
    if border:
        cell.border = border
    # Apply fill/border to all cells in merge range
    for c in range(col_start, col_end + 1):
        mcell = ws.cell(row=row, column=c)
        if fill:
            mcell.fill = fill
        if border:
            mcell.border = border
    if row_height:
        ws.row_dimensions[row].height = row_height
    return row


def write_section_header(ws, row, text):
    """Write a dark blue section header spanning all columns."""
    merge_write(ws, row, 1, NUM_COLS, text, FONT_SECTION, FILL_DARK_BLUE, ALIGN_LEFT_CENTER, BORDER_THIN, 30)
    return row + 1


def write_subsection_header(ws, row, text):
    """Write a light blue subsection header spanning all columns."""
    merge_write(ws, row, 1, NUM_COLS, text, FONT_SUBSECTION, FILL_LIGHT_BLUE, ALIGN_LEFT_CENTER, BORDER_THIN, 25)
    return row + 1


def write_na_row(ws, row, text):
    """Write a 'Not Applicable' merged italic row."""
    merge_write(ws, row, 1, NUM_COLS, text, FONT_NA, FILL_LIGHT_GRAY, ALIGN_LEFT_WRAP, BORDER_THIN, 22)
    return row + 1


def write_col_headers(ws, row, headers):
    """Write column headers with gray background."""
    for col_idx, header_text in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_idx, value=header_text)
        cell.font = FONT_COL_HEADER
        cell.fill = FILL_GRAY
        cell.alignment = ALIGN_CENTER
        cell.border = BORDER_THIN
    return row + 1


def write_data_cell(ws, row, col, value, center=False):
    """Write a single data cell."""
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = FONT_DATA
    cell.alignment = ALIGN_CENTER if center else ALIGN_LEFT_WRAP
    cell.border = BORDER_THIN
    return cell


# ─── Volume Incentive Section Builder ────────────────────────────────────────

def write_volume_section(ws, row, section_num, slabs):
    """Write the volume incentive subsection with slab table + merged qualification columns."""
    num_slabs = len(slabs)

    # Subsection header
    row = write_subsection_header(ws, row, f"{section_num}.1 Volume Incentive")

    # Column headers
    vol_headers = [
        "Sr.\nNo.",
        "TOD Slab (MT)",
        "Incentive Rate\n(\u20b9/mt)",
        "Payout\nFrequency",
        "Qualification\nCriteria",
        "Half-Yearly\nPayout Rule",
        "Yearly\nPayout Rule",
        "Monthly Minimum\nRequirement",
        "Maximum\nPayout Cap",
    ]
    header_row = row
    for col_idx, h in enumerate(vol_headers, 1):
        cell = ws.cell(row=row, column=col_idx, value=h)
        cell.font = FONT_COL_HEADER
        cell.fill = FILL_GRAY
        cell.alignment = ALIGN_CENTER
        cell.border = BORDER_THIN
    ws.row_dimensions[row].height = 35
    row += 1

    # Slab data rows
    first_data_row = row
    for i, (slab_text, rate) in enumerate(slabs, 1):
        write_data_cell(ws, row, 1, i, center=True)
        write_data_cell(ws, row, 2, slab_text, center=True)
        write_data_cell(ws, row, 3, rate, center=True)
        # Cols D-I: borders on every row (merge will handle content)
        for c in range(4, NUM_COLS + 1):
            cell = ws.cell(row=row, column=c)
            cell.border = BORDER_THIN
        row += 1
    last_data_row = row - 1

    # Vertically merge columns D-I across all slab rows
    merged_data = [
        (4, VOL_PAYOUT_FREQ),
        (5, VOL_QUAL),
        (6, VOL_HY_RULE),
        (7, VOL_YEARLY_RULE),
        (8, VOL_MONTHLY),
        (9, VOL_CAP),
    ]
    for col, text in merged_data:
        if num_slabs > 1:
            ws.merge_cells(
                start_row=first_data_row, start_column=col,
                end_row=last_data_row, end_column=col
            )
        cell = ws.cell(row=first_data_row, column=col, value=text)
        cell.font = FONT_DATA
        cell.alignment = ALIGN_LEFT_WRAP
        cell.border = BORDER_THIN

    return row + 1  # blank row after


# ─── Consistency Incentive Section Builder ───────────────────────────────────

def write_consistency_section(ws, row, section_num):
    """Write the consistency incentive subsection."""
    row = write_subsection_header(ws, row, f"{section_num}.2 Consistency Incentive")

    headers = [
        "Incentive Rate\n(\u20b9/mt)",
        "Threshold",
        "Payout\nFrequency",
        "Qualification\nCriteria",
        "Monthly\nRequirement",
        "Lean Period\nRules",
        "Enhancement\nRules",
        "Maximum\nPayout Cap",
        "",
    ]
    row = write_col_headers(ws, row, headers)

    # Data row
    data = [
        (1, 100, True),
        (2, "\u226522.5% quarterly\npro-rata achievement", False),
        (3, "Yearly", True),
        (4, CONSIST_QUAL, False),
        (5, CONSIST_MONTHLY, False),
        (6, CONSIST_LEAN, False),
        (7, CONSIST_ENHANCEMENT, False),
        (8, VOL_CAP, False),
    ]
    for col, val, center in data:
        write_data_cell(ws, row, col, val, center=center)
    # Col I empty
    ws.cell(row=row, column=9).border = BORDER_THIN
    ws.row_dimensions[row].height = 80
    row += 1

    return row + 1  # blank row


# ─── Loyalty Incentive Section Builder ───────────────────────────────────────

def write_loyalty_section(ws, row, section_num):
    """Write the loyalty incentive subsection."""
    row = write_subsection_header(ws, row, f"{section_num}.3 Loyalty Incentive")

    headers = [
        "Incentive Rate\n(\u20b9/mt)",
        "Threshold",
        "Payout\nFrequency",
        "Qualification\nCriteria",
        "Yearly Payout\nRule",
        "Maximum\nPayout Cap",
        "", "", "",
    ]
    row = write_col_headers(ws, row, headers)

    data = [
        (1, 50, True),
        (2, "\u2265110% of FY 24-25\nSigned/Amended Volume", False),
        (3, "Annually", True),
        (4, LOYALTY_QUAL, False),
        (5, LOYALTY_YEARLY, False),
        (6, VOL_CAP, False),
    ]
    for col, val, center in data:
        write_data_cell(ws, row, col, val, center=center)
    for c in range(7, NUM_COLS + 1):
        ws.cell(row=row, column=c).border = BORDER_THIN
    ws.row_dimensions[row].height = 60
    row += 1

    return row + 1


# ─── MSME Section Builder ───────────────────────────────────────────────────

def write_msme_section_applicable(ws, row, section_num):
    """Write MSME incentive subsection (applicable - for OEM)."""
    row = write_subsection_header(ws, row, f"{section_num}.4 MSME Incentive")

    headers = [
        "Incentive Rate\n(\u20b9/mt)",
        "Payout\nFrequency",
        "Qualification\nfor Signing",
        "Payout\nCriteria",
        "Upper Cap",
        "Special\nConditions",
        "", "", "",
    ]
    row = write_col_headers(ws, row, headers)

    data = [
        (1, 300, True),
        (2, "Annually", True),
        (3, MSME_QUAL, False),
        (4, MSME_PAYOUT, False),
        (5, MSME_CAP, False),
        (6, "MSME Incentive of \u20b9300/mt applicable over and above existing policy. UDYAM certificate required.", False),
    ]
    for col, val, center in data:
        write_data_cell(ws, row, col, val, center=center)
    for c in range(7, NUM_COLS + 1):
        ws.cell(row=row, column=c).border = BORDER_THIN
    ws.row_dimensions[row].height = 80
    row += 1

    return row + 1


def write_msme_section_na(ws, row, section_num):
    """Write MSME incentive subsection (not applicable - for Retail)."""
    row = write_subsection_header(ws, row, f"{section_num}.4 MSME Incentive")
    row = write_na_row(ws, row, "Not Applicable \u2014 MSME Incentive is for OEM/MSME customers only (Non-Retail).")
    return row + 1


def write_super_dealer_na(ws, row, section_num):
    """Write Super Dealer incentive subsection (N/A)."""
    row = write_subsection_header(ws, row, f"{section_num}.5 Super Dealer Incentive")
    row = write_na_row(ws, row, "Not Applicable \u2014 Only for Retail Coated Product Family (C1, C2, C3.1, C3.2, C3.3) with yearly volume \u226560,000 MT combined.")
    return row + 1


def write_pci_na(ws, row, section_num):
    """Write Process Compliance incentive subsection (N/A)."""
    row = write_subsection_header(ws, row, f"{section_num}.6 Process Compliance Incentive (PCI)")
    row = write_na_row(ws, row, "Not Applicable \u2014 Only for Product Family C3.1 and C3.2 for Retail Customers.")
    return row + 1


# ─── Sheet 1: TMT Incentives - Consolidated ─────────────────────────────────

def create_sheet1(wb):
    ws = wb.active
    ws.title = "TMT Incentives - Consolidated"
    ws.sheet_properties.pageSetUpPr = openpyxl.worksheet.properties.PageSetupProperties(fitToPage=True)
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0

    row = 1

    # ── Title ──
    merge_write(ws, row, 1, NUM_COLS,
                "JSW Steel \u2014 TMT Incentive Summary FY 2025-26",
                FONT_TITLE, FILL_DARK_BLUE, ALIGN_LEFT_CENTER, BORDER_THIN, 40)
    row += 1

    # ── Subtitle ──
    merge_write(ws, row, 1, NUM_COLS,
                "Product Family: G  |  Product: TMT Bars in straight Length, Coils, Cut & Bend  |  Date: 5th May 2025",
                FONT_SUBTITLE, FILL_LIGHT_BLUE, ALIGN_LEFT_CENTER, BORDER_THIN, 25)
    row += 1
    row += 1  # blank

    # ══════════ SECTION 1: TMT RETAIL ══════════
    row = write_section_header(ws, row, "1. TMT RETAIL (Distribution Channel: Retail)")
    row += 1  # blank

    row = write_volume_section(ws, row, "1", TMT_RETAIL_SLABS)
    row = write_consistency_section(ws, row, "1")
    row = write_loyalty_section(ws, row, "1")
    row = write_msme_section_na(ws, row, "1")
    row = write_super_dealer_na(ws, row, "1")
    row = write_pci_na(ws, row, "1")

    row += 1  # extra blank separator

    # ══════════ SECTION 2: TMT OEMs ══════════
    row = write_section_header(ws, row, "2. TMT OEMs (Distribution Channel: OEM)")
    row += 1  # blank

    row = write_volume_section(ws, row, "2", TMT_OEM_SLABS)
    row = write_consistency_section(ws, row, "2")
    row = write_loyalty_section(ws, row, "2")
    row = write_msme_section_applicable(ws, row, "2")
    row = write_super_dealer_na(ws, row, "2")
    row = write_pci_na(ws, row, "2")

    # ── Footer ──
    row += 1
    merge_write(ws, row, 1, NUM_COLS, SOURCE_NOTE, FONT_FOOTER)

    # ── Column widths ──
    col_widths = [6, 22, 14, 18, 38, 32, 32, 30, 22]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    print(f"  Sheet 1 created (last row: {row})")


# ─── Sheet 2: Terms & Conditions ────────────────────────────────────────────

def create_sheet2(wb):
    ws = wb.create_sheet("Terms & Conditions")
    ws.sheet_properties.pageSetUpPr = openpyxl.worksheet.properties.PageSetupProperties(fitToPage=True)
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0

    row = 1

    # Title
    merge_write(ws, row, 1, NUM_COLS,
                "General Terms & Conditions \u2014 Annual Incentive Scheme FY 2025-26",
                FONT_TC_TITLE, FILL_DARK_BLUE, ALIGN_LEFT_CENTER, BORDER_THIN, 35)
    row += 1
    row += 1  # blank

    # Clauses
    for i, clause_text in enumerate(TERMS_AND_CONDITIONS, 1):
        text = f"{i}.  {clause_text}"
        merge_write(ws, row, 1, NUM_COLS, text, FONT_TC_CLAUSE, alignment=ALIGN_LEFT_WRAP, border=BORDER_THIN)
        # Auto row height based on text length
        est_lines = max(1, len(clause_text) // 120 + 1)
        ws.row_dimensions[row].height = max(20, est_lines * 16)
        row += 1

    # Footer
    row += 1
    merge_write(ws, row, 1, NUM_COLS, SOURCE_NOTE, FONT_FOOTER)

    # Column widths — make all equal for even merged distribution
    for i in range(1, NUM_COLS + 1):
        ws.column_dimensions[get_column_letter(i)].width = 16

    print(f"  Sheet 2 created ({len(TERMS_AND_CONDITIONS)} clauses, last row: {row})")


# ─── Main ────────────────────────────────────────────────────────────────────

def main():
    print("Generating TMT Incentives Excel (Executive Document Style)...")
    print(f"Output: {OUTPUT_FILE}")
    print()

    wb = openpyxl.Workbook()

    create_sheet1(wb)
    create_sheet2(wb)

    wb.save(OUTPUT_FILE)
    print()
    print(f"Successfully saved to: {OUTPUT_FILE}")


if __name__ == "__main__":
    main()
