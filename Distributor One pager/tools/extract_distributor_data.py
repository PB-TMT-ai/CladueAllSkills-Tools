"""
Extract structured data for a single distributor one-pager report.
Given a (Distributor Name, State) combo, pulls data from all relevant sheets
and returns a structured dictionary ready for docx generation.
"""

import openpyxl
from datetime import datetime
from collections import defaultdict
import sys
import json

EXCEL_PATH = r"D:\Distributor One pager\data\JSW One TMT_distributor reports.xlsx"

# FY26 months: Apr'25 through Mar'26
FY26_MONTHS = ["Apr'25", "May'25", "Jun'25", "Jul'25", "Aug'25", "Sep'25",
               "Oct'25", "Nov'25", "Dec'25", "Jan'26", "Feb'26", "Mar'26"]

# Map onboarded month strings to FY26 month index (0-based)
ONBOARD_MONTH_MAP = {
    "Apr'25": 0, "May'25": 1, "Jun'25": 2, "Jul'25": 3,
    "Aug'25": 4, "Sep'25": 5, "Oct'25": 6, "Nov'25": 7,
    "Dec'25": 8, "Jan'26": 9, "Feb'26": 10, "Mar'26": 11
}


def load_workbook():
    return openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)


def normalize(s):
    """Normalize string for comparison."""
    if s is None:
        return ""
    return str(s).strip().upper()


def safe_float(v):
    """Convert to float, return 0 if not possible."""
    if v is None:
        return 0.0
    try:
        f = float(v)
        return f if f == f else 0.0  # NaN check
    except (ValueError, TypeError):
        return 0.0


def safe_int(v):
    if v is None:
        return 0
    try:
        return int(float(v))
    except (ValueError, TypeError):
        return 0


def extract_section1_basic_details(dist_name, state, district_rows):
    """Section 1: Distributor basic details from District master."""
    dist_norm = normalize(dist_name)
    state_norm = normalize(state)

    matching = [r for r in district_rows
                if normalize(r['distributor']) == dist_norm
                and normalize(r['state']) == state_norm]

    districts_served = len(matching)
    retail_potential = sum(safe_float(r['retail_demand']) for r in matching)
    vh_h_m = sum(1 for r in matching if normalize(r['categorization']) in
                 ('VERY HIGH', 'HIGH', 'MEDIUM'))

    return {
        'distributor_name': dist_name,
        'state': state,
        'districts_served': districts_served,
        'retail_potential_month': retail_potential,
        'vh_h_m_districts': vh_h_m
    }


def extract_section2_manpower(dist_name, state, district_rows, smtm_rows=None):
    """Section 2: Sales manpower. SM/TM from smtm_rows if available, else district master."""
    dist_norm = normalize(dist_name)
    state_norm = normalize(state)
    matching = [r for r in district_rows
                if normalize(r['distributor']) == dist_norm
                and normalize(r['state']) == state_norm]

    # SM/TM from dedicated SM_TM file (preferred)
    sm = None
    tm = None
    if smtm_rows:
        sm_names = set()
        tm_names = set()
        for r in smtm_rows:
            if normalize(r['distributor']) == dist_norm and normalize(r['state']) == state_norm:
                if r['sm'] and str(r['sm']).strip() not in ('', 'None', 'nan', '-'):
                    sm_names.add(str(r['sm']).strip())
                if r['tm'] and str(r['tm']).strip() not in ('', 'None', 'nan', '-'):
                    tm_names.add(str(r['tm']).strip())
        if sm_names:
            sm = ' / '.join(sorted(sm_names))
        if tm_names:
            tm = ' / '.join(sorted(tm_names))

    # Fallback to district master
    dgo_total = 0
    dsr_total = 0
    for r in matching:
        if not sm and r['sm'] and str(r['sm']).strip() not in ('', 'None', 'nan', '-'):
            sm = str(r['sm']).strip()
        if not tm and r['tm'] and str(r['tm']).strip() not in ('', 'None', 'nan', '-'):
            tm = str(r['tm']).strip()
        dgo_total += safe_int(r['dgo'])
        dsr_total += safe_int(r['dsr'])

    return {
        'sm': sm or '-',
        'tm': tm or '-',
        'dgo': dgo_total,
        'dsr': dsr_total
    }


def extract_section3_sales_performance(dist_name, achi_rows, mou_rows):
    """Section 3: Sales performance from achi and MoU sheets."""
    dist_norm = normalize(dist_name)

    # From Distributor-wise achi
    achi_row = None
    for r in achi_rows:
        if normalize(r['name']) == dist_norm:
            achi_row = r
            break

    # From Original MoU targets
    mou_row = None
    for r in mou_rows:
        if normalize(r['name']) == dist_norm:
            mou_row = r
            break

    revised_mou = safe_float(achi_row['revised_mou']) if achi_row else 0
    achieved_total = safe_float(achi_row['total']) if achi_row else 0
    pct_achievement = (achieved_total / revised_mou * 100) if revised_mou else 0

    monthly_targets = []
    monthly_achieved = []
    monthly_pct = []

    for i in range(12):
        target = safe_float(mou_row['months'][i]) if mou_row else 0
        achieved = safe_float(achi_row['months'][i]) if achi_row else 0
        pct = (achieved / target * 100) if target else 0
        monthly_targets.append(target)
        monthly_achieved.append(achieved)
        monthly_pct.append(pct)

    target_total = safe_float(mou_row['total']) if mou_row else 0
    achieved_td = sum(monthly_achieved)  # Total to date

    return {
        'revised_mou': revised_mou,
        'achieved_total': achieved_total,
        'pct_achievement': pct_achievement,
        'monthly_targets': monthly_targets,
        'monthly_achieved': monthly_achieved,
        'monthly_pct': monthly_pct,
        'target_total': target_total,
        'achieved_td': achieved_td
    }


def extract_section4_sales_by_location(dist_name, invoice_rows):
    """Section 4: Sales by location from Invoice data."""
    dist_norm = normalize(dist_name)

    # Categories: (business_unit, address_tag) -> monthly quantities
    categories = {
        'retailer_shop': [0.0] * 12,
        'retailer_site': [0.0] * 12,
        'selfstock_warehouse': [0.0] * 12,
        'selfstock_site': [0.0] * 12,
    }

    for r in invoice_rows:
        if normalize(r['distributor']) != dist_norm:
            continue

        bu = normalize(r['business_unit'])
        at = normalize(r['address_tag'])
        qty = safe_float(r['invoiced_qty'])
        month_idx = r.get('fy26_month_idx')

        if month_idx is None or month_idx < 0 or month_idx >= 12:
            continue

        if bu == 'RETAILER' and at == 'SHOP':
            categories['retailer_shop'][month_idx] += qty
        elif bu == 'RETAILER' and at == 'SITE':
            categories['retailer_site'][month_idx] += qty
        elif bu == 'SELF-STOCKING' and at == 'WAREHOUSE':
            categories['selfstock_warehouse'][month_idx] += qty
        elif bu == 'SELF-STOCKING' and at == 'SITE':
            categories['selfstock_site'][month_idx] += qty

    # Compute parent totals and child percentages
    categories['retailer_total'] = [
        categories['retailer_shop'][m] + categories['retailer_site'][m] for m in range(12)
    ]
    categories['selfstock_total'] = [
        categories['selfstock_warehouse'][m] + categories['selfstock_site'][m] for m in range(12)
    ]
    categories['retailer_shop_pct'] = [
        (categories['retailer_shop'][m] / categories['retailer_total'][m] * 100)
        if categories['retailer_total'][m] else 0
        for m in range(12)
    ]
    categories['retailer_site_pct'] = [
        (categories['retailer_site'][m] / categories['retailer_total'][m] * 100)
        if categories['retailer_total'][m] else 0
        for m in range(12)
    ]
    categories['selfstock_warehouse_pct'] = [
        (categories['selfstock_warehouse'][m] / categories['selfstock_total'][m] * 100)
        if categories['selfstock_total'][m] else 0
        for m in range(12)
    ]
    categories['selfstock_site_pct'] = [
        (categories['selfstock_site'][m] / categories['selfstock_total'][m] * 100)
        if categories['selfstock_total'][m] else 0
        for m in range(12)
    ]

    return categories


def extract_section5_channel_performance(dist_name, dealer_rows):
    """Section 5: Channel performance from FY 26 dealer sales."""
    dist_norm = normalize(dist_name)
    matching = [r for r in dealer_rows if normalize(r['distributor']) == dist_norm]

    # Secondary sales per month
    sec_sales = [0.0] * 12
    # Transacting dealers per month
    transacting = [0] * 12
    # New dealers per month
    new_dealers = [0] * 12
    # Active dealers per month (transacted in last 3 months)
    active = [0] * 12

    for dealer in matching:
        monthly = dealer['fy26_monthly']  # list of 12 monthly sales
        onboard_month = dealer.get('onboarded_month', '')

        # Count new dealer
        if onboard_month in ONBOARD_MONTH_MAP:
            idx = ONBOARD_MONTH_MAP[onboard_month]
            new_dealers[idx] += 1

        for m in range(12):
            sale = safe_float(monthly[m])
            sec_sales[m] += sale
            if sale > 0:
                transacting[m] += 1

    # Active dealers: transacted at least once in last 3 months
    for m in range(12):
        active_set = set()
        for dealer in matching:
            monthly = dealer['fy26_monthly']
            for lookback in range(max(0, m - 2), m + 1):
                if safe_float(monthly[lookback]) > 0:
                    active_set.add(dealer['name'])
                    break
            active[m] = len(active_set)

    trans_active_ratio = []
    for m in range(12):
        ratio = (transacting[m] / active[m]) if active[m] else 0
        trans_active_ratio.append(ratio)

    sec_sales_total = sum(sec_sales)
    new_dealers_total = sum(new_dealers)

    # FY26 summary values:
    # Transacting FY26 = unique dealers who lifted in any month of FY26
    transacting_fy26 = 0
    for dealer in matching:
        if any(safe_float(dealer['fy26_monthly'][m]) > 0 for m in range(12)):
            transacting_fy26 += 1

    # Active FY26 = latest month's active value (Feb'26 = index 10)
    latest_month = 10  # Feb'26
    active_fy26 = active[latest_month]

    return {
        'sec_sales': sec_sales,
        'sec_sales_total': sec_sales_total,
        'new_dealers': new_dealers,
        'new_dealers_total': new_dealers_total,
        'transacting': transacting,
        'transacting_fy26': transacting_fy26,
        'active': active,
        'active_fy26': active_fy26,
        'trans_active_ratio': trans_active_ratio,
    }


def extract_section6_key_districts(dist_name, state, district_rows, dealer_rows):
    """Section 6: Key district performance (VH/H/M only)."""
    dist_norm = normalize(dist_name)
    state_norm = normalize(state)
    matching = [r for r in district_rows
                if normalize(r['distributor']) == dist_norm
                and normalize(r['state']) == state_norm
                and normalize(r['categorization']) in ('VERY HIGH', 'HIGH', 'MEDIUM')]

    # Get dealer data grouped by district
    dist_dealers = [d for d in dealer_rows if normalize(d['distributor']) == dist_norm]
    dealers_by_district = defaultdict(list)
    for d in dist_dealers:
        dealers_by_district[normalize(d['district'])].append(d)

    results = []
    for r in matching:
        district_name = r['district']
        district_norm = normalize(district_name)
        district_dealers = dealers_by_district.get(district_norm, [])

        potential = safe_float(r['retail_demand'])

        # Sales volume: sum FY26 sales from district master
        sales_vol = sum(safe_float(r['fy26_sales'][i]) for i in range(12))

        # # billed dealers TD: dealers with any FY26 sales
        billed_td = 0
        for d in district_dealers:
            if any(safe_float(d['fy26_monthly'][m]) > 0 for m in range(12)):
                billed_td += 1

        # Current month transacting (use latest month with data)
        current_transacting = 0
        current_active = set()
        latest_month = 10  # Feb'26 (index 10, 0-based) - last month with likely data

        for d in district_dealers:
            if safe_float(d['fy26_monthly'][latest_month]) > 0:
                current_transacting += 1
            # Active: transacted in last 3 months
            for lb in range(max(0, latest_month - 2), latest_month + 1):
                if safe_float(d['fy26_monthly'][lb]) > 0:
                    current_active.add(d['name'])
                    break

        active_count = len(current_active)

        # SoB%: FY26 Avg Monthly Sales / Counter Potential of transacting dealers
        # Only count CP of dealers who lifted (had sales > 0) in each month
        total_fy26_sales = 0
        total_transacting_cp = 0
        months_with_transactions = 0

        for m in range(12):
            month_sales = 0
            month_cp = 0
            for d in district_dealers:
                sale = safe_float(d['fy26_monthly'][m])
                if sale > 0:
                    month_sales += sale
                    month_cp += safe_float(d.get('counter_potential', 0))
            total_fy26_sales += month_sales
            total_transacting_cp += month_cp
            if month_sales > 0:
                months_with_transactions += 1

        # Average monthly sales and average monthly transacting CP
        avg_monthly_sales = (total_fy26_sales / months_with_transactions) if months_with_transactions else 0
        avg_monthly_cp = (total_transacting_cp / months_with_transactions) if months_with_transactions else 0

        sob_pct = (avg_monthly_sales / avg_monthly_cp) if avg_monthly_cp else 0

        # Reach%: counter potential of transacting dealers / retail demand per month
        # Use avg monthly transacting CP (same as SoB denominator)
        reach_pct = (avg_monthly_cp / potential) if potential else 0

        # Market share%: Reach% * SoB% (both as fractions, display as %)
        market_share = reach_pct * sob_pct

        results.append({
            'district': district_name,
            'category': r['categorization'],
            'potential': potential,
            'sales_vol': sales_vol,  # FY26 total sales from district master
            'billed_td': billed_td,
            'transacting': current_transacting,
            'active': active_count,
            'sob_pct': sob_pct,
            'reach_pct': reach_pct,
            'market_share_pct': market_share
        })

    # Sort by potential descending
    results.sort(key=lambda x: x['potential'], reverse=True)
    return results


def extract_section7_top_dealers(dist_name, dealer_rows):
    """Section 7: Top 10 dealers by FY26 total sales."""
    dist_norm = normalize(dist_name)
    matching = [d for d in dealer_rows if normalize(d['distributor']) == dist_norm]

    # Calculate FY26 total for each dealer
    for d in matching:
        d['fy26_total'] = sum(safe_float(d['fy26_monthly'][m]) for m in range(12))

    # Sort by total, take top 10
    matching.sort(key=lambda x: x['fy26_total'], reverse=True)
    top10 = matching[:10]

    results = []
    for d in top10:
        results.append({
            'name': d['name'],
            'district': d['district'],
            'segmentation': d.get('segmentation', '-'),
            'monthly': d['fy26_monthly'],
            'total': d['fy26_total']
        })

    return results


# ─── Sheet Loaders ───────────────────────────────────────────

def load_district_master(wb):
    """Load District master sheet into list of dicts."""
    ws = wb['District master']
    rows = []
    for i, row in enumerate(ws.iter_rows(min_row=4, values_only=True)):
        vals = list(row)
        if not vals[1]:  # skip empty
            continue

        # FY26 sales: cols 26-37 (indices 25-36, Apr'25-Mar'26)
        fy26_sales = [safe_float(vals[j]) if j < len(vals) else 0.0 for j in range(25, 37)]

        rows.append({
            'district': str(vals[1]).strip() if vals[1] else '',
            'state': str(vals[2]).strip() if vals[2] else '',
            'zone': str(vals[3]).strip() if vals[3] else '',
            'distributor': str(vals[4]).strip() if vals[4] else '',
            'categorization': str(vals[5]).strip() if vals[5] else '',
            'dsr': vals[6],
            'dgo': vals[7],
            'sm': vals[8],
            'tm': vals[9],
            'retail_demand': vals[10],
            'market_share': vals[11],
            'fy26_sales': fy26_sales,
        })
    return rows


def load_achi_data(wb):
    """Load FY 26_Distributor-wise achi. sheet."""
    ws = wb['FY 26_Distributor-wise achi.']
    rows = []
    for row in ws.iter_rows(min_row=4, values_only=True):
        vals = list(row)
        if not vals[1]:
            continue
        months = [safe_float(vals[j]) if j < len(vals) else 0.0 for j in range(4, 16)]
        rows.append({
            'name': str(vals[1]).strip(),
            'state': str(vals[2]).strip() if vals[2] else '',
            'revised_mou': vals[3],
            'months': months,
            'total': vals[16] if len(vals) > 16 else 0,
        })
    return rows


def load_mou_targets(wb):
    """Load FY 26_Original MoU_targets sheet."""
    ws = wb['FY 26_Original MoU_targets']
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        vals = list(row)
        if not vals[1]:
            continue
        months = [safe_float(vals[j]) if j < len(vals) else 0.0 for j in range(7, 19)]
        rows.append({
            'name': str(vals[1]).strip(),
            'code': str(vals[2]).strip() if vals[2] else '',
            'state': str(vals[3]).strip() if vals[3] else '',
            'months': months,
            'total': vals[19] if len(vals) > 19 else 0,
        })
    return rows


def load_dealer_sales(wb):
    """Load FY 26 dealer sales sheet."""
    ws = wb['FY 26 dealer sales']
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        vals = list(row)
        if not vals[4]:  # dealer name
            continue

        # FY26 monthly: cols 38-49 (indices 37-48)
        fy26_monthly = []
        for j in range(37, 49):
            fy26_monthly.append(safe_float(vals[j]) if j < len(vals) else 0.0)

        rows.append({
            'name': str(vals[4]).strip(),
            'state': str(vals[5]).strip() if vals[5] else '',
            'district': str(vals[6]).strip() if vals[6] else '',
            'distributor': str(vals[8]).strip() if vals[8] else '',
            'onboarded_month': str(vals[16]).strip() if vals[16] else '',
            'counter_potential': vals[19],
            'segmentation': str(vals[21]).strip() if len(vals) > 21 and vals[21] else '-',
            'fy26_monthly': fy26_monthly,
        })
    return rows


def load_invoice_data(wb):
    """Load Invoice sheet (filtered to FY26 only for performance)."""
    ws = wb['Invoice']
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        vals = list(row)
        if not vals[1]:
            continue

        # Parse invoice date to get FY26 month index
        inv_date = vals[1]
        month_idx = None
        try:
            if isinstance(inv_date, datetime):
                dt = inv_date
            elif isinstance(inv_date, str):
                dt = datetime.strptime(inv_date.strip(), '%Y-%m-%d')
            else:
                continue

            # FY26: Apr 2025 (idx 0) to Mar 2026 (idx 11)
            if dt.year == 2025 and dt.month >= 4:
                month_idx = dt.month - 4
            elif dt.year == 2026 and dt.month <= 3:
                month_idx = dt.month + 8
        except (ValueError, TypeError):
            continue

        if month_idx is None:
            continue  # Skip non-FY26 invoices

        rows.append({
            'distributor': str(vals[8]).strip() if vals[8] else '',
            'invoiced_qty': vals[23],
            'business_unit': str(vals[32]).strip() if vals[32] else '',
            'address_tag': str(vals[47]).strip() if len(vals) > 47 and vals[47] else '',
            'fy26_month_idx': month_idx,
        })
    return rows


SMTM_PATH = r"D:\Distributor One pager\data\District wise SM_TM name.xlsx"


def load_smtm_data():
    """Load SM/TM names from dedicated District wise SM_TM name.xlsx file."""
    wb = openpyxl.load_workbook(SMTM_PATH, read_only=True, data_only=True)
    ws = wb['Sheet1']
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        vals = list(row)
        if not vals[4]:
            continue
        rows.append({
            'district': str(vals[1]).strip() if vals[1] else '',
            'state': str(vals[2]).strip() if vals[2] else '',
            'zone': str(vals[3]).strip() if vals[3] else '',
            'distributor': str(vals[4]).strip() if vals[4] else '',
            'sm': vals[5],
            'tm': vals[6],
        })
    wb.close()
    return rows


def extract_all_from_loaded(distributor_name, state, district_rows, achi_rows, mou_rows, dealer_rows, invoice_rows, smtm_rows=None):
    """Extract all sections using pre-loaded sheet data (for batch mode)."""
    print(f"Extracting data for: {distributor_name} - {state}")

    data = {
        'section1': extract_section1_basic_details(distributor_name, state, district_rows),
        'section2': extract_section2_manpower(distributor_name, state, district_rows, smtm_rows),
        'section3': extract_section3_sales_performance(distributor_name, achi_rows, mou_rows),
        'section4': extract_section4_sales_by_location(distributor_name, invoice_rows),
        'section5': extract_section5_channel_performance(distributor_name, dealer_rows),
        'section6': extract_section6_key_districts(distributor_name, state, district_rows, dealer_rows),
        'section7': extract_section7_top_dealers(distributor_name, dealer_rows),
        'months': FY26_MONTHS,
    }

    return data


def extract_all(distributor_name, state):
    """Main extraction function - returns complete data dict for one-pager."""
    print(f"Loading workbook...")
    wb = load_workbook()

    print(f"Loading sheets...")
    district_rows = load_district_master(wb)
    achi_rows = load_achi_data(wb)
    mou_rows = load_mou_targets(wb)
    dealer_rows = load_dealer_sales(wb)

    print(f"Loading invoice data (this may take a moment)...")
    invoice_rows = load_invoice_data(wb)

    wb.close()

    smtm_rows = load_smtm_data()

    return extract_all_from_loaded(distributor_name, state, district_rows, achi_rows, mou_rows, dealer_rows, invoice_rows, smtm_rows)


if __name__ == '__main__':
    # Test with sample distributor
    dist = sys.argv[1] if len(sys.argv) > 1 else "NIKUNJ UDYOG"
    state = sys.argv[2] if len(sys.argv) > 2 else "HARYANA"

    data = extract_all(dist, state)

    # Print summary
    s1 = data['section1']
    print(f"\n{'='*60}")
    print(f"Distributor: {s1['distributor_name']}")
    print(f"State: {s1['state']}")
    print(f"Districts served: {s1['districts_served']}")
    print(f"Retail potential: {s1['retail_potential_month']}")
    print(f"VH/H/M districts: {s1['vh_h_m_districts']}")

    s2 = data['section2']
    print(f"\nSM: {s2['sm']}, TM: {s2['tm']}")
    print(f"DGO: {s2['dgo']}, DSR: {s2['dsr']}")

    s3 = data['section3']
    print(f"\nRevised MoU: {s3['revised_mou']}")
    print(f"Achieved: {s3['achieved_total']:.1f}")
    print(f"% Achievement: {s3['pct_achievement']:.1f}%")
    print(f"Monthly targets: {s3['monthly_targets']}")
    print(f"Monthly achieved: {[round(x,1) for x in s3['monthly_achieved']]}")

    s5 = data['section5']
    print(f"\nSec sales total: {s5['sec_sales_total']:.1f}")
    print(f"New dealers: {s5['new_dealers']}")
    print(f"Transacting: {s5['transacting']}")
    print(f"Active: {s5['active']}")

    s6 = data['section6']
    print(f"\nKey districts ({len(s6)}):")
    for d in s6:
        print(f"  {d['district']} ({d['category']}): pot={d['potential']}, "
              f"SoB={d['sob_pct']*100:.1f}%, Reach={d['reach_pct']*100:.1f}%, MS={d['market_share_pct']*100:.2f}%")

    s7 = data['section7']
    print(f"\nTop {len(s7)} dealers:")
    for d in s7:
        print(f"  {d['name']} ({d['district']}): {d['total']:.1f} MT, seg={d['segmentation']}")
