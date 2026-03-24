"""
Tool: inspect_excel_files.py
Purpose: Read all Excel files from North-Marketing-Execution-Reports and print
         sheet names, column headers, row counts, and a sample of rows for each.
Usage: python tools/inspect_excel_files.py <directory_path>
"""

import sys
import os
import openpyxl
import json

def inspect_file(filepath):
    """Return metadata about each sheet in an Excel file."""
    result = {"file": os.path.basename(filepath), "sheets": []}
    try:
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                result["sheets"].append({"name": sheet_name, "rows": 0, "headers": [], "sample": []})
                continue

            # Find the header row (first non-empty row)
            header_row = None
            header_idx = 0
            for i, row in enumerate(rows):
                if any(cell is not None for cell in row):
                    header_row = [str(cell) if cell is not None else "" for cell in row]
                    header_idx = i
                    break

            # Get sample data rows (up to 3)
            data_rows = rows[header_idx + 1: header_idx + 4]
            sample = []
            for row in data_rows:
                sample.append([str(cell) if cell is not None else "" for cell in row])

            result["sheets"].append({
                "name": sheet_name,
                "total_rows": len(rows) - header_idx - 1,
                "headers": header_row,
                "sample": sample
            })
        wb.close()
    except Exception as e:
        result["error"] = str(e)
    return result

def main():
    directory = sys.argv[1] if len(sys.argv) > 1 else "/home/user/North-Marketing-Execution-Reports"

    xlsx_files = sorted([f for f in os.listdir(directory) if f.endswith(('.xlsx', '.xls'))])

    print(f"Found {len(xlsx_files)} Excel files\n")
    print("=" * 100)

    for fname in xlsx_files:
        filepath = os.path.join(directory, fname)
        info = inspect_file(filepath)

        print(f"\n📄 FILE: {info['file']}")
        if "error" in info:
            print(f"   ❌ Error: {info['error']}")
            continue

        for sheet in info["sheets"]:
            print(f"   Sheet: '{sheet['name']}' | Rows: {sheet.get('total_rows', 0)}")
            if sheet.get("headers"):
                # Clean headers for display
                headers = [h for h in sheet["headers"] if h.strip()]
                print(f"   Headers: {headers}")
            if sheet.get("sample"):
                print(f"   Sample row: {sheet['sample'][0][:8]}...")
        print("-" * 100)

if __name__ == "__main__":
    main()
