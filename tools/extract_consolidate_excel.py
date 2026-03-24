"""
Tool: extract_consolidate_excel.py
Purpose: Extract and consolidate all Excel data from North-Marketing-Execution-Reports
         into categorized CSV summaries with UNIFIED column schemas.
Input: Directory path containing Excel files
Output: Consolidated CSVs in .tmp/ directory + summary stats to stdout
"""

import sys
import os
import pandas as pd
import openpyxl
from datetime import datetime

REPORT_DIR = sys.argv[1] if len(sys.argv) > 1 else "/home/user/North-Marketing-Execution-Reports"
OUTPUT_DIR = "/home/user/CladueAllSkills-Tools/.tmp"

os.makedirs(OUTPUT_DIR, exist_ok=True)

# ── Unified schemas ──────────────────────────────────────────────────────────

WP_COLUMNS = [
    'serial_no', 'region', 'dealer_name', 'state', 'city', 'location',
    'phone', 'district', 'status', 'latitude', 'longitude',
    'wall_number', 'painting_type', 'wall_tenure_months',
    'rented_or_free', 'rent_amount', 'execution_date',
    'remarks', 'plan_sqft', 'actual_sqft', 'size',
    'source_file', 'source_sheet'
]

IMPACT_COLUMNS = [
    'serial_no', 'region', 'painting_type', 'state', 'city', 'location',
    'site_name', 'status', 'latitude', 'longitude',
    'wall_number', 'wall_tenure_months', 'rented_or_free',
    'execution_date', 'remarks', 'plan_sqft', 'actual_sqft',
    'width', 'height',
    'source_file', 'source_sheet'
]

GSB_COLUMNS = [
    'serial_no', 'dealer_name', 'state', 'city_district', 'phone',
    'address', 'gst_no', 'board_type', 'width', 'height',
    'qty', 'sqft', 'rate', 'amount',
    'source_file'
]

INSHOP_COLUMNS = [
    'serial_no', 'dealer_name', 'area', 'district', 'phone',
    'oneway_desc', 'oneway_sqft', 'vinyl_desc', 'vinyl_sqft',
    'sunboard_desc', 'sunboard_sqft', 'gsb',
    'source_file'
]

BILL_COLUMNS = [
    'serial_no', 'dealer_name', 'wall_address', 'painting_sqft',
    'gsb_size', 'nlb_size', 'remarks',
    'source_file'
]


def read_sheet_raw(filepath, sheet_name=None):
    """Read an Excel sheet and return rows as list of lists, with detected header index."""
    try:
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        ws = wb[sheet_name] if sheet_name else wb[wb.sheetnames[0]]
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
        if not rows:
            return [], 0
        # Find header row
        for i, row in enumerate(rows):
            row_str = " ".join([str(c).lower() for c in row if c is not None])
            if any(k in row_str for k in ['s.no', 's.n.', 'sr.', 'region', 'dealer', 'city', 'district']):
                return rows, i
        return rows, 0
    except Exception as e:
        print(f"  ⚠ Error reading {os.path.basename(filepath)} [{sheet_name}]: {e}")
        return [], 0


def cell(row, idx):
    """Safely get cell value from row by index."""
    if idx is not None and idx < len(row):
        v = row[idx]
        return str(v).strip() if v is not None else ''
    return ''


def find_col(headers, *keywords):
    """Find column index matching any keyword (case-insensitive)."""
    for i, h in enumerate(headers):
        hl = str(h).lower().strip() if h else ''
        for kw in keywords:
            if kw in hl:
                return i
    return None


# ── DEALER WALL PAINTING ─────────────────────────────────────────────────────

def extract_wp_structured(filepath, sheet_name, source_label):
    """Extract dealer WP data into unified schema."""
    rows, hi = read_sheet_raw(filepath, sheet_name)
    if not rows:
        return []

    headers = rows[hi]
    data_rows = [r for r in rows[hi+1:] if any(c is not None and str(c).strip() for c in r)]

    # Map columns
    c_sno = find_col(headers, 's.no', 's.n.', 'sr.')
    c_region = find_col(headers, 'region')
    c_dealer = find_col(headers, 'dealer', 'area')
    c_state = find_col(headers, 'state')
    c_city = find_col(headers, 'city')
    c_location = find_col(headers, 'location')
    c_phone = find_col(headers, 'phone', 'contact')
    c_district = find_col(headers, 'district')
    c_status = find_col(headers, 'status')
    c_lat = find_col(headers, 'latitude', 'lat')
    c_lon = find_col(headers, 'longitude', 'lon')
    c_wallno = find_col(headers, 'wall/shop', 'wall number', 'wall no')
    c_type = find_col(headers, 'painting type', 'type')
    c_tenure = find_col(headers, 'tenure')
    c_rented = find_col(headers, 'rented', 'free wall')
    c_rent_amt = find_col(headers, 'amount of rent', 'rent')
    c_exec_date = find_col(headers, 'execution date', 'execution')
    c_remarks = find_col(headers, 'remark', 'comment')
    c_plan = find_col(headers, 'plan sq', 'planned sq')
    c_actual = find_col(headers, 'actual sq')
    c_size = find_col(headers, 'size')

    result = []
    for row in data_rows:
        sno = cell(row, c_sno)
        # Skip sub-header rows that got mixed in
        if sno.lower() in ('s.no', 'sr.', 's.n.', ''):
            if not cell(row, c_dealer):
                continue

        result.append({
            'serial_no': sno,
            'region': cell(row, c_region),
            'dealer_name': cell(row, c_dealer),
            'state': cell(row, c_state),
            'city': cell(row, c_city),
            'location': cell(row, c_location),
            'phone': cell(row, c_phone),
            'district': cell(row, c_district),
            'status': cell(row, c_status),
            'latitude': cell(row, c_lat),
            'longitude': cell(row, c_lon),
            'wall_number': cell(row, c_wallno),
            'painting_type': cell(row, c_type),
            'wall_tenure_months': cell(row, c_tenure),
            'rented_or_free': cell(row, c_rented),
            'rent_amount': cell(row, c_rent_amt),
            'execution_date': cell(row, c_exec_date),
            'remarks': cell(row, c_remarks),
            'plan_sqft': cell(row, c_plan),
            'actual_sqft': cell(row, c_actual),
            'size': cell(row, c_size),
            'source_file': os.path.basename(filepath),
            'source_sheet': source_label,
        })
    return result


def consolidate_dealer_wp():
    """Consolidate all dealer wall/shop painting files."""
    files_config = [
        ('Bill no 145 Dealer WP (1).xlsx', 'Sheet1', 'Bill 145'),
        ('JSW Complete Work 21.6.2025 (1).xlsx', 'Sheet1', 'Complete Work Jun25'),
        ('JSW_Dealer WP_Final (1) (1).xlsx', 'HR', 'Dealer WP - HR'),
        ('JSW_Dealer WP_Final (1) (1).xlsx', 'DL', 'Dealer WP - Delhi'),
        ('JSW One Tmt Dealer Wall & Shop Painting 1 March 2026 (2) (1).xlsx', 'Sheet1', 'Dealer WP Mar26'),
        ('PB & Jammu Wall painting (1) (2).xlsx', 'WP', 'PB & Jammu WP'),
        ('UP WP (3).xlsx', 'WP', 'UP WP v3'),
        ('UP WP (4).xlsx', 'WP', 'UP WP v4'),
        ('WP - 100625 (4).xlsx', 'HR', 'WP Jun25 - HR'),
        ('WP - 100625 (4).xlsx', 'NCR', 'WP Jun25 - NCR'),
        ('WP - 100625 (4).xlsx', 'RAj', 'WP Jun25 - Rajasthan'),
        ('WP-200625 (5).xlsx', 'Haryana', 'WP Jun25v2 - HR'),
        ('WP-200625 (5).xlsx', 'Rajasthan', 'WP Jun25v2 - Rajasthan'),
        ('WP-200625 (5).xlsx', 'NCR - UP', 'WP Jun25v2 - NCR-UP'),
        ('WP-200625 (5).xlsx', 'NCR-Delhi', 'WP Jun25v2 - Delhi'),
        ('Bill No 321 (1).xlsx', 'HR', 'Bill 321 - HR'),
    ]

    all_rows = []
    for fname, sheet, label in files_config:
        fpath = os.path.join(REPORT_DIR, fname)
        if not os.path.exists(fpath):
            print(f"  ⚠ Missing: {fname}")
            continue
        extracted = extract_wp_structured(fpath, sheet, label)
        print(f"  ✓ {label}: {len(extracted)} rows")
        all_rows.extend(extracted)

    return pd.DataFrame(all_rows, columns=WP_COLUMNS)


# ── IMPACT WALL ───────────────────────────────────────────────────────────────

def extract_impact_structured(filepath, sheet_name, source_label):
    """Extract impact wall data into unified schema."""
    rows, hi = read_sheet_raw(filepath, sheet_name)
    if not rows:
        return []

    headers = rows[hi]
    data_rows = [r for r in rows[hi+1:] if any(c is not None and str(c).strip() for c in r)]

    c_sno = find_col(headers, 's.no', 's.n.')
    c_region = find_col(headers, 'region')
    c_type = find_col(headers, 'painting type', 'type')
    c_state = find_col(headers, 'state')
    c_city = find_col(headers, 'city')
    c_location = find_col(headers, 'location')
    c_site = find_col(headers, 'site name', 'site')
    c_status = find_col(headers, 'status')
    c_lat = find_col(headers, 'latitude')
    c_lon = find_col(headers, 'longitude')
    c_wallno = find_col(headers, 'wall/shop', 'wall number')
    c_tenure = find_col(headers, 'tenure')
    c_rented = find_col(headers, 'rented')
    c_exec = find_col(headers, 'execution')
    c_remarks = find_col(headers, 'remark', 'comment')
    c_plan = find_col(headers, 'plan')
    c_actual = find_col(headers, 'actual')
    c_w = find_col(headers, 'w')
    c_h = find_col(headers, 'h')

    result = []
    for row in data_rows:
        sno = cell(row, c_sno)
        if sno.lower() in ('s.no', 'sr.', ''):
            if not cell(row, c_city):
                continue

        result.append({
            'serial_no': sno,
            'region': cell(row, c_region),
            'painting_type': cell(row, c_type),
            'state': cell(row, c_state),
            'city': cell(row, c_city),
            'location': cell(row, c_location),
            'site_name': cell(row, c_site),
            'status': cell(row, c_status),
            'latitude': cell(row, c_lat),
            'longitude': cell(row, c_lon),
            'wall_number': cell(row, c_wallno),
            'wall_tenure_months': cell(row, c_tenure),
            'rented_or_free': cell(row, c_rented),
            'execution_date': cell(row, c_exec),
            'remarks': cell(row, c_remarks),
            'plan_sqft': cell(row, c_plan),
            'actual_sqft': cell(row, c_actual),
            'width': cell(row, c_w),
            'height': cell(row, c_h),
            'source_file': os.path.basename(filepath),
            'source_sheet': source_label,
        })
    return result


def consolidate_impact_wall():
    """Consolidate all impact wall files."""
    files_config = [
        ('Impact WP - 100625 (4).xlsx', 'HR', 'Impact Jun25 - HR'),
        ('Impact WP - 100625 (4).xlsx', 'Ghaziabad', 'Impact Jun25 - Ghaziabad'),
        ('JSW IMPACT - 080725.xlsx', 'Haryana', 'Impact Jul25 - HR'),
        ('JSW Impact Wall Xls Jaipur Area (1).xlsx', 'Sheet1', 'Impact - Jaipur'),
        ('UP Impact Wall (3).xlsx', 'Sheet1', 'Impact - UP'),
        ('JSW One Tmt Impact Wall 01 March 2026 (1) (1).xlsx', 'Sheet1', 'Impact Mar26'),
        ('JSW One Tmt Impact Wall 31 December 2025 (3).xlsx', 'Sheet1', 'Impact Dec25'),
    ]

    all_rows = []
    for fname, sheet, label in files_config:
        fpath = os.path.join(REPORT_DIR, fname)
        if not os.path.exists(fpath):
            print(f"  ⚠ Missing: {fname}")
            continue
        extracted = extract_impact_structured(fpath, sheet, label)
        print(f"  ✓ {label}: {len(extracted)} rows")
        all_rows.extend(extracted)

    return pd.DataFrame(all_rows, columns=IMPACT_COLUMNS)


# ── GSB / NLB BOARDS ─────────────────────────────────────────────────────────

def extract_gsb_structured(filepath, source_label):
    """Extract GSB/NLB board data into unified schema."""
    rows, hi = read_sheet_raw(filepath)
    if not rows:
        return []

    headers = rows[hi]
    data_rows = [r for r in rows[hi+1:] if any(c is not None and str(c).strip() for c in r)]

    c_sno = find_col(headers, 's.no', 's.n.', 'sr.')
    c_dealer = find_col(headers, 'dealer', 'delear', 'name of the dealer', 'subdealer name')
    c_state = find_col(headers, 'state')
    c_city = find_col(headers, 'city', 'district', 'area')
    c_phone = find_col(headers, 'phone', 'mobile')
    c_address = find_col(headers, 'address', 'full address')
    c_gst = find_col(headers, 'gst')
    c_type = find_col(headers, 'type of board', 'type')
    c_w = find_col(headers, 'w')
    c_h = find_col(headers, 'h')
    c_qty = find_col(headers, 'qty')
    c_sqft = find_col(headers, 'sqft', 'sq.ft', 'sq. ft')
    c_rate = find_col(headers, 'rate')
    c_amount = find_col(headers, 'amount', 'cost', 'total')

    result = []
    for row in data_rows:
        sno = cell(row, c_sno)
        if sno.lower() in ('s.no', 'sr.', ''):
            if not cell(row, c_dealer):
                continue

        result.append({
            'serial_no': sno,
            'dealer_name': cell(row, c_dealer),
            'state': cell(row, c_state),
            'city_district': cell(row, c_city),
            'phone': cell(row, c_phone),
            'address': cell(row, c_address),
            'gst_no': cell(row, c_gst),
            'board_type': cell(row, c_type),
            'width': cell(row, c_w),
            'height': cell(row, c_h),
            'qty': cell(row, c_qty),
            'sqft': cell(row, c_sqft),
            'rate': cell(row, c_rate),
            'amount': cell(row, c_amount),
            'source_file': os.path.basename(filepath),
        })
    return result


def consolidate_gsb_nlb():
    """Consolidate all GSB/NLB board files."""
    files = [
        ('D2R NLB LIST UPDATED TILL JUNE 25 FINAL (1) (1) (1).xlsx', 'D2R NLB Jun25'),
        ('Delhi HR Boards_5th Feb_Invoice Final (1).xlsx', 'Delhi HR Boards Feb'),
        ('GSB -  June 2025_300625 (2).xlsx', 'GSB Jun25'),
        ('GSB+ NLB (1).xlsx', 'GSB+NLB'),
        ('JSW GSB NLB Board FINAL 22 jan 2026 (3).xlsx', 'GSB NLB Jan26'),
        ('JSW One Tmt Dealer GSB ,NLB Installation 8 July 2025.xlsx', 'GSB NLB Jul25'),
    ]

    all_rows = []
    for fname, label in files:
        fpath = os.path.join(REPORT_DIR, fname)
        if not os.path.exists(fpath):
            continue
        extracted = extract_gsb_structured(fpath, label)
        print(f"  ✓ {label}: {len(extracted)} rows")
        all_rows.extend(extracted)

    return pd.DataFrame(all_rows, columns=GSB_COLUMNS)


# ── IN-SHOP BRANDING ─────────────────────────────────────────────────────────

def consolidate_inshop():
    """Extract in-shop branding data."""
    fpath = os.path.join(REPORT_DIR, 'inshop (PB, HR, RJ, Delhi ) (1).xlsx')
    if not os.path.exists(fpath):
        return pd.DataFrame(columns=INSHOP_COLUMNS)

    rows, hi = read_sheet_raw(fpath)
    headers = rows[hi]
    data_rows = [r for r in rows[hi+1:] if any(c is not None and str(c).strip() for c in r)]

    result = []
    for row in data_rows:
        result.append({
            'serial_no': cell(row, 0),
            'dealer_name': cell(row, 1),
            'area': cell(row, 2),
            'district': cell(row, 3),
            'phone': cell(row, 4),
            'oneway_desc': cell(row, 5),
            'oneway_sqft': cell(row, 6),
            'vinyl_desc': cell(row, 7),
            'vinyl_sqft': cell(row, 8),
            'sunboard_desc': cell(row, 9),
            'sunboard_sqft': cell(row, 10),
            'gsb': cell(row, 11),
            'source_file': os.path.basename(fpath),
        })
    print(f"  ✓ In-shop Branding: {len(result)} rows")
    return pd.DataFrame(result, columns=INSHOP_COLUMNS)


# ── BILLS ─────────────────────────────────────────────────────────────────────

def consolidate_bills():
    """Extract bill/invoice summary data."""
    files = [
        'Bill no 119 (1).xlsx',
        'Bill no 121 (1).xlsx',
    ]

    all_rows = []
    for fname in files:
        fpath = os.path.join(REPORT_DIR, fname)
        if not os.path.exists(fpath):
            continue

        rows, hi = read_sheet_raw(fpath)
        headers = rows[hi]
        data_rows = [r for r in rows[hi+1:] if any(c is not None and str(c).strip() for c in r)]

        c_sno = find_col(headers, 'sr.', 's.no')
        c_dealer = find_col(headers, 'dealer')
        c_address = find_col(headers, 'wall address', 'address')
        c_painting = find_col(headers, 'painting')
        c_gsb = find_col(headers, 'gsb')
        c_nlb = find_col(headers, 'nlb')

        for row in data_rows:
            sno = cell(row, c_sno)
            if sno.lower() in ('sr.no.', 's.no', ''):
                if not cell(row, c_dealer):
                    continue
            all_rows.append({
                'serial_no': sno,
                'dealer_name': cell(row, c_dealer),
                'wall_address': cell(row, c_address),
                'painting_sqft': cell(row, c_painting),
                'gsb_size': cell(row, c_gsb),
                'nlb_size': cell(row, c_nlb),
                'remarks': '',
                'source_file': fname,
            })
        print(f"  ✓ {fname}: {len(data_rows)} rows")

    return pd.DataFrame(all_rows, columns=BILL_COLUMNS)


# ── SUMMARY ───────────────────────────────────────────────────────────────────

def print_summary(name, df, state_col=None):
    """Print summary for a category."""
    print(f"\n{'='*80}")
    print(f"  {name}")
    print(f"{'='*80}")
    print(f"  Rows: {len(df)} | Columns: {len(df.columns)}")
    print(f"  Schema: {list(df.columns)}")

    if state_col and state_col in df.columns:
        states = df[state_col].replace('', pd.NA).dropna()
        if len(states) > 0:
            print(f"\n  By {state_col}:")
            for s, ct in states.value_counts().head(10).items():
                print(f"    {s}: {ct}")

    if 'source_file' in df.columns:
        print(f"\n  Sources ({df['source_file'].nunique()}):")
        for f, ct in df.groupby('source_file').size().items():
            print(f"    {f}: {ct}")


# ── MAIN ──────────────────────────────────────────────────────────────────────

def main():
    print("=" * 80)
    print("  NORTH MARKETING EXECUTION REPORTS - UNIFIED DATA CONSOLIDATION")
    print(f"  Source: {REPORT_DIR}")
    print(f"  Output: {OUTPUT_DIR}")
    print(f"  Run at: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("=" * 80)

    # 1. Dealer Wall Painting
    print("\n[1/5] Dealer Wall/Shop Painting...")
    wp = consolidate_dealer_wp()
    wp.to_csv(os.path.join(OUTPUT_DIR, "unified_dealer_wall_painting.csv"), index=False)
    print_summary("DEALER WALL/SHOP PAINTING", wp, 'state')

    # 2. Impact Wall
    print("\n[2/5] Impact Wall Painting...")
    impact = consolidate_impact_wall()
    impact.to_csv(os.path.join(OUTPUT_DIR, "unified_impact_wall.csv"), index=False)
    print_summary("IMPACT WALL PAINTING", impact, 'state')

    # 3. GSB/NLB
    print("\n[3/5] GSB/NLB Boards...")
    gsb = consolidate_gsb_nlb()
    gsb.to_csv(os.path.join(OUTPUT_DIR, "unified_gsb_nlb_boards.csv"), index=False)
    print_summary("GSB/NLB BOARDS", gsb, 'state')

    # 4. In-shop
    print("\n[4/5] In-shop Branding...")
    inshop = consolidate_inshop()
    inshop.to_csv(os.path.join(OUTPUT_DIR, "unified_inshop_branding.csv"), index=False)
    print_summary("IN-SHOP BRANDING", inshop, 'area')

    # 5. Bills
    print("\n[5/5] Bills/Invoices...")
    bills = consolidate_bills()
    bills.to_csv(os.path.join(OUTPUT_DIR, "unified_bills.csv"), index=False)
    print_summary("BILLS/INVOICES", bills)

    # Final summary
    print("\n" + "=" * 80)
    print("  CONSOLIDATION COMPLETE")
    print("=" * 80)
    totals = {
        "Dealer Wall/Shop Painting": len(wp),
        "Impact Wall Painting": len(impact),
        "GSB/NLB Boards": len(gsb),
        "In-shop Branding": len(inshop),
        "Bills/Invoices": len(bills),
    }
    for name, count in totals.items():
        print(f"  {name}: {count} rows")
    print(f"  TOTAL: {sum(totals.values())} rows")
    print(f"\n  Output files:")
    for f in sorted(os.listdir(OUTPUT_DIR)):
        if f.startswith('unified_') and f.endswith('.csv'):
            size = os.path.getsize(os.path.join(OUTPUT_DIR, f))
            print(f"    {f} ({size:,} bytes)")


if __name__ == "__main__":
    main()
